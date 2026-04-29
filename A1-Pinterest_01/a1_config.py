"""
Load secrets, settings, and prompts with repo-wide + per-pipeline + per-site layers:

- Keys:  config/shared_keys.json → (optional) <project>/config/keys.json → (optional) same keys on
  the active site row in config/sites.json → environment variables
- Settings:  config/shared_settings.json → (optional) <project>/config/settings.json →
  (optional) site["settings"] in config/sites.json
- Prompts:  config/prompts/{name}.json → (optional) <project>/config/prompts/{name}.json →
  (optional) config/site_prompts/{site["prompts_dir"]}/{name}.json →
  (optional) site["prompts"][name] in config/sites.json

On the active site row, set "no_shared_settings": true to skip config/shared_settings.json, or
"no_shared_prompts": true to skip config/prompts/ (shared repo files only; project A1/.../config/prompts
still load).

Templates: copy config/shared_keys.example.json -> shared_keys.json, shared_settings.example.json ->
shared_settings.json; see config/sites.example.json. Optional: A1-Pinterest_01/config/keys.json or
settings.json; per-site file prompts: prompts_dir + config/site_prompts/{id}/*.json.

If set, environment variables override JSON for the same setting (e.g. OPENAI_API_KEY, OPENAI_MODEL,
USEAPI_NET_API_TOKEN or USEAPI_TOKEN, USEAPI_MJ_CHANNEL or USEAPI_MIDJOURNEY_CHANNEL, CLOUDFLARE_ACCOUNT_ID,
R2_ACCESS_KEY_ID, R2_SECRET_ACCESS_KEY, R2_BUCKET, R2_PUBLIC_BASE_URL, WP_URL, WP_USER, WP_APP_PASSWORD;
also PINTEREST_SITE_ID, PINTEREST_OUT_DIR, OPENAI_REQUEST_TIMEOUT, OPENAI_RETRY_MAX_TRIES in child scripts).

All merges are deep (nested dicts combine). The active site is selected by PINTEREST_SITE_ID.
"""
from __future__ import annotations

import json
import os
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, Optional

A1_DIR = Path(__file__).resolve().parent
CONFIG_DIR = A1_DIR / "config"
PROMPTS_DIR = CONFIG_DIR / "prompts"
REPO_ROOT = A1_DIR.parent
# Repo config: OpenAI, UseAPI, R2. Optional <project>/config/keys.json merges on top; missing files = skip layer.
REPO_CONFIG_DIR = REPO_ROOT / "config"
SHARED_KEYS_PATH = REPO_CONFIG_DIR / "shared_keys.json"
SHARED_SETTINGS_PATH = REPO_CONFIG_DIR / "shared_settings.json"
SHARED_PROMPTS_DIR = REPO_CONFIG_DIR / "prompts"
SITE_PROMPTS_DIR = REPO_CONFIG_DIR / "site_prompts"
SITES_PATH = REPO_CONFIG_DIR / "sites.json"


def _deep_merge(base: Dict[str, Any], extra: Optional[Dict[str, Any]]) -> Dict[str, Any]:
    if not extra:
        return dict(base) if base else {}
    out: Dict[str, Any] = dict(base) if base else {}
    for k, v in extra.items():
        if k in out and isinstance(out[k], dict) and isinstance(v, dict):
            out[k] = _deep_merge(out[k], v)
        else:
            out[k] = v
    return out


def _read_json(path: Path) -> Dict[str, Any]:
    if not path.is_file():
        return {}
    with open(path, "r", encoding="utf-8") as f:
        data = json.load(f)
    return data if isinstance(data, dict) else {}


def load_settings() -> Dict[str, Any]:
    """
    Merges config/shared_settings.json, project config/settings.json, then
    the active site row's "settings" object in config/sites.json (if present).
    If the site has no_shared_settings: true, the shared file is not used (start from project local).
    """
    site = get_active_site()
    use_shared = not (
        isinstance(site, dict) and (site.get("no_shared_settings") is True)
    )
    shared = _read_json(SHARED_SETTINGS_PATH) if use_shared else {}
    local = _read_json(CONFIG_DIR / "settings.json")
    data = _deep_merge(shared, local)
    st = (site or {}).get("settings") if isinstance(site, dict) else None
    if isinstance(st, dict) and st:
        data = _deep_merge(data, st)
    return data


def load_keys() -> Dict[str, Any]:
    """
    Merges config/shared_keys.json with (optional) A1-Pinterest_01/config/keys.json if that file exists.
    Project values override shared. Start from config/shared_keys.example.json -> copy to shared_keys.json.

    For a run with PINTEREST_SITE_ID set, any non-empty OpenAI / UseAPI / R2 / WordPress fields on that
    row in config/sites.json override shared and project keys for that site only.
    """
    shared = _read_json(SHARED_KEYS_PATH)
    keys_path = CONFIG_DIR / "keys.json"
    local = _read_json(keys_path) if keys_path.is_file() else {}
    data: Dict[str, Any] = {**shared, **local}

    # Blanks in project keys.json must not erase shared_keys.json (only non-empty local overrides)
    for k in (
        "openai_api_key",
        "openai_model",
        "useapi_token",
        "useapi_midjourney_channel",
        "r2_account_id",
        "r2_access_key_id",
        "r2_secret_access_key",
        "r2_bucket",
        "r2_public_base_url",
    ):
        v = data.get(k)
        if isinstance(v, str) and not v.strip() and shared.get(k) and str(shared.get(k, "")).strip():
            data[k] = shared[k]

    # Per-site row in config/sites.json overrides shared + project (when PINTEREST_SITE_ID is set)
    site = get_active_site()
    if isinstance(site, dict):
        for k in (
            "openai_api_key",
            "openai_model",
            "useapi_token",
            "useapi_midjourney_channel",
            "r2_account_id",
            "r2_access_key_id",
            "r2_secret_access_key",
            "r2_bucket",
            "r2_public_base_url",
            "wordpress_url",
            "wordpress_user",
            "wordpress_app_password",
        ):
            v = site.get(k)
            if v is not None and str(v).strip() != "":
                data[k] = v

    # --- env overrides (standard names) ---
    if os.environ.get("OPENAI_API_KEY"):
        data["openai_api_key"] = os.environ["OPENAI_API_KEY"]
    if os.environ.get("OPENAI_MODEL"):
        data["openai_model"] = os.environ["OPENAI_MODEL"]

    if os.environ.get("USEAPI_NET_API_TOKEN") or os.environ.get("USEAPI_TOKEN"):
        data["useapi_token"] = os.environ.get("USEAPI_NET_API_TOKEN") or os.environ.get("USEAPI_TOKEN", "")
    if os.environ.get("USEAPI_MJ_CHANNEL") or os.environ.get("USEAPI_MIDJOURNEY_CHANNEL"):
        v = os.environ.get("USEAPI_MJ_CHANNEL") or os.environ.get("USEAPI_MIDJOURNEY_CHANNEL", "")
        if v:
            data["useapi_midjourney_channel"] = v

    for a, b in [
        ("CLOUDFLARE_ACCOUNT_ID", "r2_account_id"),
        ("R2_ACCESS_KEY_ID", "r2_access_key_id"),
        ("R2_SECRET_ACCESS_KEY", "r2_secret_access_key"),
        ("R2_BUCKET", "r2_bucket"),
        ("R2_PUBLIC_BASE_URL", "r2_public_base_url"),
    ]:
        if os.environ.get(a):
            data[b] = os.environ[a]

    if os.environ.get("WP_URL"):
        data["wordpress_url"] = os.environ["WP_URL"]
    if os.environ.get("WP_USER"):
        data["wordpress_user"] = os.environ["WP_USER"]
    if os.environ.get("WP_APP_PASSWORD"):
        data["wordpress_app_password"] = os.environ["WP_APP_PASSWORD"]

    return data


def load_prompts(name: str) -> Dict[str, Any]:
    """
    Merges (deep): repo config/prompts/{name}.json, project config/prompts/{name}.json,
    then optional file config/site_prompts/{site.prompts_dir}/{name}.json,
    then optional site['prompts'][name] in config/sites.json.
    `name` is the filename without .json, e.g. a1_start, a2_json, a4_articles.
    """
    site = get_active_site()
    use_shared = not (
        isinstance(site, dict) and (site.get("no_shared_prompts") is True)
    )
    sh = _read_json(SHARED_PROMPTS_DIR / f"{name}.json") if use_shared else {}
    loc = _read_json(PROMPTS_DIR / f"{name}.json")
    data = _deep_merge(sh, loc)
    if not isinstance(site, dict):
        return data
    sub = site.get("prompts_dir")
    if isinstance(sub, str) and sub.strip():
        p = SITE_PROMPTS_DIR / sub.strip() / f"{name}.json"
        data = _deep_merge(data, _read_json(p))
    pm = site.get("prompts")
    if isinstance(pm, dict) and name in pm and isinstance(pm[name], dict):
        data = _deep_merge(data, pm[name])
    return data


def load_prompts_excluding_row_inline(name: str) -> Dict[str, Any]:
    """
    Same file layers as load_prompts (shared, project, site prompts_dir files) but does not
    apply site['prompts'][name] from the sites.json row — for comparing “baseline” vs inline row.
    """
    site = get_active_site()
    use_shared = not (
        isinstance(site, dict) and (site.get("no_shared_prompts") is True)
    )
    sh = _read_json(SHARED_PROMPTS_DIR / f"{name}.json") if use_shared else {}
    loc = _read_json(PROMPTS_DIR / f"{name}.json")
    data = _deep_merge(sh, loc)
    if not isinstance(site, dict):
        return data
    sub = site.get("prompts_dir")
    if isinstance(sub, str) and sub.strip():
        p = SITE_PROMPTS_DIR / sub.strip() / f"{name}.json"
        data = _deep_merge(data, _read_json(p))
    return data


def read_excel_with_retry(
    path: str, *, sheet_name=0, retries: int = 10, delay: float = 0.4
):
    """
    Read an .xlsx while another process (or Excel) may briefly lock the file.
    Close the workbook in Excel if errors persist.
    """
    import time
    import pandas as pd

    last: Optional[Exception] = None
    for i in range(retries):
        try:
            return pd.read_excel(path, sheet_name=sheet_name)
        except PermissionError as e:
            last = e
            time.sleep(delay * (1 + i))
    assert last is not None
    raise last


def to_excel_with_retry(
    df, path: str, *, index: bool = False, retries: int = 10, delay: float = 0.4
) -> None:
    import time

    last: Optional[Exception] = None
    for i in range(retries):
        try:
            df.to_excel(path, index=index)
            return
        except PermissionError as e:
            last = e
            time.sleep(delay * (1 + i))
    assert last is not None
    raise last


def get_openai_model(settings: Optional[Dict[str, Any]] = None, keys: Optional[Dict[str, Any]] = None) -> str:
    s = settings if settings is not None else load_settings()
    k = keys if keys is not None else load_keys()
    return (k.get("openai_model") or s.get("openai_model") or "gpt-4o-mini").strip()


def set_openai_key_from_keys(keys: Optional[Dict[str, Any]] = None) -> str:
    import openai

    k = keys if keys is not None else load_keys()
    key = (k.get("openai_api_key") or os.environ.get("OPENAI_API_KEY") or "").strip()
    if not key:
        raise RuntimeError(
            "Missing openai_api_key: set in config/shared_keys.json, this project's config/keys.json, "
            "a site field in config/sites.json, or environment OPENAI_API_KEY."
        )
    openai.api_key = key
    return key


def load_sites() -> Dict[str, Any]:
    return _read_json(SITES_PATH)


def active_site_id() -> str:
    return (os.environ.get("PINTEREST_SITE_ID") or "").strip()


def get_active_site() -> Dict[str, Any]:
    """
    Resolves the site row for the current run (PINTEREST_SITE_ID) or a single default.
    When no sites.json: default output dir matches legacy A1 layout.
    """
    raw = load_sites()
    sites = raw.get("sites")
    sid = active_site_id()
    default_out = "A1-Pinterest_01-out"
    if isinstance(sites, list) and len(sites) > 0:
        for s in sites:
            if not isinstance(s, dict):
                continue
            if sid and str(s.get("id", "")).strip() == sid:
                d = dict(s)
                d.setdefault("out_dir", (d.get("out_dir") or f"{d.get('id', 'out')}-out").strip())
                return d
        s0 = dict(sites[0])
        s0.setdefault("out_dir", (s0.get("out_dir") or f"{s0.get('id', 'out')}-out").strip())
        if not sid:
            return s0
    if sid:
        out = (os.environ.get("PINTEREST_OUT_DIR") or f"{sid}-out").strip()
        return {"id": sid, "out_dir": out}
    return {"id": "default", "out_dir": default_out}


def all_output_dir() -> str:
    s = get_active_site()
    sub = (s.get("out_dir") or f"{s.get('id', 'out')}-out").strip()
    return str((REPO_ROOT / "ALL" / sub).resolve())


def all_output_join(*parts: str) -> str:
    base = Path(all_output_dir())
    p = base
    for a in parts:
        p = p / a
    return str(p.resolve())


def resolve_start_titles_excel() -> str:
    """
    Titles for A.1-START: STARTS/{start_file} if set, else STARTS/{site_id}.xlsx, else STARTS/START1.xlsx
    """
    override = (os.environ.get("PINTEREST_START_FILE_OVERRIDE") or "").strip()
    if override:
        p = Path(override)
        if not p.is_absolute():
            p = (REPO_ROOT / override).resolve()
        if p.is_file():
            return str(p)
    site = get_active_site()
    d = REPO_ROOT / "STARTS"
    sid = str(site.get("id", "default"))
    if site.get("start_file"):
        c = d / str(site["start_file"])
        if c.is_file():
            return str(c)
    for name in (f"{sid}.xlsx", "START1.xlsx"):
        c = d / name
        if c.is_file():
            return str(c)
    return str(d / f"{sid}.xlsx")


def _source_start_should_sync_usage(source_path: str) -> bool:
    """Updates START*.xlsx in STARTS/, not allocator temp sheets."""
    try:
        p = Path(source_path).resolve()
        sd = (REPO_ROOT / "STARTS").resolve()
        p.relative_to(sd)
        if "_runtime_global_start" in p.parts:
            return False
        return p.suffix.lower() == ".xlsx"
    except (ValueError, OSError):
        return False


def apply_usage_to_start_workbook(
    source_path: str, row_used_success: Optional[Dict[int, bool]] = None
) -> None:
    """
    Writes per-row columns on SOURCE STARTS workbook (pandas row index i ⇒ Excel sheet row == i + 2).

    Columns: used (1/0), used_at ("YYYY-mm-dd HH:MM:SS"), used_project.

    Args:
      row_used_success: map Excel row index (≥2) → True if Recipe non‑empty after A.1-START run.
      If omitted or empty, returns without doing IO.
    """
    if not source_path or not row_used_success:
        return
    if not _source_start_should_sync_usage(source_path):
        return
    fp = Path(source_path).resolve()
    if not fp.is_file():
        return

    site = get_active_site()
    sid = str(site.get("id", "") or "").strip()
    label = (
        str(site.get("display_name", "") or "").strip()
        or sid
        or "default"
    )
    proj_label = label

    stamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    try:
        from openpyxl import load_workbook
    except ImportError:
        raise RuntimeError("openpyxl is required for STARTS usage tracking") from None

    wb = load_workbook(fp)

    try:
        sh = wb.active
        max_row = max(2, int(sh.max_row or 2))
        scan_hi = max(int(sh.max_column or 1), 16)
        header_to_col: Dict[str, int] = {}
        for c in range(1, scan_hi + 1):
            hv = sh.cell(row=1, column=c).value
            if hv is None:
                continue
            header_to_col[str(hv).strip().lower()] = c

        title_col = header_to_col.get("title")
        if title_col is None:
            for c in range(1, scan_hi + 1):
                hv = sh.cell(row=1, column=c).value
                if hv is not None and str(hv).strip().lower() == "title":
                    title_col = c
                    header_to_col["title"] = c
                    break
        if title_col is None:
            title_col = 1

        def _hdr(c: int) -> str:
            v = sh.cell(row=1, column=c).value
            if v is None:
                return ""
            return str(v).strip()

        def ensure_col(name: str) -> int:
            lk = name.strip().lower()
            c0 = header_to_col.get(lk)
            if c0:
                return c0
            new_c = int(sh.max_column or 1) + 1
            sh.cell(row=1, column=new_c, value=name)
            header_to_col[lk] = new_c
            return new_c

        def prefer_after_title(name: str, offset: int) -> int:
            """Always prefer Title+offset for usage metadata columns."""
            lk = name.strip().lower()
            c = title_col + offset
            h = _hdr(c).lower()
            if h == "" or h == lk:
                sh.cell(row=1, column=c, value=name)
                header_to_col[lk] = c
                return c
            c0 = header_to_col.get(lk)
            if c0:
                return c0
            return ensure_col(name)

        used_col = prefer_after_title("used", 1)
        used_at_col = prefer_after_title("used_at", 2)
        used_project_col = prefer_after_title("used_project", 3)

        for r in range(2, max_row + 1):
            ok = row_used_success.get(r)
            if ok is None:
                continue
            if ok:
                sh.cell(row=r, column=used_col, value=1)
                sh.cell(row=r, column=used_at_col, value=stamp)
                sh.cell(row=r, column=used_project_col, value=proj_label)
            else:
                sh.cell(row=r, column=used_col, value=0)
                sh.cell(row=r, column=used_at_col, value="")
                sh.cell(row=r, column=used_project_col, value="")

        wb.save(str(fp))
    finally:
        wb.close()


def _mask_secret(value: str, *, head: int = 4, tail: int = 4) -> str:
    s = (value or "").strip()
    if not s:
        return ""
    if len(s) <= head + tail + 1:
        return "(set)"
    return s[:head] + "…" + s[-tail:]


def _redact_keys_for_view(d: Dict[str, Any]) -> Dict[str, Any]:
    out = dict(d)
    sens = (
        "openai_api_key",
        "useapi_token",
        "r2_secret_access_key",
        "r2_access_key_id",
        "wordpress_app_password",
    )
    for k in sens:
        v = out.get(k)
        if v is not None and str(v).strip():
            out[k] = _mask_secret(str(v))
        elif v is not None and not str(v).strip():
            out[k] = ""
    return out


def _value_outline_for_view(
    val: Any,
    *,
    max_str: int = 120,
    _depth: int = 0,
    _max_depth: int = 10,
) -> Any:
    """
    Recursively show dict/list/string values in a size-bounded, dashboard-safe way
    (strings truncated; no more '<dict, N chars>' placeholders).
    """
    if _depth >= _max_depth:
        return "…"
    if val is None or isinstance(val, bool):
        return val
    if isinstance(val, (int, float)):
        return val
    if isinstance(val, str):
        s = val
        if len(s) > max_str:
            return s[: max_str - 1] + "…"
        return s
    if isinstance(val, list):
        if not val:
            return []
        return [
            _value_outline_for_view(x, max_str=max_str, _depth=_depth + 1, _max_depth=_max_depth)
            for x in val[:5]
        ] + (["…"] if len(val) > 5 else [])
    if isinstance(val, dict):
        return {
            k: _value_outline_for_view(v, max_str=max_str, _depth=_depth + 1, _max_depth=_max_depth)
            for k, v in val.items()
        }
    return str(val)


def _active_site_dict_for_view(site: Dict[str, Any]) -> Dict[str, Any]:
    """Site row for Info tab: inline prompts are outlined (readable), not collapsed to a char count."""
    out = dict(site)
    pm = out.get("prompts")
    if isinstance(pm, dict):
        out["prompts"] = {
            str(name): _value_outline_for_view(val)
            for name, val in pm.items()
        }
    return out


# Secrets that may appear on the site row in sites.json; mask like merged keys
_ACTIVE_SITE_REDACT_KEYS = (
    "wordpress_app_password",
    "openai_api_key",
    "useapi_token",
    "r2_secret_access_key",
    "r2_access_key_id",
)


def keys_provenance() -> Dict[str, str]:
    """
    Human-readable **where the merged value comes from** for each key in load_keys().

    Precedence matches load_keys: environment → this site row in sites.json →
    A1-.../config/keys.json → config/shared_keys.json. WordPress fields are not in
    shared_keys; they come from the site row or WP_* env only.
    """
    site = get_active_site() if isinstance(get_active_site(), dict) else {}
    m = load_keys()
    out: Dict[str, str] = {}
    if not m:
        return out

    def _nz(d: object, k: str) -> bool:
        if not isinstance(d, dict) or k not in d:
            return False
        return str(d.get(k) or "").strip() != ""

    def _v_eq(a: object, b: object) -> bool:
        return str(a or "") == str(b or "")

    API_KEYS = (
        "openai_api_key",
        "openai_model",
        "useapi_token",
        "useapi_midjourney_channel",
        "r2_account_id",
        "r2_access_key_id",
        "r2_secret_access_key",
        "r2_bucket",
        "r2_public_base_url",
    )
    local_path = CONFIG_DIR / "keys.json"
    local = _read_json(local_path) if local_path.is_file() else {}
    shared = _read_json(SHARED_KEYS_PATH)

    for k in m.keys():
        if k == "openai_api_key" and os.environ.get("OPENAI_API_KEY"):
            out[k] = "Environment: OPENAI_API_KEY (highest priority)"
        elif k == "openai_model" and os.environ.get("OPENAI_MODEL"):
            out[k] = "Environment: OPENAI_MODEL"
        elif k == "useapi_token" and (
            os.environ.get("USEAPI_TOKEN") or os.environ.get("USEAPI_NET_API_TOKEN")
        ):
            out[k] = "Environment: USEAPI_TOKEN or USEAPI_NET_API_TOKEN"
        elif k == "useapi_midjourney_channel" and (
            os.environ.get("USEAPI_MJ_CHANNEL")
            or os.environ.get("USEAPI_MIDJOURNEY_CHANNEL")
        ):
            out[k] = "Environment: USEAPI_MJ_CHANNEL / USEAPI_MIDJOURNEY_CHANNEL"
        elif k == "r2_account_id" and os.environ.get("CLOUDFLARE_ACCOUNT_ID"):
            out[k] = "Environment: CLOUDFLARE_ACCOUNT_ID"
        elif k == "r2_access_key_id" and os.environ.get("R2_ACCESS_KEY_ID"):
            out[k] = "Environment: R2_ACCESS_KEY_ID"
        elif k == "r2_secret_access_key" and os.environ.get("R2_SECRET_ACCESS_KEY"):
            out[k] = "Environment: R2_SECRET_ACCESS_KEY"
        elif k == "r2_bucket" and os.environ.get("R2_BUCKET"):
            out[k] = "Environment: R2_BUCKET"
        elif k == "r2_public_base_url" and os.environ.get("R2_PUBLIC_BASE_URL"):
            out[k] = "Environment: R2_PUBLIC_BASE_URL"
        elif k == "wordpress_url" and os.environ.get("WP_URL"):
            out[k] = "Environment: WP_URL"
        elif k == "wordpress_user" and os.environ.get("WP_USER"):
            out[k] = "Environment: WP_USER"
        elif k == "wordpress_app_password" and os.environ.get("WP_APP_PASSWORD"):
            out[k] = "Environment: WP_APP_PASSWORD"
    for k in m.keys():
        if k in out:
            continue
        if _nz(site, k):
            out[k] = (
                f'Overridden: this project row in config/sites.json (field "{k}")  '
                "wins over shared_keys.json and A1-…/config/keys.json"
            )
            continue
        if k in API_KEYS:
            if _nz(local, k) and _v_eq(m.get(k), local.get(k)):
                out[k] = (
                    f'From A1-Pinterest_01/config/keys.json (field "{k}")  '
                    "project override of shared_keys"
                )
            elif _nz(shared, k) and _v_eq(m.get(k), shared.get(k)):
                out[k] = (
                    f'From config/shared_keys.json (field "{k}")  '
                    "not set on this site row; value is the repo default for this key"
                )
            elif _nz(local, k):
                out[k] = f'From A1-Pinterest_01/config/keys.json (field "{k}")'
            elif _nz(shared, k):
                out[k] = f'From config/shared_keys.json (field "{k}")'
            else:
                out[k] = "merged (no JSON source found for this key)"
        elif k in ("wordpress_url", "wordpress_user", "wordpress_app_password"):
            out[k] = "Not in shared_keys. Add in sites.json or use WP_URL / WP_USER / WP_APP_PASSWORD"
        else:
            out[k] = "other"
    return out


_INLINE_PROMPT_NAMES = (
    "a1_start",
    "a2_json",
    "a2_prompt",
    "a4_articles",
    "a5_pin_data",
    "a8_pin_bulk",
)


def _prompt_leaf_field_kind(key: str, val: Any) -> str:
    """
    How to edit one leaf in the site-editor form, inferred from the repo prompt JSON value.
    """
    if isinstance(val, bool):
        return "bool"
    if isinstance(val, (int, float)) and not isinstance(val, bool):
        return "number"
    if isinstance(val, str):
        lk = (key or "").lower()
        if len(val) > 200 or any(
            x in lk
            for x in (
                "system",
                "user_template",
                "json_schema",
                "user",
                "template",
                "schema",
                "prefix",
                "suffix",
                "intro",
                "mid",
            )
        ):
            return "textarea"
        if len(val) > 80:
            return "textarea"
        return "text"
    if isinstance(val, list):
        if not val:
            return "json"
        if all(isinstance(x, str) for x in val):
            return "lines"
        return "json"
    return "json"


def _flatten_repo_prompt_to_fields(d: Any, prefix: str, out: list) -> None:
    if isinstance(d, dict):
        for k, v in d.items():
            sub = f"{prefix}.{k}" if prefix else str(k)
            if isinstance(v, dict):
                if v:
                    _flatten_repo_prompt_to_fields(v, sub, out)
            else:
                out.append(
                    {
                        "path": sub,
                        "kind": _prompt_leaf_field_kind(k, v),
                        "label": sub,
                    }
                )


def prompts_inline_field_schema() -> Dict[str, Any]:
    """
    For each config/prompts/{name}.json, list of {path, kind, label} for the dashboard
    to render one input per field (builds the same JSON shape on save).
    """
    result: Dict[str, Any] = {}
    for n in _INLINE_PROMPT_NAMES:
        raw = _read_json(SHARED_PROMPTS_DIR / f"{n}.json")
        if not raw:
            result[n] = []
            continue
        rows: list = []
        _flatten_repo_prompt_to_fields(raw, "", rows)
        rows.sort(key=lambda r: (r.get("path") or ""))
        result[n] = rows
    return result


def _flatten_merged_prompt_to_map(d: Any, prefix: str, out: Dict[str, Any]) -> None:
    """Same leaves as the form schema, but store merged values (from load_prompts)."""
    if isinstance(d, dict):
        for k, v in d.items():
            sub = f"{prefix}.{k}" if prefix else str(k)
            if isinstance(v, dict):
                if v:
                    _flatten_merged_prompt_to_map(v, sub, out)
            else:
                out[sub] = v


def prompts_effective_by_path() -> Dict[str, Dict[str, Any]]:
    """
    For each inline prompt name, flat dot-path  merged value the pipeline would use
    (shared + project + site files + this row) — for dashboard placeholders when
    a field is not overridden on the site row.
    """
    result: Dict[str, Dict[str, Any]] = {}
    for n in _INLINE_PROMPT_NAMES:
        p = load_prompts(n)
        if not isinstance(p, dict) or not p:
            result[n] = {}
            continue
        flat: Dict[str, Any] = {}
        _flatten_merged_prompt_to_map(p, "", flat)
        result[n] = flat
    return result


def prompts_excluding_row_inline_by_path() -> Dict[str, Dict[str, Any]]:
    """
    Same paths as prompts_effective_by_path but merged from files + prompts_dir only,
    not sites.json row `prompts` — for tooltips when the row does override a field.
    """
    result: Dict[str, Dict[str, Any]] = {}
    for n in _INLINE_PROMPT_NAMES:
        p = load_prompts_excluding_row_inline(n)
        if not isinstance(p, dict) or not p:
            result[n] = {}
            continue
        flat: Dict[str, Any] = {}
        _flatten_merged_prompt_to_map(p, "", flat)
        result[n] = flat
    return result


def _prompt_block_hint(name: str) -> str:
    site = get_active_site() if isinstance(get_active_site(), dict) else {}
    use_shared = not (isinstance(site, dict) and (site.get("no_shared_prompts") is True))
    bits: list[str] = []
    if use_shared and _read_json(SHARED_PROMPTS_DIR / f"{name}.json"):
        bits.append("config/prompts/{}.json".format(name))
    if _read_json(PROMPTS_DIR / f"{name}.json"):
        bits.append("A1…/config/prompts/{}.json".format(name))
    if isinstance(site, dict) and str(site.get("prompts_dir") or "").strip():
        p = SITE_PROMPTS_DIR / str(site.get("prompts_dir")).strip() / f"{name}.json"
        if p.is_file() and _read_json(p):
            bits.append("config/site_prompts/…/{}.json".format(name))
    pm = site.get("prompts") if isinstance(site, dict) else None
    if (
        isinstance(pm, dict)
        and name in pm
        and isinstance(pm.get(name), dict)
        and pm.get(name)
    ):
        bits.append("this row → prompts → {}".format(name))
    if not bits:
        return "no files yet; add repo prompts or type inline below"
    return " ← ".join(bits)


def resolved_runtime_snapshot() -> Dict[str, Any]:
    """
    Merged settings/keys and active site (with secrets redacted) for the dashboard /api/site-config.
    Depends on the current PINTEREST_SITE_ID and the rest of the environment, like worker subprocesses.
    """
    site = get_active_site()
    site_view = _active_site_dict_for_view(site) if isinstance(site, dict) else site
    if isinstance(site_view, dict):
        for k in _ACTIVE_SITE_REDACT_KEYS:
            if k in site_view and site_view.get(k) is not None and str(site_view.get(k) or "").strip():
                site_view[k] = _mask_secret(str(site_view[k]))  # type: ignore[index]

    st = load_settings()
    k_all = load_keys()
    k_view = _redact_keys_for_view(k_all)
    pr_names = ("a1_start", "a2_json", "a2_prompt", "a4_articles", "a5_pin_data", "a8_pin_bulk")
    prompts_summary: Dict[str, Any] = {}
    for n in pr_names:
        p = load_prompts(n)
        if isinstance(p, dict):
            prompts_summary[n] = {
                "top_level_keys": list(p.keys()),
                "merged_outline": _value_outline_for_view(p),
            }
        else:
            prompts_summary[n] = {"top_level_keys": [], "merged_outline": {}}
    use_shared_st = not (isinstance(site, dict) and (site.get("no_shared_settings") is True))
    use_shared_pr = not (isinstance(site, dict) and (site.get("no_shared_prompts") is True))
    return {
        "pinterest_site_id": os.environ.get("PINTEREST_SITE_ID", ""),
        "all_output_dir": all_output_dir(),
        "resolve_start_titles_excel": resolve_start_titles_excel(),
        "how_merge_works": {
            "settings": "config/shared_settings.json, then A1-.../config/settings.json, then this site row settings (or skip shared if no_shared_settings).",
            "keys": "config/shared_keys.json, then project keys.json, then non-empty fields on this site row, then env overrides.",
            "wordpress": "URL / user / app password: only from this site row in sites.json and WP_URL / WP_USER / WP_APP_PASSWORD (not from shared_keys.json; that file is OpenAI / UseAPI / R2).",
            "prompts_files": "config/prompts, project prompts, config/site_prompts/prompts_dir, then site row prompts[…] (or skip shared dir if no_shared_prompts).",
            "uses_shared_settings_file": use_shared_st,
            "uses_shared_prompts_dir": use_shared_pr,
        },
        "openai_model_effective": get_openai_model(),
        "settings_top_level_keys": sorted(st.keys(), key=str),
        "keys_merged_field_names": sorted([x for x in k_all.keys() if not str(x).startswith("_")], key=str),
        "keys_provenance": keys_provenance(),
        "prompts_form_hints": {n: _prompt_block_hint(n) for n in pr_names},
        "prompts_inline_field_schema": prompts_inline_field_schema(),
        "active_site": site_view,
        "settings": st,
        "keys": k_view,
        "prompts_effective_merged": prompts_summary,
        "prompts_summary": prompts_summary,
        "prompts_effective_by_path": prompts_effective_by_path(),
        "prompts_excluding_row_inline_by_path": prompts_excluding_row_inline_by_path(),
    }
