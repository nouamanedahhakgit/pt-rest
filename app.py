import os
import shutil
import subprocess
import threading
import queue
import json
import re
import copy
import fnmatch
import importlib.util
from flask import (
    Flask,
    Response,
    request,
    jsonify,
    stream_with_context,
    render_template,
    redirect,
    url_for,
    flash,
    send_from_directory,
    abort,
)
import openpyxl
import openai
import sys
import time
import random
import logging
from datetime import datetime
from typing import Callable, Any, Optional, Dict, List


app = Flask(__name__)
app.secret_key = "dev"

# ------------------------------------------------------------
# !!! WARNING: This is not recommended for production usage !!!
# ------------------------------------------------------------
# من الأفضل تستعمل ENV بدل ما تكتب المفتاح مباشرة
openai.api_key = os.getenv("OPENAI_API_KEY",
                           "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")  # <--- Your key here

import openai_chat_compat

openai_chat_compat.install()

# ===== Retry: إعداد اللوجينغ
logging.basicConfig(level=logging.INFO, format="%(asctime)s [%(levelname)s] %(message)s")

# ===== Retry: OpenAI exceptions (SDK 1.x; legacy openai.error was removed)
try:
    from openai import (
        APIConnectionError,
        APITimeoutError,
        RateLimitError,
        InternalServerError,
        APIError as OpenAIAPIError,
    )

    DEFAULT_RETRY_EXCEPTIONS = (
        RateLimitError,
        APIConnectionError,
        APITimeoutError,
        InternalServerError,
    )
except ImportError:  # openai < 1 (unlikely if install() ran)
    from openai import error as openai_error  # type: ignore

    OpenAIAPIError = openai_error.APIError  # type: ignore
    DEFAULT_RETRY_EXCEPTIONS = (
        openai_error.RateLimitError,
        openai_error.ServiceUnavailableError,
        openai_error.APIError,
        openai_error.Timeout,
        openai_error.TryAgain,
    )


def _subprocess_env(env_extra=None):
    """Windows: child Python often uses cp1252 for stdout; emoji in logs then crash unless UTF-8 mode is on."""
    env = os.environ.copy()
    if sys.platform == "win32":
        env.setdefault("PYTHONUTF8", "1")
    if env_extra:
        env.update(env_extra)
    return env


def _job_subprocess_env(base_env: Optional[dict], job_env: Optional[dict]) -> dict:
    """
    Build subprocess env for one dashboard project job.
    Strip parent Flask site vars first so global parallel runs cannot inherit
    the wrong PINTEREST_SITE_ID / PINTEREST_OUT_DIR from the server process.
    """
    env = _subprocess_env(base_env)
    for key in ("PINTEREST_SITE_ID", "PINTEREST_OUT_DIR"):
        env.pop(key, None)
    if job_env:
        env.update(job_env)
    return env


# Child scripts log UTF-8 (emoji, etc.); Windows defaults to cp1252 → UnicodeDecodeError on the pipe.
_SUBPROCESS_STDOUT_KWARGS = {
    "stdout": subprocess.PIPE,
    "stderr": subprocess.STDOUT,
    "text": True,
    "encoding": "utf-8",
    "errors": "replace",
    "bufsize": 1,
}

# A.4-ARTICLES: install compat + retry on ChatCompletion before runpy (used by all runners, not only parallel).
# Top level must be column 0 in -c string.
_A4_ARTICLES_BOOTSTRAP = r"""import os, sys, time, random, socket, runpy
import openai
_root = os.path.abspath(os.path.join(os.getcwd(), ".."))
if _root not in sys.path:
    sys.path.insert(0, _root)
import openai_chat_compat
openai_chat_compat.install()

from openai import (
    APIConnectionError,
    APITimeoutError,
    RateLimitError,
    InternalServerError,
    APIError,
)

_orig = openai.ChatCompletion.create
MAX_TRIES = int(os.getenv("OPENAI_RETRY_MAX_TRIES", "6"))
TIMEOUT = int(os.getenv("OPENAI_REQUEST_TIMEOUT", "180"))

def _retryable(e):
    if isinstance(e, (APIConnectionError, APITimeoutError, RateLimitError, InternalServerError, socket.timeout)):
        return True
    if isinstance(e, APIError):
        code = getattr(e, "status_code", None) or getattr(e, "http_status", None)
        return code in (500, 502, 503, 504)
    return False

def _create_with_retry(*args, **kwargs):
    if "request_timeout" not in kwargs:
        kwargs["request_timeout"] = TIMEOUT

    for attempt in range(1, MAX_TRIES + 1):
        try:
            return _orig(*args, **kwargs)
        except Exception as e:
            if not _retryable(e):
                raise
            wait = min(60, (2 ** (attempt - 1)) + random.random())
            print(f"[OpenAI Retry {attempt}/{MAX_TRIES}] {type(e).__name__}: {e} | waiting {wait:.1f}s...", flush=True)
            time.sleep(wait)

    raise RuntimeError("OpenAI request failed after multiple retries.")

openai.ChatCompletion.create = _create_with_retry
runpy.run_path("A.4-ARTICLES.py", run_name="__main__")
"""


def _popen_pipeline_script(folder_abs: str, script: str, base_env: dict) -> subprocess.Popen:
    """Start a pipeline .py; A.4-ARTICLES uses bootstrap -c for OpenAI retry wrapper."""
    if script == "A.4-ARTICLES.py":
        cmd = [sys.executable, "-u", "-c", _A4_ARTICLES_BOOTSTRAP]
    else:
        cmd = [sys.executable, "-u", script]
    return subprocess.Popen(
        cmd,
        cwd=folder_abs,
        env=base_env,
        **_SUBPROCESS_STDOUT_KWARGS,
    )


def call_with_retries(
        func: Callable,
        *args,
        max_attempts: Optional[int] = None,  # None => إلى أن ينجح
        initial_delay: float = 1.0,
        max_delay: float = 60.0,
        backoff_factor: float = 2.0,
        jitter: float = 0.3,
        retry_on_exceptions: tuple = DEFAULT_RETRY_EXCEPTIONS,
        retry_for_statuses: tuple = (500, 502, 503, 504),
        **kwargs,
) -> Any:
    attempt = 0
    delay = initial_delay
    while True:
        attempt += 1
        try:
            logging.info(f"Attempt {attempt} calling {getattr(func, '__name__', str(func))} ...")
            return func(*args, **kwargs)
        except Exception as e:
            status = getattr(e, "status_code", None) or getattr(e, "http_status", None)
            is_retry_exc = isinstance(e, retry_on_exceptions) or (
                isinstance(e, OpenAIAPIError)
                and status is not None
                and status in retry_for_statuses
            )
            is_retry_status = (status in retry_for_statuses) if status is not None else False
            logging.warning(f"Attempt {attempt} failed: {type(e).__name__}: {e}")
            if not (is_retry_exc or is_retry_status):
                logging.error("Not configured to retry this error. Raising.")
                raise
            if (max_attempts is not None) and (attempt >= max_attempts):
                logging.error(f"Exceeded max_attempts={max_attempts}. Raising.")
                raise
            # backoff + jitter
            jitter_amt = delay * jitter
            sleep_s = max(0.1, min(delay + random.uniform(-jitter_amt, jitter_amt), max_delay))
            logging.info(f"Sleeping {sleep_s:.2f}s before next attempt.")
            time.sleep(sleep_s)
            delay = min(delay * backoff_factor, max_delay)


def chat_completion_with_retry(messages, model="gpt-4o-mini", **kwargs):
    return call_with_retries(
        openai.ChatCompletion.create,
        model=model,
        messages=messages,
        temperature=kwargs.pop("temperature", 0.7),
        max_attempts=None,  # عاود حتى ينجح
        **kwargs,
    )


# -------------------- Repo project discovery (unlimited) --------------------
_APP_ROOT = os.path.dirname(os.path.abspath(__file__))
_APP_CONFIG = os.path.join(_APP_ROOT, "config")
_SKIP_PROJECT_PARENTS = {".git", "ALL", "node_modules", "Save CSV", "__pycache__"}


def _natural_sort_key(name: str):
    return [int(t) if t.isdigit() else t.lower() for t in re.split(r"(\d+)", name)]


def _is_pipeline_project(name: str) -> bool:
    return os.path.isfile(os.path.join(_APP_ROOT, name, "A.1-START.py"))


def _load_app_projects_file() -> dict:
    path = os.path.join(_APP_CONFIG, "projects.json")
    if not os.path.isfile(path):
        return {}
    try:
        with open(path, "r", encoding="utf-8") as f:
            data = json.load(f)
        return data if isinstance(data, dict) else {}
    except (json.JSONDecodeError, OSError):
        return {}


def _load_sites_file_app() -> dict:
    path = os.path.join(_APP_CONFIG, "sites.json")
    if not os.path.isfile(path):
        return {}
    try:
        with open(path, "r", encoding="utf-8") as f:
            d = json.load(f)
        return d if isinstance(d, dict) else {}
    except (json.JSONDecodeError, OSError):
        return {}


def _use_json_sites() -> bool:
    s = _load_sites_file_app().get("sites")
    return isinstance(s, list) and len(s) > 0


# -------------------- Global themes (static-site themes for CF UPLOAD) --------------------
_THEMES_DIR = os.path.join(_APP_ROOT, "themes")


def _themes_dir() -> str:
    return _THEMES_DIR


def _list_theme_slugs() -> list:
    """List subfolders of themes/ that contain a theme.json (sorted alphabetically)."""
    root = _themes_dir()
    if not os.path.isdir(root):
        return []
    out = []
    for name in sorted(os.listdir(root)):
        sub = os.path.join(root, name)
        if not os.path.isdir(sub):
            continue
        if os.path.isfile(os.path.join(sub, "theme.json")):
            out.append(name)
    return out


def _read_theme_meta(slug: str) -> dict:
    """Return theme.json for a slug, augmented with on-disk file presence checks."""
    slug = (slug or "").strip()
    if not slug:
        return {}
    folder = os.path.join(_themes_dir(), slug)
    meta_path = os.path.join(folder, "theme.json")
    if not os.path.isfile(meta_path):
        return {}
    try:
        with open(meta_path, "r", encoding="utf-8") as f:
            meta = json.load(f)
    except (json.JSONDecodeError, OSError):
        meta = {}
    if not isinstance(meta, dict):
        meta = {}
    meta.setdefault("slug", slug)
    meta.setdefault("display_name", slug)
    meta["_folder"] = folder
    meta["_has_index"] = os.path.isfile(os.path.join(folder, "index.html"))
    meta["_has_article"] = os.path.isfile(os.path.join(folder, "article.html"))
    meta["_has_static"] = os.path.isdir(os.path.join(folder, "static"))
    return meta


def _list_themes_full() -> list:
    """All themes as dicts, including a synthesized 'valid' flag."""
    out = []
    for slug in _list_theme_slugs():
        meta = _read_theme_meta(slug)
        if not meta:
            continue
        meta["valid"] = bool(meta.get("_has_index") and meta.get("_has_article"))
        out.append(meta)
    return out


def _theme_exists(slug: str) -> bool:
    slug = (slug or "").strip()
    if not slug:
        return False
    return slug in set(_list_theme_slugs())


def _pipeline_code_folder() -> str:
    return (str(_load_sites_file_app().get("pipeline_code_folder") or "A1-Pinterest_01").strip() or "A1-Pinterest_01")


def _safe_log_dom_id(s: str) -> str:
    x = re.sub(r"[^a-zA-Z0-9_-]+", "_", (s or "").strip().strip("_"))
    return x or "log"


def _site_row_public_title(s: dict) -> str:
    """Name shown in dashboard: A1-Pinterest_01, B1-Pinterest_51, …"""
    t = s.get("display_name") or s.get("name") or s.get("id") or "site"
    return str(t).strip()


def flat_run_units() -> list:
    """
    One entry per 'project' the UI runs: either one row per real folder, or
    (when config/sites.json has sites) one row per site using the same code folder.

    - label: title for headers + ?project= (e.g. A1-Pinterest_01)
    - log_id: safe id for #log_{log_id} and SSE routing (getElementById)
    - env: PINTEREST_SITE_ID stays the internal `id` from JSON
    """
    if not _use_json_sites():
        out = []
        for f in PROJECT_FOLDERS:
            out.append(
                {
                    "folder": f,
                    "label": f,
                    "log_id": _safe_log_dom_id(f),
                    "env": {},
                }
            )
        return out
    root = _pipeline_code_folder()
    root_path = os.path.join(_APP_ROOT, root)
    if not os.path.isdir(root_path):
        return [
            {
                "folder": f,
                "label": f,
                "log_id": _safe_log_dom_id(f),
                "env": {},
            }
            for f in PROJECT_FOLDERS
        ]
    out = []
    for s in _load_sites_file_app().get("sites", []):
        if not isinstance(s, dict):
            continue
        sid = str(s.get("id", "")).strip()
        if not sid:
            continue
        title = _site_row_public_title(s)
        log_id = (str(s.get("log_id", "")).strip() or _safe_log_dom_id(title) or _safe_log_dom_id(sid))
        out_dir = str(s.get("out_dir", "") or f"{sid}-out").strip()
        out.append(
            {
                "folder": root,
                "label": title,
                "log_id": log_id,
                "env": {"PINTEREST_SITE_ID": sid, "PINTEREST_OUT_DIR": out_dir},
            }
        )
    return out


def all_out_name_for_label(label: str) -> str:
    """ALL/<out_dir>/... — label can be display_name or id from sites.json."""
    if _use_json_sites():
        for s in _load_sites_file_app().get("sites", []):
            if not isinstance(s, dict):
                continue
            sid = str(s.get("id", "")).strip()
            stitle = _site_row_public_title(s) if s else ""
            if sid == label or stitle == label:
                return (s.get("out_dir") or f"{s.get('id')}-out").strip()  # type: ignore[union-attr]
    return f"{label}-out"


def _project_excel_path_by_out_dir(out_dir: str) -> str:
    """
    Unified pipeline workbook:
    - Prefer Recipes.xlsx (new single-file flow)
    - Fallback to images.xlsx for backward compatibility
    """
    base = os.path.join(os.getcwd(), "ALL", out_dir)
    recipes_path = os.path.join(base, "Recipes.xlsx")
    if os.path.exists(recipes_path):
        return recipes_path
    return os.path.join(base, "images.xlsx")


def _is_filled_excel_value(v) -> bool:
    if v is None:
        return False
    s = str(v).strip()
    if not s:
        return False
    return s.lower() != "nan"


def _json_for_inline_script(obj) -> str:
    """
    JSON safe to paste inside HTML <script>...</script>.
    Escapes '<' so '</script>' in data cannot terminate the script tag.
    """
    return json.dumps(obj, ensure_ascii=False, separators=(",", ":")).replace("<", "\\u003c")


_STARTS_TOTAL_CACHE = {"at": 0.0, "value": 0}
_PROJECT_STATS_CACHE: dict = {}
_GLOBAL_START_ROTATION = {"cursor": 0}


def _total_titles_in_starts_cached(ttl_seconds: float = 180.0) -> int:
    now = time.time()
    at = float(_STARTS_TOTAL_CACHE.get("at", 0.0) or 0.0)
    if now - at <= ttl_seconds:
        return int(_STARTS_TOTAL_CACHE.get("value", 0) or 0)

    starts_dir = os.path.join(_APP_ROOT, "STARTS")
    if not os.path.isdir(starts_dir):
        _STARTS_TOTAL_CACHE["at"] = now
        _STARTS_TOTAL_CACHE["value"] = 0
        return 0

    total = 0
    xlsx_files = [
        f
        for f in os.listdir(starts_dir)
        if f.lower().endswith(".xlsx")
        and not f.startswith("~$")
        and not f.startswith("._")
        and os.path.isfile(os.path.join(starts_dir, f))
    ]
    for fn in xlsx_files:
        fp = os.path.join(starts_dir, fn)
        try:
            for rec in _read_titles_from_start_workbook(fp):
                if str(rec.get("title", "")).strip():
                    total += 1
        except Exception:
            continue

    _STARTS_TOTAL_CACHE["at"] = now
    _STARTS_TOTAL_CACHE["value"] = int(total)
    return int(total)


def _project_column_stats(project_label: str) -> dict:
    out_dir = all_out_name_for_label(project_label)
    file_path = _project_excel_path_by_out_dir(out_dir)
    if not os.path.exists(file_path):
        return {"ok": False, "error": "excel_not_found", "file_path": file_path}
    try:
        mtime = float(os.path.getmtime(file_path))
    except OSError:
        mtime = 0.0
    global_total = _total_titles_in_starts_cached()
    ck = str(project_label)
    cached = _PROJECT_STATS_CACHE.get(ck)
    # Strong cache: if workbook file didn't change, reuse computed stats.
    # Previous 8s TTL forced expensive re-scan loops even with unchanged files.
    if isinstance(cached, dict) and float(cached.get("mtime", -1.0)) == mtime:
        return dict(cached.get("data") or {})
    try:
        wb = openpyxl.load_workbook(file_path, data_only=True, read_only=True)
    except Exception as e:
        return {"ok": False, "error": f"open_failed: {e}", "file_path": file_path}
    try:
        sh = wb.active
        max_col = int(sh.max_column or 0)
        max_row = int(sh.max_row or 0)
        if max_col <= 0:
            return {"ok": True, "total_titles": 0, "columns": []}
        # Fast path: read headers + rows with values_only to avoid per-cell access cost.
        header_values = next(
            sh.iter_rows(min_row=1, max_row=1, min_col=1, max_col=max_col, values_only=True),
            tuple(),
        )
        headers = []
        for i in range(max_col):
            hv = header_values[i] if i < len(header_values) else None
            h = str(hv).strip() if hv is not None else f"Column_{i+1}"
            headers.append(h or f"Column_{i+1}")

        total_rows = max(0, max_row - 1)
        filled_counts = [0 for _ in range(max_col)]
        title_filled_rows = 0
        title_col_idx = -1
        for i, h in enumerate(headers):
            if str(h).strip().lower() == "title":
                title_col_idx = i
                break
        if total_rows > 0:
            for row_vals in sh.iter_rows(
                min_row=2,
                max_row=max_row,
                min_col=1,
                max_col=max_col,
                values_only=True,
            ):
                ln = len(row_vals)
                for i in range(max_col):
                    v = row_vals[i] if i < ln else None
                    if _is_filled_excel_value(v):
                        filled_counts[i] += 1
                if title_col_idx >= 0:
                    tv = row_vals[title_col_idx] if title_col_idx < ln else None
                    if _is_filled_excel_value(tv):
                        title_filled_rows += 1
        # Denominator is number of rows with a non-empty Title.
        total_for_stats = int(title_filled_rows if title_filled_rows > 0 else total_rows)
        # Show full pipeline coverage: expected columns first, then any extra columns from the sheet.
        expected_columns = [
            "Title",
            "Recipe",
            "Generated At",
            "Json Recipe",
            "Prompt",
            "Prompt Image Ingredients",
            "main_image",
            "image_1",
            "image_2",
            "image_3",
            "image_4",
            "statu",
            "error",
            "main_image_ingredients",
            "image_ing_1",
            "image_ing_2",
            "image_ing_3",
            "image_ing_4",
            "statu_ing",
            "article",
            "recipe_title_pin",
            "pinterest_title",
            "pinterest_description",
            "pinterest_keywords",
            "rank_math_focus_keyword",
            "rank_math_description",
            "rank_math_pillar_content",
            "category",
            "categories",
            "pinterest_image",
            "status",
            "post_url",
        ]

        by_lower = {str(h).strip().lower(): i for i, h in enumerate(headers)}
        cols = []
        seen_lowers = set()

        for name in expected_columns:
            lk = name.strip().lower()
            seen_lowers.add(lk)
            if lk in by_lower:
                idx = by_lower[lk]
                cols.append({"name": headers[idx], "filled": int(filled_counts[idx]), "total": int(total_for_stats)})
            else:
                cols.append({"name": name, "filled": 0, "total": int(total_for_stats)})

        for i, name in enumerate(headers):
            lk = str(name).strip().lower()
            if lk in seen_lowers:
                continue
            cols.append({"name": name, "filled": int(filled_counts[i]), "total": int(total_for_stats)})
        data = {
            "ok": True,
            "file_path": file_path,
            "total_titles": int(total_rows),
            "global_total_titles": int(global_total),
            "columns": cols,
        }
        _PROJECT_STATS_CACHE[ck] = {
            "at": time.time(),
            "mtime": mtime,
            "global_total_titles": int(global_total),
            "data": data,
        }
        return data
    finally:
        wb.close()


def _all_dir_path() -> str:
    return os.path.join(_APP_ROOT, "ALL")


def _value_to_all_url(val: str) -> str:
    """If `val` points to a file inside ALL/, return a /files/all/... URL.

    Accepts absolute or relative paths using either / or \\ separators.
    Returns an empty string when the value does not resolve inside ALL/.
    """
    s = str(val or "").strip().strip('"').strip("'")
    if not s:
        return ""
    norm = s.replace("\\", "/")
    all_root = os.path.abspath(_all_dir_path())
    rel = ""
    if os.path.isabs(s) or (len(s) >= 2 and s[1] == ":"):
        try:
            ap = os.path.abspath(s)
        except (OSError, ValueError):
            return ""
        try:
            r = os.path.relpath(ap, all_root)
        except ValueError:
            return ""
        if r.startswith("..") or os.path.isabs(r):
            return ""
        rel = r
    else:
        low = norm.lower()
        marker = "all/"
        idx = low.find("/" + marker)
        if idx >= 0:
            rel = norm[idx + 1 + len(marker):]
        elif low.startswith(marker):
            rel = norm[len(marker):]
        else:
            return ""
    rel = rel.replace("\\", "/").lstrip("/")
    if not rel or ".." in rel.split("/"):
        return ""
    target = os.path.abspath(os.path.join(all_root, rel))
    if os.path.commonpath([target, all_root]) != all_root:
        return ""
    if not os.path.isfile(target):
        return ""
    from urllib.parse import quote
    return "/files/all/" + quote(rel, safe="/")


def _project_column_details(project_label: str, column_name: str) -> dict:
    out_dir = all_out_name_for_label(project_label)
    file_path = _project_excel_path_by_out_dir(out_dir)
    if not os.path.exists(file_path):
        return {"ok": False, "error": "excel_not_found", "file_path": file_path}
    try:
        wb = openpyxl.load_workbook(file_path, data_only=True, read_only=True)
    except Exception as e:
        return {"ok": False, "error": f"open_failed: {e}", "file_path": file_path}
    try:
        sh = wb.active
        max_col = int(sh.max_column or 0)
        max_row = int(sh.max_row or 0)
        if max_col <= 0:
            return {"ok": True, "column": column_name, "rows": [], "total_titles": 0}

        headers = []
        for c in range(1, max_col + 1):
            hv = sh.cell(row=1, column=c).value
            h = str(hv).strip() if hv is not None else f"Column_{c}"
            headers.append(h or f"Column_{c}")

        hmap = {str(h).strip().lower(): i + 1 for i, h in enumerate(headers)}
        req_col = hmap.get(str(column_name or "").strip().lower())
        title_col = hmap.get("title", 1)
        total_rows = max(0, max_row - 1)

        rows = []
        for r in range(2, max_row + 1):
            title_v = sh.cell(row=r, column=title_col).value if title_col else None
            title_s = str(title_v).strip() if title_v is not None else ""
            if req_col:
                val = sh.cell(row=r, column=req_col).value
                val_s = str(val).strip() if val is not None else ""
                filled = _is_filled_excel_value(val)
            else:
                val_s = ""
                filled = False
            rows.append(
                {
                    "row": r - 1,
                    "title": title_s,
                    "value": val_s,
                    "value_url": _value_to_all_url(val_s),
                    "filled": bool(filled),
                }
            )

        return {
            "ok": True,
            "column": column_name,
            "column_exists": bool(req_col),
            "rows": rows,
            "total_titles": int(total_rows),
            "file_path": file_path,
        }
    finally:
        wb.close()


def _starts_dir_path() -> str:
    return os.path.join(_APP_ROOT, "STARTS")


def _safe_start_file_path(file_name: str) -> str:
    base = os.path.basename(str(file_name or "").strip())
    if not base:
        raise ValueError("Missing file name")
    if base != str(file_name or "").strip():
        raise ValueError("Invalid file name")
    if not base.lower().endswith(".xlsx"):
        raise ValueError("File must end with .xlsx")
    if base.startswith("~$"):
        raise ValueError("Temporary Excel files are not allowed")
    starts_dir = _starts_dir_path()
    os.makedirs(starts_dir, exist_ok=True)
    fp = os.path.abspath(os.path.join(starts_dir, base))
    if os.path.dirname(fp) != os.path.abspath(starts_dir):
        raise ValueError("Invalid file path")
    return fp


def _project_start_file_path(project_label: str) -> str:
    """Per-project START workbook path under ALL/<out_dir>/."""
    if not project_label or not is_allowed_project_label(project_label):
        raise ValueError("Unknown project")
    out_dir = all_out_name_for_label(project_label)
    base = os.path.join(_APP_ROOT, "ALL", out_dir)
    os.makedirs(base, exist_ok=True)

    start_name = "START.xlsx"
    if _use_json_sites():
        d = _load_sites_file_app()
        sites = d.get("sites") if isinstance(d, dict) else None
        if isinstance(sites, list):
            for s in sites:
                if not isinstance(s, dict):
                    continue
                title = _site_row_public_title(s)
                sid = str(s.get("id", "") or "").strip()
                lid = str(s.get("log_id", "") or "").strip()
                if project_label in {title, sid, lid}:
                    nm = str(s.get("start_file", "") or "").strip()
                    if nm:
                        start_name = os.path.basename(nm) or "START.xlsx"
                    break
    return os.path.join(base, start_name)


def _unique_headers(header_values: list) -> list:
    """
    Make duplicate Excel headers unique for JSON/table rendering.
    Example: used, used, used -> used, used.1, used.2
    """
    out = []
    counts = {}
    for raw in header_values:
        base = str(raw).strip() if raw is not None else ""
        if not base:
            base = "Column"
        n = int(counts.get(base, 0))
        if n <= 0:
            name = base
        else:
            name = f"{base}.{n}"
        counts[base] = n + 1
        out.append(name)
    return out


@app.route("/api/starts-files")
def api_starts_files():
    mode = (request.args.get("mode") or "projects").strip().lower()
    if mode != "legacy":
        items = []
        for u in flat_run_units():
            project = str(u.get("label", "") or "").strip()
            if not project:
                continue
            try:
                fp = _project_start_file_path(project)
            except ValueError:
                continue
            exists = os.path.isfile(fp)
            row_count = 0
            title_count = 0
            if exists:
                try:
                    wb = openpyxl.load_workbook(fp, data_only=True, read_only=True)
                    try:
                        sh = wb.active
                        max_row = int(sh.max_row or 0)
                        row_count = max(0, max_row - 1)
                        max_col = int(sh.max_column or 0)
                        title_col = 1
                        for c in range(1, max_col + 1):
                            hv = sh.cell(row=1, column=c).value
                            if hv is not None and str(hv).strip().lower() == "title":
                                title_col = c
                                break
                        for r in range(2, max_row + 1):
                            v = sh.cell(row=r, column=title_col).value
                            if _is_filled_excel_value(v):
                                title_count += 1
                    finally:
                        wb.close()
                except Exception:
                    pass
            items.append(
                {
                    "project": project,
                    "name": os.path.basename(fp),
                    "row_count": int(row_count),
                    "title_count": int(title_count),
                    "path": fp,
                    "exists": bool(exists),
                }
            )
        return jsonify({"ok": True, "mode": "projects", "items": items})

    pattern = (request.args.get("pattern") or "START*.xlsx").strip() or "START*.xlsx"
    starts_dir = _starts_dir_path()
    if not os.path.isdir(starts_dir):
        return jsonify({"ok": True, "items": [], "starts_dir": starts_dir})
    items = []
    for fn in sorted(os.listdir(starts_dir), key=_natural_sort_key):
        if not fn.lower().endswith(".xlsx"):
            continue
        if fn.startswith("~$") or fn.startswith("._") or fn.startswith("_"):
            continue
        if not fnmatch.fnmatch(fn.lower(), pattern.lower()):
            continue
        fp = os.path.join(starts_dir, fn)
        row_count = 0
        title_count = 0
        try:
            wb = openpyxl.load_workbook(fp, data_only=True, read_only=True)
            try:
                sh = wb.active
                max_row = int(sh.max_row or 0)
                row_count = max(0, max_row - 1)
                max_col = int(sh.max_column or 0)
                title_col = 1
                for c in range(1, max_col + 1):
                    hv = sh.cell(row=1, column=c).value
                    if hv is not None and str(hv).strip().lower() == "title":
                        title_col = c
                        break
                for r in range(2, max_row + 1):
                    v = sh.cell(row=r, column=title_col).value
                    if _is_filled_excel_value(v):
                        title_count += 1
            finally:
                wb.close()
        except Exception:
            pass
        items.append(
            {
                "name": fn,
                "row_count": int(row_count),
                "title_count": int(title_count),
                "path": fp,
            }
        )
    return jsonify({"ok": True, "items": items, "starts_dir": starts_dir, "pattern": pattern})


@app.route("/api/starts-read")
def api_starts_read():
    project = (request.args.get("project") or "").strip()
    file_name = (request.args.get("file") or "").strip()
    try:
        fp = _project_start_file_path(project) if project else _safe_start_file_path(file_name)
    except ValueError as e:
        return jsonify({"ok": False, "error": str(e)}), 400
    if not os.path.isfile(fp):
        return jsonify({"ok": False, "error": "File not found", "file": file_name}), 404

    limit = int(request.args.get("limit", 200) or 200)
    offset = int(request.args.get("offset", 0) or 0)
    if limit < 1:
        limit = 1
    if limit > 2000:
        limit = 2000
    if offset < 0:
        offset = 0

    wb = openpyxl.load_workbook(fp, data_only=True, read_only=True)
    try:
        sh = wb.active
        max_col = int(sh.max_column or 0)
        max_row = int(sh.max_row or 0)
        raw_headers = []
        if max_col > 0:
            for c in range(1, max_col + 1):
                hv = sh.cell(row=1, column=c).value
                h = str(hv).strip() if hv is not None else f"Column_{c}"
                raw_headers.append(h or f"Column_{c}")
        headers = _unique_headers(raw_headers)

        start_excel_row = 2 + offset
        end_excel_row = min(max_row, start_excel_row + limit - 1)
        rows = []
        if max_row >= 2 and max_col > 0 and start_excel_row <= end_excel_row:
            for r in range(start_excel_row, end_excel_row + 1):
                row_obj = {"excel_row": int(r)}
                for c in range(1, max_col + 1):
                    key = headers[c - 1]
                    row_obj[key] = sh.cell(row=r, column=c).value
                rows.append(row_obj)
        return jsonify(
            {
                "ok": True,
                "project": project,
                "file": file_name,
                "headers": headers,
                "rows": rows,
                "total_rows": max(0, max_row - 1),
                "offset": offset,
                "limit": limit,
            }
        )
    finally:
        wb.close()


@app.route("/api/starts-create", methods=["POST"])
def api_starts_create():
    data = request.get_json(force=True, silent=True) or {}
    file_name = str(data.get("file") or "").strip()
    headers = data.get("headers")
    if not isinstance(headers, list) or not headers:
        headers = ["Title"]
    hdrs = [str(h).strip() for h in headers if str(h or "").strip()]
    if not hdrs:
        hdrs = ["Title"]
    if not any(h.lower() == "title" for h in hdrs):
        hdrs.insert(0, "Title")
    try:
        fp = _safe_start_file_path(file_name)
    except ValueError as e:
        return jsonify({"ok": False, "error": str(e)}), 400
    if os.path.exists(fp):
        return jsonify({"ok": False, "error": "File already exists", "file": file_name}), 409
    wb = openpyxl.Workbook()
    try:
        sh = wb.active
        sh.title = "Titles"
        for i, h in enumerate(hdrs, start=1):
            sh.cell(row=1, column=i, value=h)
        wb.save(fp)
    finally:
        wb.close()
    return jsonify({"ok": True, "file": file_name, "path": fp, "headers": hdrs})


@app.route("/api/starts-add-rows", methods=["POST"])
def api_starts_add_rows():
    data = request.get_json(force=True, silent=True) or {}
    project = str(data.get("project") or "").strip()
    file_name = str(data.get("file") or "").strip()
    titles = data.get("titles")
    rows = data.get("rows")
    try:
        fp = _project_start_file_path(project) if project else _safe_start_file_path(file_name)
    except ValueError as e:
        return jsonify({"ok": False, "error": str(e)}), 400
    if not os.path.isfile(fp):
        return jsonify({"ok": False, "error": "File not found", "file": file_name}), 404

    normalized_rows = []
    if isinstance(titles, list):
        for t in titles:
            s = str(t or "").strip()
            if s:
                normalized_rows.append({"Title": s})
    if isinstance(rows, list):
        for one in rows:
            if isinstance(one, dict):
                obj = {}
                for k, v in one.items():
                    kk = str(k or "").strip()
                    if not kk:
                        continue
                    obj[kk] = v
                if obj:
                    normalized_rows.append(obj)
    if not normalized_rows:
        return jsonify({"ok": False, "error": "Provide titles[] or rows[]"}), 400

    wb = openpyxl.load_workbook(fp)
    try:
        sh = wb.active
        max_col = int(sh.max_column or 1)
        header_to_col = {}
        for c in range(1, max_col + 1):
            hv = sh.cell(row=1, column=c).value
            if hv is None:
                continue
            header_to_col[str(hv).strip().lower()] = c
        if "title" not in header_to_col:
            sh.cell(row=1, column=1, value="Title")
            header_to_col["title"] = 1
            if max_col < 1:
                max_col = 1

        def ensure_col(name: str) -> int:
            nonlocal max_col
            lk = str(name).strip().lower()
            c0 = header_to_col.get(lk)
            if c0:
                return c0
            max_col = max(max_col, int(sh.max_column or 1)) + 1
            sh.cell(row=1, column=max_col, value=name)
            header_to_col[lk] = max_col
            return max_col

        start_row = int(sh.max_row or 1) + 1
        for i, row_obj in enumerate(normalized_rows):
            rr = start_row + i
            for k, v in row_obj.items():
                col = ensure_col(k)
                sh.cell(row=rr, column=col, value=v)
        wb.save(fp)
        return jsonify(
            {
                "ok": True,
                "project": project,
                "file": file_name,
                "added_rows": len(normalized_rows),
                "start_excel_row": int(start_row),
                "end_excel_row": int(start_row + len(normalized_rows) - 1),
            }
        )
    finally:
        wb.close()


@app.route("/api/starts-update-row", methods=["POST"])
def api_starts_update_row():
    data = request.get_json(force=True, silent=True) or {}
    project = str(data.get("project") or "").strip()
    file_name = str(data.get("file") or "").strip()
    excel_row = int(data.get("excel_row", 0) or 0)
    values = data.get("values")
    values_by_col = data.get("values_by_col")
    if excel_row < 2:
        return jsonify({"ok": False, "error": "excel_row must be >= 2"}), 400
    if (not isinstance(values, dict) or not values) and (not isinstance(values_by_col, dict) or not values_by_col):
        return jsonify({"ok": False, "error": "values or values_by_col is required"}), 400
    try:
        fp = _project_start_file_path(project) if project else _safe_start_file_path(file_name)
    except ValueError as e:
        return jsonify({"ok": False, "error": str(e)}), 400
    if not os.path.isfile(fp):
        return jsonify({"ok": False, "error": "File not found", "file": file_name}), 404

    wb = openpyxl.load_workbook(fp)
    try:
        sh = wb.active
        max_col = int(sh.max_column or 1)
        header_to_col = {}
        for c in range(1, max_col + 1):
            hv = sh.cell(row=1, column=c).value
            if hv is None:
                continue
            header_to_col[str(hv).strip().lower()] = c

        def ensure_col(name: str) -> int:
            nonlocal max_col
            lk = str(name or "").strip().lower()
            if not lk:
                raise ValueError("Invalid column name")
            c0 = header_to_col.get(lk)
            if c0:
                return c0
            max_col = max(max_col, int(sh.max_column or 1)) + 1
            sh.cell(row=1, column=max_col, value=str(name).strip())
            header_to_col[lk] = max_col
            return max_col

        if isinstance(values_by_col, dict) and values_by_col:
            for k, v in values_by_col.items():
                try:
                    col = int(k)
                except (TypeError, ValueError):
                    continue
                if col < 1:
                    continue
                sh.cell(row=excel_row, column=col, value=v)
        if isinstance(values, dict) and values:
            for k, v in values.items():
                col = ensure_col(str(k))
                sh.cell(row=excel_row, column=col, value=v)
        wb.save(fp)
        return jsonify({"ok": True, "project": project, "file": file_name, "excel_row": excel_row})
    finally:
        wb.close()


@app.route("/api/starts-delete-rows", methods=["POST"])
def api_starts_delete_rows():
    data = request.get_json(force=True, silent=True) or {}
    project = str(data.get("project") or "").strip()
    file_name = str(data.get("file") or "").strip()
    rows = data.get("excel_rows")
    if not isinstance(rows, list) or not rows:
        return jsonify({"ok": False, "error": "excel_rows list is required"}), 400
    try:
        fp = _project_start_file_path(project) if project else _safe_start_file_path(file_name)
    except ValueError as e:
        return jsonify({"ok": False, "error": str(e)}), 400
    if not os.path.isfile(fp):
        return jsonify({"ok": False, "error": "File not found", "file": file_name}), 404

    to_delete = []
    for r in rows:
        try:
            rr = int(r)
        except (TypeError, ValueError):
            continue
        if rr >= 2:
            to_delete.append(rr)
    to_delete = sorted(set(to_delete), reverse=True)
    if not to_delete:
        return jsonify({"ok": False, "error": "No valid excel rows to delete"}), 400

    wb = openpyxl.load_workbook(fp)
    try:
        sh = wb.active
        deleted = 0
        max_row = int(sh.max_row or 1)
        for rr in to_delete:
            if 2 <= rr <= max_row:
                sh.delete_rows(rr, 1)
                deleted += 1
        wb.save(fp)
        return jsonify({"ok": True, "project": project, "file": file_name, "deleted_rows": deleted})
    finally:
        wb.close()


@app.route("/api/starts-clear-file", methods=["POST"])
def api_starts_clear_file():
    data = request.get_json(force=True, silent=True) or {}
    project = str(data.get("project") or "").strip()
    file_name = str(data.get("file") or "").strip()
    try:
        fp = _project_start_file_path(project) if project else _safe_start_file_path(file_name)
    except ValueError as e:
        return jsonify({"ok": False, "error": str(e)}), 400
    if not os.path.isfile(fp):
        return jsonify({"ok": False, "error": "File not found", "file": file_name}), 404

    wb = openpyxl.load_workbook(fp)
    try:
        sh = wb.active
        max_row = int(sh.max_row or 1)
        max_col = int(sh.max_column or 1)
        cleared = 0
        if max_row >= 2:
            for r in range(2, max_row + 1):
                for c in range(1, max_col + 1):
                    sh.cell(row=r, column=c, value=None)
                cleared += 1
        wb.save(fp)
        return jsonify({"ok": True, "project": project, "file": file_name, "cleared_rows": int(cleared)})
    finally:
        wb.close()


@app.route("/api/recipes-files")
def api_recipes_files():
    items = []
    for u in flat_run_units():
        label = str(u.get("label", "") or "").strip()
        if not label:
            continue
        out_dir = all_out_name_for_label(label)
        file_path = _project_excel_path_by_out_dir(out_dir)
        exists = os.path.isfile(file_path)
        row_count = 0
        title_count = 0
        if exists:
            try:
                wb = openpyxl.load_workbook(file_path, data_only=True, read_only=True)
                try:
                    sh = wb.active
                    max_row = int(sh.max_row or 0)
                    max_col = int(sh.max_column or 0)
                    row_count = max(0, max_row - 1)
                    title_col = 1
                    for c in range(1, max_col + 1):
                        hv = sh.cell(row=1, column=c).value
                        if hv is not None and str(hv).strip().lower() == "title":
                            title_col = c
                            break
                    for r in range(2, max_row + 1):
                        if _is_filled_excel_value(sh.cell(row=r, column=title_col).value):
                            title_count += 1
                finally:
                    wb.close()
            except Exception:
                pass
        items.append(
            {
                "project": label,
                "out_dir": out_dir,
                "file_path": file_path,
                "exists": bool(exists),
                "row_count": int(row_count),
                "title_count": int(title_count),
            }
        )
    return jsonify({"ok": True, "items": items})


@app.route("/api/recipes-read")
def api_recipes_read():
    project = (request.args.get("project") or "").strip()
    if not project or not is_allowed_project_label(project):
        return jsonify({"ok": False, "error": "Unknown project"}), 404
    out_dir = all_out_name_for_label(project)
    file_path = _project_excel_path_by_out_dir(out_dir)
    if not os.path.isfile(file_path):
        return jsonify({"ok": False, "error": "Recipes file not found", "project": project}), 404

    limit = int(request.args.get("limit", 200) or 200)
    offset = int(request.args.get("offset", 0) or 0)
    limit = max(1, min(limit, 2000))
    offset = max(0, offset)

    wb = openpyxl.load_workbook(file_path, data_only=True, read_only=True)
    try:
        sh = wb.active
        max_col = int(sh.max_column or 0)
        max_row = int(sh.max_row or 0)
        raw_headers = []
        if max_col > 0:
            for c in range(1, max_col + 1):
                hv = sh.cell(row=1, column=c).value
                h = str(hv).strip() if hv is not None else f"Column_{c}"
                raw_headers.append(h or f"Column_{c}")
        headers = _unique_headers(raw_headers)

        rows = []
        start_excel_row = 2 + offset
        end_excel_row = min(max_row, start_excel_row + limit - 1)
        if max_col > 0 and max_row >= 2 and start_excel_row <= end_excel_row:
            for r in range(start_excel_row, end_excel_row + 1):
                row_obj = {"excel_row": int(r)}
                for c in range(1, max_col + 1):
                    row_obj[headers[c - 1]] = sh.cell(row=r, column=c).value
                rows.append(row_obj)
        return jsonify(
            {
                "ok": True,
                "project": project,
                "file_path": file_path,
                "headers": headers,
                "rows": rows,
                "total_rows": max(0, max_row - 1),
                "offset": offset,
                "limit": limit,
            }
        )
    finally:
        wb.close()


@app.route("/api/recipes-update-row", methods=["POST"])
def api_recipes_update_row():
    data = request.get_json(force=True, silent=True) or {}
    project = str(data.get("project") or "").strip()
    excel_row = int(data.get("excel_row", 0) or 0)
    values = data.get("values")
    values_by_col = data.get("values_by_col")
    if not project or not is_allowed_project_label(project):
        return jsonify({"ok": False, "error": "Unknown project"}), 404
    if excel_row < 2:
        return jsonify({"ok": False, "error": "excel_row must be >= 2"}), 400
    if (not isinstance(values, dict) or not values) and (not isinstance(values_by_col, dict) or not values_by_col):
        return jsonify({"ok": False, "error": "values or values_by_col is required"}), 400

    out_dir = all_out_name_for_label(project)
    file_path = _project_excel_path_by_out_dir(out_dir)
    if not os.path.isfile(file_path):
        return jsonify({"ok": False, "error": "Recipes file not found", "project": project}), 404

    wb = openpyxl.load_workbook(file_path)
    try:
        sh = wb.active
        max_col = int(sh.max_column or 1)
        header_to_col = {}
        for c in range(1, max_col + 1):
            hv = sh.cell(row=1, column=c).value
            if hv is None:
                continue
            header_to_col[str(hv).strip().lower()] = c

        def ensure_col(name: str) -> int:
            nonlocal max_col
            lk = str(name or "").strip().lower()
            if not lk:
                raise ValueError("Invalid column name")
            c0 = header_to_col.get(lk)
            if c0:
                return c0
            max_col = max(max_col, int(sh.max_column or 1)) + 1
            sh.cell(row=1, column=max_col, value=str(name).strip())
            header_to_col[lk] = max_col
            return max_col

        if isinstance(values_by_col, dict) and values_by_col:
            for k, v in values_by_col.items():
                try:
                    col = int(k)
                except (TypeError, ValueError):
                    continue
                if col < 1:
                    continue
                sh.cell(row=excel_row, column=col, value=v)
        if isinstance(values, dict) and values:
            for k, v in values.items():
                col = ensure_col(str(k))
                sh.cell(row=excel_row, column=col, value=v)
        wb.save(file_path)
        return jsonify({"ok": True, "project": project, "excel_row": excel_row})
    finally:
        wb.close()


@app.route("/api/recipes-delete-rows", methods=["POST"])
def api_recipes_delete_rows():
    data = request.get_json(force=True, silent=True) or {}
    project = str(data.get("project") or "").strip()
    rows = data.get("excel_rows")
    if not project or not is_allowed_project_label(project):
        return jsonify({"ok": False, "error": "Unknown project"}), 404
    if not isinstance(rows, list) or not rows:
        return jsonify({"ok": False, "error": "excel_rows list is required"}), 400

    out_dir = all_out_name_for_label(project)
    file_path = _project_excel_path_by_out_dir(out_dir)
    if not os.path.isfile(file_path):
        return jsonify({"ok": False, "error": "Recipes file not found", "project": project}), 404

    to_delete = []
    for r in rows:
        try:
            rr = int(r)
        except (TypeError, ValueError):
            continue
        if rr >= 2:
            to_delete.append(rr)
    to_delete = sorted(set(to_delete), reverse=True)
    if not to_delete:
        return jsonify({"ok": False, "error": "No valid excel rows to delete"}), 400

    wb = openpyxl.load_workbook(file_path)
    try:
        sh = wb.active
        deleted = 0
        max_row = int(sh.max_row or 1)
        for rr in to_delete:
            if 2 <= rr <= max_row:
                sh.delete_rows(rr, 1)
                deleted += 1
        wb.save(file_path)
        return jsonify({"ok": True, "project": project, "deleted_rows": deleted})
    finally:
        wb.close()


@app.route("/api/recipes-clear-file", methods=["POST"])
def api_recipes_clear_file():
    data = request.get_json(force=True, silent=True) or {}
    project = str(data.get("project") or "").strip()
    if not project or not is_allowed_project_label(project):
        return jsonify({"ok": False, "error": "Unknown project"}), 404

    out_dir = all_out_name_for_label(project)
    file_path = _project_excel_path_by_out_dir(out_dir)
    if not os.path.isfile(file_path):
        return jsonify({"ok": False, "error": "Recipes file not found", "project": project}), 404

    wb = openpyxl.load_workbook(file_path)
    try:
        sh = wb.active
        max_row = int(sh.max_row or 1)
        max_col = int(sh.max_column or 1)
        cleared = 0
        if max_row >= 2:
            for r in range(2, max_row + 1):
                for c in range(1, max_col + 1):
                    sh.cell(row=r, column=c, value=None)
                cleared += 1
        wb.save(file_path)
        return jsonify({"ok": True, "project": project, "cleared_rows": int(cleared)})
    finally:
        wb.close()


def _normalize_script_jobs(script_names) -> list:
    """
    (folder, script) | (folder, script, env) | 4- or 5-tuple
    -> (folder, script, env, log_id, line_label)
    log_id: SSE "folder" key and #log_{log_id}
    line_label: text in log lines
    """
    out = []
    for item in script_names:
        if len(item) == 2:
            f, s = item[0], item[1]
            out.append((f, s, {}, _safe_log_dom_id(f), f))
        elif len(item) == 3:
            f, s, e = item[0], item[1], item[2] or {}
            line = f"{f}[{e.get('PINTEREST_SITE_ID')}]" if e.get("PINTEREST_SITE_ID") else f
            out.append((f, s, e, _safe_log_dom_id(line), line))
        elif len(item) == 4:
            f, s, e, a = item[0], item[1], item[2] or {}, item[3]
            out.append((f, s, e, _safe_log_dom_id(a), a))
        else:
            f, s, e, log_id, line_label = item[0], item[1], item[2] or {}, item[3], item[4]
            out.append((f, s, e, log_id, line_label))
    return out


def jobs_for_script(script: str) -> list:
    return [
        (u["folder"], script, u.get("env") or {}, u["log_id"], u["label"])
        for u in flat_run_units()
    ]


def jobs_for_imagine_unit_group(units: list) -> list:
    return [
        (u["folder"], "A.3-IMAGINE.py", u.get("env") or {}, u["log_id"], u["label"])
        for u in (units or [])
    ]


def _unit_by_label(label: str):
    for u in flat_run_units():
        if u["label"] == label or u.get("log_id") == label:
            return u
    return None


def flat_ui_labels() -> list:
    return [u["label"] for u in flat_run_units()]


def is_allowed_project_label(label: str) -> bool:
    return any(u["label"] == label or u.get("log_id") == label for u in flat_run_units())


def _site_index_by_project_label(label: str) -> Optional[int]:
    """Index into config/sites.json sites[] for this dashboard project label (or log_id)."""
    d = _load_sites_file_app()
    if not isinstance(d, dict):
        return None
    sites = d.get("sites")
    if not isinstance(sites, list):
        return None
    u = _unit_by_label(label)
    if not u:
        return None
    e = u.get("env") or {}
    sid = str(e.get("PINTEREST_SITE_ID", "")).strip()
    for i, s in enumerate(sites):
        if not isinstance(s, dict):
            continue
        if sid and str(s.get("id", "")).strip() == sid:
            return i
        stitle = _site_row_public_title(s)
        if stitle == label or str(s.get("id", "")).strip() == label or str(s.get("log_id", "")).strip() == label:
            return i
    return None


SITES_FILE_PATH = os.path.join(_APP_CONFIG, "sites.json")

# Keys saved as plain strings in the form (optional empty = drop or keep secret, see _site_from_form)
_SITE_FORM_STRING_KEYS = (
    "id",
    "display_name",
    "out_dir",
    "start_file",
    "templates_dir",
    "log_id",
    "prompts_dir",
    "openai_api_key",
    "openai_model",
    "useapi_token",
    "useapi_midjourney_channel",
    "r2_account_id",
    "r2_access_key_id",
    "r2_secret_access_key",
    "r2_bucket",
    "r2_public_base_url",
    "wordpress_url",
    "wordpress_user",
    "wordpress_app_password",
    "theme_slug",
    "cloudflare_api_token",
    "cloudflare_account_id",
    "cloudflare_project_name",
)

# If user leaves these blank in the form, keep previous file values (same as password UX)
_SITE_SECRET_REUSE_KEYS = (
    "openai_api_key",
    "useapi_token",
    "useapi_midjourney_channel",
    "r2_access_key_id",
    "r2_secret_access_key",
    "r2_account_id",
    "r2_bucket",
    "r2_public_base_url",
    "wordpress_app_password",
    "cloudflare_api_token",
    "cloudflare_account_id",
)


def _normalize_sites_doc(doc: dict) -> dict:
    """Fix common sites.json shape issues before write (e.g. category_id_mapping as JSON string)."""
    if not isinstance(doc, dict):
        return doc
    sites = doc.get("sites")
    if not isinstance(sites, list):
        return doc
    folder = str(doc.get("pipeline_code_folder", "") or "").strip() or "A1-Pinterest_01"
    mod = _a1_config_module_for_pipeline_folder(folder)
    norm_fn = getattr(mod, "normalize_category_id_mapping", None) if mod else None
    if not norm_fn:
        return doc
    for s in sites:
        if not isinstance(s, dict):
            continue
        st = s.get("settings")
        if not isinstance(st, dict):
            continue
        parsed = norm_fn(st.get("category_id_mapping"))
        if parsed:
            st["category_id_mapping"] = parsed
    return doc


def _write_sites_doc(doc: dict) -> None:
    doc = _normalize_sites_doc(doc)
    os.makedirs(os.path.dirname(SITES_FILE_PATH) or ".", exist_ok=True)
    txt = json.dumps(doc, ensure_ascii=False, indent=2) + "\n"
    tmp = SITES_FILE_PATH + ".tmp"
    with open(tmp, "w", encoding="utf-8") as f:
        f.write(txt)
    os.replace(tmp, SITES_FILE_PATH)


def _ensure_project_output_files_for_sites(doc: Optional[dict] = None) -> None:
    """
    Ensure each site has ALL/<out_dir>/ with START.xlsx (or site start_file)
    and Recipes.xlsx.
    """
    d = doc if isinstance(doc, dict) else _load_sites_file_app()
    sites = d.get("sites") if isinstance(d, dict) else None
    if not isinstance(sites, list):
        return

    pipeline_folder = ""
    if isinstance(d, dict):
        pipeline_folder = str(d.get("pipeline_code_folder", "") or "").strip() or "A1-Pinterest_01"
    pipeline_root = os.path.join(_APP_ROOT, pipeline_folder)

    for s in sites:
        if not isinstance(s, dict):
            continue
        sid = str(s.get("id", "") or "").strip()
        out_dir = str(s.get("out_dir", "") or "").strip() or (f"{sid}-out" if sid else "")
        if not out_dir:
            continue
        target_dir = os.path.join(_APP_ROOT, "ALL", out_dir)
        os.makedirs(target_dir, exist_ok=True)

        start_name = str(s.get("start_file", "") or "").strip() or "START.xlsx"
        start_name = os.path.basename(start_name) or "START.xlsx"
        start_path = os.path.join(target_dir, start_name)
        recipes_path = os.path.join(target_dir, "Recipes.xlsx")

        if not os.path.isfile(start_path):
            wb = openpyxl.Workbook()
            try:
                sh = wb.active
                sh.title = "Titles"
                sh.cell(row=1, column=1, value="Title")
                wb.save(start_path)
            finally:
                wb.close()

        if not os.path.isfile(recipes_path):
            wb = openpyxl.Workbook()
            try:
                sh = wb.active
                sh.title = "Recipes"
                sh.cell(row=1, column=1, value="Title")
                sh.cell(row=1, column=2, value="Recipe")
                sh.cell(row=1, column=3, value="Generated At")
                wb.save(recipes_path)
            finally:
                wb.close()

        templates_sub = str(s.get("templates_dir", "") or "").strip() or "templates-html"
        templates_sub = os.path.basename(templates_sub) or "templates-html"
        target_templates_dir = os.path.join(target_dir, templates_sub)
        os.makedirs(target_templates_dir, exist_ok=True)

        # Source folder name matches the site's templates_dir setting first
        # (so projects using "templates-html" get the html templates, and
        # projects using "templates" get the legacy ones). Falls back to the
        # other if the primary doesn't exist on the source side.
        source_candidates = [templates_sub]
        if templates_sub not in ("templates", "templates-html"):
            source_candidates += ["templates-html", "templates"]
        elif templates_sub == "templates-html":
            source_candidates += ["templates"]
        else:
            source_candidates += ["templates-html"]

        template_source_dir = ""
        for cand in source_candidates:
            p = os.path.join(pipeline_root, cand)
            if os.path.isdir(p):
                template_source_dir = p
                break

        if template_source_dir and os.path.isdir(template_source_dir):
            for nm in os.listdir(template_source_dir):
                src = os.path.join(template_source_dir, nm)
                dst = os.path.join(target_templates_dir, nm)
                if os.path.isdir(src):
                    if not os.path.exists(dst):
                        try:
                            shutil.copytree(src, dst)
                        except OSError:
                            pass
                elif os.path.isfile(src) and not os.path.exists(dst):
                    try:
                        shutil.copy2(src, dst)
                    except OSError:
                        pass

        # A.6b HTML pins: each project keeps its own ALL/<out>/templates-html/
        html_target = os.path.join(target_dir, "templates-html")
        os.makedirs(html_target, exist_ok=True)
        html_src = os.path.join(pipeline_root, "templates-html")
        if os.path.isdir(html_src):
            for nm in os.listdir(html_src):
                if not nm.lower().endswith(".html"):
                    continue
                src = os.path.join(html_src, nm)
                dst = os.path.join(html_target, nm)
                if os.path.isfile(src) and not os.path.exists(dst):
                    try:
                        shutil.copy2(src, dst)
                    except OSError:
                        pass


def _next_default_site_id(sites: list) -> str:
    nmax = 0
    for s in sites:
        if not isinstance(s, dict):
            continue
        sid = str(s.get("id", "")).strip()
        m = re.match(r"^p?(\d+)$", sid, re.I)
        if m:
            nmax = max(nmax, int(m.group(1)))
    for k in range(nmax + 1, 999):
        candidate = f"p{k:02d}"
        if not any(isinstance(s, dict) and str(s.get("id", "")).strip() == candidate for s in sites):
            return candidate
    return "p99"


def _site_from_form(form, i: int, old: Optional[dict]) -> dict:
    old = old or {}
    p = f"site_{i}_"
    s: dict = {}
    for k in _SITE_FORM_STRING_KEYS:
        v = (form.get(p + k) or "").strip()
        if v:
            s[k] = v
    if form.get(p + "no_shared_settings") == "1":
        s["no_shared_settings"] = True
    if form.get(p + "no_shared_prompts") == "1":
        s["no_shared_prompts"] = True
    settings_field_prefix = p + "setting__"
    has_settings_fields = False
    settings_from_fields: dict = {}

    def _coerce_field_value(raw: str):
        txt = str(raw or "").strip()
        if txt == "":
            return None
        if (txt.startswith("{") and txt.endswith("}")) or (txt.startswith("[") and txt.endswith("]")):
            try:
                return json.loads(txt)
            except Exception:
                return txt
        lk = txt.lower()
        if lk == "true":
            return True
        if lk == "false":
            return False
        if re.fullmatch(r"-?\d+", txt):
            try:
                return int(txt)
            except Exception:
                return txt
        if re.fullmatch(r"-?\d+\.\d+", txt):
            try:
                return float(txt)
            except Exception:
                return txt
        return txt

    def _set_nested_value(dst: dict, path_parts: list, value):
        cur = dst
        for part in path_parts[:-1]:
            node = cur.get(part)
            if not isinstance(node, dict):
                node = {}
                cur[part] = node
            cur = node
        cur[path_parts[-1]] = value

    for fk in form.keys():
        if not str(fk).startswith(settings_field_prefix):
            continue
        has_settings_fields = True
        payload = str(fk)[len(settings_field_prefix):]
        bits = [b for b in payload.split("__") if b]
        if not bits:
            continue
        vv = _coerce_field_value(form.get(fk))
        if vv is None:
            continue
        _set_nested_value(settings_from_fields, bits, vv)

    t = (form.get(p + "settings_json") or "").strip()
    if has_settings_fields:
        if settings_from_fields:
            s["settings"] = settings_from_fields
    elif t:
        try:
            s["settings"] = json.loads(t)
        except json.JSONDecodeError as e:
            raise ValueError(f"Site {i + 1} settings JSON: {e}") from e
    else:
        if "settings" in old:
            s["settings"] = old["settings"]
    t2 = (form.get(p + "prompts_json") or "").strip()
    prompt_field_prefix = p + "prompt__"
    has_prompt_fields = False
    prompts_from_fields: dict = {}

    def _coerce_prompt_value(raw: str):
        txt = str(raw or "").strip()
        if txt == "":
            return None
        if (txt.startswith("{") and txt.endswith("}")) or (txt.startswith("[") and txt.endswith("]")):
            try:
                return json.loads(txt)
            except Exception:
                return txt
        lk = txt.lower()
        if lk == "true":
            return True
        if lk == "false":
            return False
        if re.fullmatch(r"-?\d+", txt):
            try:
                return int(txt)
            except Exception:
                return txt
        if re.fullmatch(r"-?\d+\.\d+", txt):
            try:
                return float(txt)
            except Exception:
                return txt
        return txt

    def _set_nested_prompt(dst: dict, path_parts: list, value):
        cur = dst
        for part in path_parts[:-1]:
            node = cur.get(part)
            if not isinstance(node, dict):
                node = {}
                cur[part] = node
            cur = node
        cur[path_parts[-1]] = value

    for fk in form.keys():
        if not str(fk).startswith(prompt_field_prefix):
            continue
        has_prompt_fields = True
        payload = str(fk)[len(prompt_field_prefix):]
        bits = [b for b in payload.split("__") if b]
        if len(bits) < 2:
            continue
        prompt_name = bits[0]
        path_parts = bits[1:]
        vv = _coerce_prompt_value(form.get(fk))
        if vv is None:
            continue
        root = prompts_from_fields.get(prompt_name)
        if not isinstance(root, dict):
            root = {}
            prompts_from_fields[prompt_name] = root
        _set_nested_prompt(root, path_parts, vv)

    if has_prompt_fields:
        if prompts_from_fields:
            s["prompts"] = prompts_from_fields
        # If prompt fields are present but all empty, treat as clearing inline prompts.
    elif t2:
        try:
            s["prompts"] = json.loads(t2)
        except json.JSONDecodeError as e:
            raise ValueError(f"Site {i + 1} prompts JSON: {e}") from e
    else:
        if "prompts" in old:
            s["prompts"] = old["prompts"]
    for k in _SITE_SECRET_REUSE_KEYS:
        if s.get(k):
            continue
        if k in old and (old.get(k) is not None) and str(old.get(k) or "").strip() != "":
            s[k] = old[k]
    return s


def _default_new_site(sites: list) -> dict:
    sid = _next_default_site_id(sites)
    return {
        "id": sid,
        "display_name": f"New site ({sid})",
        "out_dir": f"{sid}-out",
        "templates_dir": "templates",
        "wordpress_url": "https://",
        "wordpress_user": "",
        "wordpress_app_password": "",
    }


_A1_CONFIG_MODULES: dict = {}
_site_config_lock = threading.Lock()


def _a1_config_module_for_pipeline_folder(folder: str):
    p = os.path.join(_APP_ROOT, folder, "a1_config.py")
    if not os.path.isfile(p):
        return None
    if folder in _A1_CONFIG_MODULES:
        return _A1_CONFIG_MODULES[folder]
    mname = "a1_config__" + re.sub(r"[^a-zA-Z0-9_]+", "_", folder)
    spec = importlib.util.spec_from_file_location(mname, p)
    if spec is None or spec.loader is None:
        return None
    mod = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(mod)  # type: ignore[union-attr]
    _A1_CONFIG_MODULES[folder] = mod
    return mod


def load_project_folders() -> list:
    """
    - If config/projects.json has non-empty "folders", use that order (only existing dirs).
    - Otherwise every direct subfolder that contains A.1-START.py, sorted naturally.
    """
    pc = _load_app_projects_file()
    explicit = pc.get("folders")
    if isinstance(explicit, list) and len(explicit) > 0:
        return [f for f in explicit if f and os.path.isdir(os.path.join(_APP_ROOT, f))]
    names: list = []
    for name in os.listdir(_APP_ROOT):
        if name in _SKIP_PROJECT_PARENTS or name.startswith("."):
            continue
        p = os.path.join(_APP_ROOT, name)
        if not os.path.isdir(p):
            continue
        if not _is_pipeline_project(name):
            continue
        names.append(name)
    names.sort(key=_natural_sort_key)
    return names


def _split_folder_groups(folders: list, n: int) -> list:
    """Split `folders` into n consecutive index ranges (n UI slots); sizes differ by at most 1."""
    L = len(folders)
    if n <= 0:
        return []
    if L == 0:
        return [[] for _ in range(n)]
    base, rem = divmod(L, n)
    out, start = [], 0
    for i in range(n):
        sz = base + (1 if i < rem else 0)
        out.append(folders[start : start + sz])
        start += sz
    return out


def _split_unit_groups(units: list, n: int) -> list:
    """Like _split_folder_groups but each chunk is a list of run-unit dicts."""
    L = len(units)
    if n <= 0:
        return []
    if L == 0:
        return [[] for _ in range(n)]
    base, rem = divmod(L, n)
    out, start = [], 0
    for i in range(n):
        sz = base + (1 if i < rem else 0)
        out.append(units[start : start + sz])
        start += sz
    return out


PROJECT_FOLDERS = load_project_folders()
if not PROJECT_FOLDERS and os.path.isdir(os.path.join(_APP_ROOT, "A1-Pinterest_01")):
    PROJECT_FOLDERS = ["A1-Pinterest_01"]

_PC = _load_app_projects_file()
_s2 = _PC.get("start2_folders")
if isinstance(_s2, list):
    START2_PROJECTS = [x for x in _s2 if isinstance(x, str) and x]
else:
    START2_PROJECTS = []


def _unit_in_start2_list(u: dict) -> bool:
    s2 = set(START2_PROJECTS)
    if not s2:
        return False
    if u.get("label") in s2 or u.get("log_id") in s2:
        return True
    sid = (u.get("env") or {}).get("PINTEREST_SITE_ID", "")
    return sid in s2


def jobs_for_start1_all_except_s2() -> list:
    o = []
    for u in flat_run_units():
        if _unit_in_start2_list(u):
            continue
        o.append(
            (u["folder"], "A.1-START.py", u.get("env") or {}, u["log_id"], u["label"])
        )
    return o


def _read_titles_from_start_workbook(path: str) -> list:
    titles: list = []
    try:
        wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    except Exception:
        return titles
    try:
        sh = wb.active
        title_col = 1
        used_col = 0
        max_col = max(1, int(sh.max_column or 1))
        for c in range(1, max_col + 1):
            hv = sh.cell(row=1, column=c).value
            if hv is None:
                continue
            hk = str(hv).strip().lower()
            if hk == "title":
                title_col = c
                break
        for c in range(1, max_col + 1):
            hv = sh.cell(row=1, column=c).value
            if hv is None:
                continue
            if str(hv).strip().lower() == "used":
                used_col = c
                break

        def _is_used_true(v) -> bool:
            if v is None:
                return False
            if isinstance(v, bool):
                return v is True
            if isinstance(v, (int, float)):
                return int(v) == 1
            s = str(v).strip().lower()
            if not s:
                return False
            return s in {"1", "true", "yes", "y", "used"}

        max_row = int(sh.max_row or 1)
        for r in range(2, max_row + 1):
            if used_col > 0 and _is_used_true(sh.cell(row=r, column=used_col).value):
                continue
            v = sh.cell(row=r, column=title_col).value
            if v is None:
                continue
            t = str(v).strip()
            if t:
                titles.append({"title": t, "source_row": r})
    finally:
        wb.close()
    return titles


def _write_start_usage_columns(starts_dir: str, xlsx_files: list, used_map: dict, stamp: str) -> None:
    """
    Update each STARTS workbook with run usage metadata:
    - used: 1 for assigned, 0 otherwise
    - used_at: timestamp for assigned rows
    - used_project: project label for assigned rows
    """
    for fn in xlsx_files:
        fp = os.path.join(starts_dir, fn)
        try:
            wb = openpyxl.load_workbook(fp)
        except Exception:
            continue
        try:
            sh = wb.active
            max_col = max(1, int(sh.max_column or 1))
            max_row = int(sh.max_row or 1)

            header_to_col = {}
            for c in range(1, max_col + 1):
                hv = sh.cell(row=1, column=c).value
                if hv is None:
                    continue
                header_to_col[str(hv).strip().lower()] = c

            def ensure_col(name: str) -> int:
                lk = name.strip().lower()
                c0 = header_to_col.get(lk)
                if c0:
                    return c0
                new_c = int(sh.max_column or 1) + 1
                sh.cell(row=1, column=new_c, value=name)
                header_to_col[lk] = new_c
                return new_c

            used_col = ensure_col("used")
            used_at_col = ensure_col("used_at")
            used_project_col = ensure_col("used_project")

            by_row = used_map.get(fn, {})
            for r in range(2, max_row + 1):
                rec = by_row.get(r)
                if rec:
                    sh.cell(row=r, column=used_col, value=1)
                    sh.cell(row=r, column=used_at_col, value=stamp)
                    sh.cell(row=r, column=used_project_col, value=rec.get("project", ""))
                else:
                    sh.cell(row=r, column=used_col, value=0)
                    sh.cell(row=r, column=used_at_col, value="")
                    sh.cell(row=r, column=used_project_col, value="")

            wb.save(fp)
        finally:
            wb.close()


def _build_global_start_runtime_jobs(base_jobs: list) -> tuple[list, dict]:
    """
    Build temp START files for this click only (no sites.json change):
    - Reads all titles from STARTS/*.xlsx
    - Distributes titles evenly over current jobs
    - Writes per-job runtime .xlsx and passes override through env
    - Writes one usage report with Used=true/false
    """
    jobs = list(base_jobs or [])
    starts_dir = os.path.join(_APP_ROOT, "STARTS")
    if not jobs:
        return jobs, {"mode": "no_jobs", "titles_total": 0}
    if not os.path.isdir(starts_dir):
        return jobs, {"mode": "missing_starts_folder", "titles_total": 0}

    xlsx_files = [
        f
        for f in os.listdir(starts_dir)
        if f.lower().endswith(".xlsx")
        and not f.startswith("~$")
        and not f.startswith("._")
        and not f.startswith("_")
        and os.path.isfile(os.path.join(starts_dir, f))
    ]
    xlsx_files.sort(key=_natural_sort_key)
    if not xlsx_files:
        return jobs, {"mode": "no_xlsx_files", "titles_total": 0}

    title_rows: list = []
    for fn in xlsx_files:
        fp = os.path.join(starts_dir, fn)
        for rec in _read_titles_from_start_workbook(fp):
            title_rows.append(
                {
                    "title": rec["title"],
                    "source_file": fn,
                    "source_row": int(rec.get("source_row", 0) or 0),
                }
            )

    if not title_rows:
        return jobs, {"mode": "no_titles_found", "titles_total": 0}

    n_jobs = len(jobs)
    chunks = [[] for _ in range(n_jobs)]
    start_idx = int(_GLOBAL_START_ROTATION.get("cursor", 0) or 0) % max(1, n_jobs)
    for i, row in enumerate(title_rows):
        chunks[(start_idx + i) % n_jobs].append(row)
    _GLOBAL_START_ROTATION["cursor"] = (start_idx + 1) % max(1, n_jobs)

    runtime_root = os.path.join(starts_dir, "_runtime_global_start")
    os.makedirs(runtime_root, exist_ok=True)
    for name in os.listdir(runtime_root):
        p = os.path.join(runtime_root, name)
        if os.path.isfile(p):
            try:
                os.remove(p)
            except OSError:
                pass

    patched_jobs = []
    usage_rows = []
    per_project_counts = {}
    for i, job in enumerate(jobs):
        folder, script, env, log_id, line_label = job
        rows = chunks[i]
        wb = openpyxl.Workbook()
        sh = wb.active
        sh.title = "Titles"
        sh.cell(row=1, column=1, value="Title")
        sh.cell(row=1, column=2, value="source_file")
        sh.cell(row=1, column=3, value="source_row")
        for r, item in enumerate(rows, start=2):
            sh.cell(row=r, column=1, value=item["title"])
            sh.cell(row=r, column=2, value=item.get("source_file", ""))
            sh.cell(row=r, column=3, value=int(item.get("source_row", 0) or 0))
            usage_rows.append(
                {
                    "Title": item["title"],
                    "Source File": item["source_file"],
                    "Assigned Project": line_label,
                    "Used": True,
                }
            )
        out_name = f"{i+1:04d}_{_safe_log_dom_id(str(line_label))}.xlsx"
        out_path = os.path.join(runtime_root, out_name)
        wb.save(out_path)
        wb.close()
        env2 = dict(env or {})
        env2["PINTEREST_START_FILE_OVERRIDE"] = out_path
        patched_jobs.append((folder, script, env2, log_id, line_label))
        per_project_counts[str(line_label)] = len(rows)

    used_titles = set()
    for c in chunks:
        for item in c:
            used_titles.add((item["title"], item["source_file"]))
    for item in title_rows:
        k = (item["title"], item["source_file"])
        if k in used_titles:
            continue
        usage_rows.append(
            {
                "Title": item["title"],
                "Source File": item["source_file"],
                "Assigned Project": "",
                "Used": False,
            }
        )

    usage_path = os.path.join(runtime_root, "_global_usage.xlsx")
    wb_u = openpyxl.Workbook()
    sh_u = wb_u.active
    sh_u.title = "Usage"
    headers = ["Title", "Source File", "Assigned Project", "Used"]
    for c, h in enumerate(headers, start=1):
        sh_u.cell(row=1, column=c, value=h)
    for r, row in enumerate(usage_rows, start=2):
        sh_u.cell(row=r, column=1, value=row["Title"])
        sh_u.cell(row=r, column=2, value=row["Source File"])
        sh_u.cell(row=r, column=3, value=row["Assigned Project"])
        sh_u.cell(row=r, column=4, value=row["Used"])
    wb_u.save(usage_path)
    wb_u.close()

    return patched_jobs, {
        "mode": "allocated",
        "titles_total": len(title_rows),
        "projects_total": n_jobs,
        "rotation_start_index": int(start_idx),
        "runtime_dir": runtime_root,
        "usage_file": usage_path,
        "per_project_counts": per_project_counts,
    }


def jobs_for_start2_only() -> list:
    return [
        (u["folder"], "A.1-START.py", u.get("env") or {}, u["log_id"], u["label"])
        for u in flat_run_units()
        if _unit_in_start2_list(u)
    ]


FLAT_UNITS = flat_run_units()
IMAGINE_NUM_GROUPS = int(_PC.get("imagine_num_groups", 17))
IMAGINE_SLOT_GROUPS = _split_unit_groups(FLAT_UNITS, IMAGINE_NUM_GROUPS)
# stream-imagine-group1 … = slots of sites (or folders) in order
IMAGINE_GROUP1 = IMAGINE_SLOT_GROUPS[0] if len(IMAGINE_SLOT_GROUPS) > 0 else []
IMAGINE_GROUP2 = IMAGINE_SLOT_GROUPS[1] if len(IMAGINE_SLOT_GROUPS) > 1 else []
IMAGINE_GROUP3 = IMAGINE_SLOT_GROUPS[2] if len(IMAGINE_SLOT_GROUPS) > 2 else []
IMAGINE_GROUP4 = IMAGINE_SLOT_GROUPS[3] if len(IMAGINE_SLOT_GROUPS) > 3 else []
IMAGINE_GROUP5 = IMAGINE_SLOT_GROUPS[4] if len(IMAGINE_SLOT_GROUPS) > 4 else []
IMAGINE_GROUP6 = IMAGINE_SLOT_GROUPS[5] if len(IMAGINE_SLOT_GROUPS) > 5 else []
IMAGINE_GROUP7 = IMAGINE_SLOT_GROUPS[6] if len(IMAGINE_SLOT_GROUPS) > 6 else []
IMAGINE_GROUP8 = IMAGINE_SLOT_GROUPS[7] if len(IMAGINE_SLOT_GROUPS) > 7 else []
IMAGINE_GROUP9 = IMAGINE_SLOT_GROUPS[8] if len(IMAGINE_SLOT_GROUPS) > 8 else []
IMAGINE_GROUP10 = IMAGINE_SLOT_GROUPS[9] if len(IMAGINE_SLOT_GROUPS) > 9 else []
IMAGINE_GROUP11 = IMAGINE_SLOT_GROUPS[10] if len(IMAGINE_SLOT_GROUPS) > 10 else []
IMAGINE_GROUP12 = IMAGINE_SLOT_GROUPS[11] if len(IMAGINE_SLOT_GROUPS) > 11 else []
IMAGINE_GROUP13 = IMAGINE_SLOT_GROUPS[12] if len(IMAGINE_SLOT_GROUPS) > 12 else []
IMAGINE_GROUP14 = IMAGINE_SLOT_GROUPS[13] if len(IMAGINE_SLOT_GROUPS) > 13 else []
IMAGINE_GROUP15 = IMAGINE_SLOT_GROUPS[14] if len(IMAGINE_SLOT_GROUPS) > 14 else []
IMAGINE_GROUP16 = IMAGINE_SLOT_GROUPS[15] if len(IMAGINE_SLOT_GROUPS) > 15 else []
IMAGINE_GROUP17 = IMAGINE_SLOT_GROUPS[16] if len(IMAGINE_SLOT_GROUPS) > 16 else []
_tuples = _PC.get("imagine_all_ranges")
if isinstance(_tuples, list) and _tuples:
    IMAGINE_ALL_RANGES = [tuple(x) for x in _tuples]  # type: ignore
else:
    _n = len(FLAT_UNITS)
    if _n <= 0:
        IMAGINE_ALL_RANGES = []
    elif _n == 1:
        IMAGINE_ALL_RANGES = [(1, 1)]
    else:
        _mid = (_n + 1) // 2
        IMAGINE_ALL_RANGES = [(1, _mid), (_mid + 1, _n)]

# -------------------- Global list to track running subprocesses --------------------
running_processes = []


# -------------------- 1) Concurrency / SSE Logging --------------------
def generate_log_parallel(script_names, env_extra=None):
    """
    Runs scripts concurrently in separate threads and yields SSE log lines.
    (No concurrency cap)
    script_names: list of 2- to 5-tuples; final form (folder, script, env, log_id, line_label).
    """
    q = queue.Queue()
    threads = []
    base_env = _subprocess_env(env_extra)
    jobs = _normalize_script_jobs(script_names)

    def run_process(folder, script, job_env, log_id, line_label):
        folder_abs = os.path.join(os.getcwd(), folder)
        script_path = os.path.join(folder_abs, script)
        if not os.path.exists(script_path):
            q.put(
                {
                    "folder": log_id,
                    "line": f"Script {line_label}/{script} not found.",
                }
            )
            return
        be = _job_subprocess_env(base_env, job_env)
        q.put({"folder": log_id, "line": f"Running {line_label}/{script}..."})
        proc = _popen_pipeline_script(folder_abs, script, be)

        running_processes.append(proc)
        for line in proc.stdout:
            q.put({"folder": log_id, "line": line.rstrip()})
        proc.stdout.close()
        proc.wait()
        try:
            running_processes.remove(proc)
        except ValueError:
            pass
        q.put({"folder": log_id, "line": f"Finished {line_label}/{script}"})

    for folder, script, job_env, log_id, line_label in jobs:
        t = threading.Thread(
            target=run_process, args=(folder, script, job_env, log_id, line_label)
        )
        t.start()
        threads.append(t)

    while any(t.is_alive() for t in threads) or not q.empty():
        try:
            msg = q.get(timeout=0.1)
            yield f"data: {json.dumps(msg)}\n\n"
        except queue.Empty:
            pass

    yield "data: " + json.dumps({"folder": "all", "line": "Finished all processes."}) + "\n\n"


def generate_log_sequential(script_names, env_extra=None, delay_between_jobs_s: float = 0.0):
    """
    Runs scripts one by one (strict sequence) and streams SSE log lines.
    Optional delay_between_jobs_s is applied between finished job N and starting job N+1.
    """
    jobs = _normalize_script_jobs(script_names)
    base_env = _subprocess_env(env_extra)

    for idx, (folder, script, job_env, log_id, line_label) in enumerate(jobs, start=1):
        folder_abs = os.path.join(os.getcwd(), folder)
        script_path = os.path.join(folder_abs, script)
        if not os.path.exists(script_path):
            yield "data: " + json.dumps(
                {"folder": log_id, "line": f"Script {line_label}/{script} not found (SKIPPED)."}
            ) + "\n\n"
            continue

        be = _job_subprocess_env(base_env, job_env)
        yield "data: " + json.dumps(
            {"folder": log_id, "line": f"Running {line_label}/{script}... [{idx}/{len(jobs)}]"}
        ) + "\n\n"
        proc = _popen_pipeline_script(folder_abs, script, be)
        running_processes.append(proc)
        for line in proc.stdout:
            yield "data: " + json.dumps({"folder": log_id, "line": line.rstrip()}) + "\n\n"
        proc.stdout.close()
        proc.wait()
        try:
            running_processes.remove(proc)
        except ValueError:
            pass
        yield "data: " + json.dumps(
            {"folder": log_id, "line": f"Finished {line_label}/{script} (exit={proc.returncode})."}
        ) + "\n\n"

        if delay_between_jobs_s > 0 and idx < len(jobs):
            wait_s = int(max(0, delay_between_jobs_s))
            yield "data: " + json.dumps(
                {
                    "folder": log_id,
                    "line": f"Safety wait: sleeping {wait_s}s before next job.",
                }
            ) + "\n\n"
            time.sleep(wait_s)

    yield "data: " + json.dumps({"folder": "all", "line": "Finished all processes."}) + "\n\n"

def generate_log_imagine_all_grouped(env_extra=None):
    """
    Run A.3-IMAGINE.py for ALL projects in GROUPS:
    - المجموعات كيتحددو ف IMAGINE_ALL_RANGES (1-based indices)
    - كل مجموعة كتخدم ف Thread بوحدها (Parallel)
    - داخل كل مجموعة المشاريع كيتخدمو واحد مور واحد (Sequential)
    """
    q = queue.Queue()
    base_env = _subprocess_env(env_extra)

    # نحضرو اللوائح ديال الفولدرات لكل مجموعة
    groups = []
    for i, (start_idx_1, end_idx_1) in enumerate(IMAGINE_ALL_RANGES, start=1):
        start_idx_1 = max(1, start_idx_1)
        end_idx_1 = min(len(FLAT_UNITS), end_idx_1)
        if start_idx_1 > end_idx_1:
            continue
        start0 = start_idx_1 - 1
        end0 = end_idx_1
        uchunk = FLAT_UNITS[start0:end0]
        if not uchunk:
            continue
        label = f"Group {i} ({start_idx_1}-{end_idx_1})"
        groups.append((label, uchunk))

    def run_group(label, units):
        for unit in units:
            script = "A.3-IMAGINE.py"
            folder = unit["folder"]
            job_env = _job_subprocess_env(base_env, unit.get("env") or {})
            line_label = unit.get("label", folder)
            log_id = unit.get("log_id", _safe_log_dom_id(line_label))
            folder_abs = os.path.join(os.getcwd(), folder)
            script_path = os.path.join(folder_abs, script)
            if not os.path.exists(script_path):
                q.put(
                    {
                        "folder": log_id,
                        "line": f"[{label}] Script {line_label}/{script} not found.",
                    }
                )
                continue

            q.put({"folder": log_id, "line": f"[{label}] Running {line_label}/{script}..."})
            proc = _popen_pipeline_script(folder_abs, script, job_env)
            running_processes.append(proc)
            for line in proc.stdout:
                q.put({"folder": log_id, "line": f"[{label}] " + line.rstrip()})
            proc.stdout.close()
            proc.wait()
            try:
                running_processes.remove(proc)
            except ValueError:
                pass
            q.put(
                {
                    "folder": log_id,
                    "line": f"[{label}] Finished {line_label}/{script} (exit={proc.returncode}).",
                }
            )

        q.put({"folder": "all", "line": f"[{label}] Finished all projects in this group."})

    threads = []
    for label, folders in groups:
        t = threading.Thread(target=run_group, args=(label, folders), daemon=True)
        t.start()
        threads.append(t)

    # نخرجو اللوغات حتى يساليو جميع الثريدات
    while any(t.is_alive() for t in threads) or not q.empty():
        try:
            msg = q.get(timeout=0.1)
            yield "data: " + json.dumps(msg) + "\n\n"
        except queue.Empty:
            pass

    yield "data: " + json.dumps({"folder": "all", "line": "Finished all IMAGINE ALL groups."}) + "\n\n"


# -------------------- 2) Pool Runner (Fixed-size concurrency) --------------------
def generate_log_pool(script_names, max_concurrency=10, env_extra=None):
    """
    Fixed-size pool: keeps up to max_concurrency running; if one finishes or is SKIPPED,
    immediately starts the next.
    """
    q = queue.Queue()
    lock = threading.Lock()
    job_list = _normalize_script_jobs(script_names)
    scripts_iter = iter(job_list)
    active_threads = []
    base_env = _subprocess_env(env_extra)

    def run_process(folder, script, job_env, log_id, line_label):
        folder_abs = os.path.join(os.getcwd(), folder)
        script_path = os.path.join(folder_abs, script)
        if not os.path.exists(script_path):
            q.put(
                {
                    "folder": log_id,
                    "line": f"Script {line_label}/{script} not found (SKIPPED).",
                }
            )
        else:
            be = _job_subprocess_env(base_env, job_env)
            q.put({"folder": log_id, "line": f"Running {line_label}/{script}..."})
            proc = _popen_pipeline_script(folder_abs, script, be)

            running_processes.append(proc)
            for line in proc.stdout:
                q.put({"folder": log_id, "line": line.rstrip()})
            proc.stdout.close()
            proc.wait()
            try:
                running_processes.remove(proc)
            except ValueError:
                pass
            q.put({"folder": log_id, "line": f"Finished {line_label}/{script}"})

        # refill slot
        with lock:
            try:
                nxt = next(scripts_iter)
                t = threading.Thread(
                    target=run_process,
                    args=(nxt[0], nxt[1], nxt[2], nxt[3], nxt[4]),
                )
                t.start()
                active_threads.append(t)
            except StopIteration:
                pass

    # prime pool
    for _ in range(min(max_concurrency, len(job_list))):
        try:
            f, s, e, lid, llab = next(scripts_iter)
            t = threading.Thread(target=run_process, args=(f, s, e, lid, llab))
            t.start()
            active_threads.append(t)
        except StopIteration:
            break

    # stream logs
    while any(t.is_alive() for t in active_threads) or not q.empty():
        try:
            msg = q.get(timeout=0.2)
            yield f"data: {json.dumps(msg)}\n\n"
        except queue.Empty:
            pass

    yield "data: " + json.dumps({"folder": "all", "line": "Finished all processes."}) + "\n\n"


# -------------------- 2b) BATCH Runner (Strict batches: wait for N to finish, then next N) --------------------
def generate_log_in_batches(script_names, batch_size=3, env_extra=None):
    """
    Runs scripts in strict batches of size `batch_size`.
    - Start at most `batch_size` processes.
    - Wait until ALL of them finish.
    - Then start the next batch.
    Streams all stdout/stderr lines via SSE.
    """
    q = queue.Queue()
    base_env = _subprocess_env(env_extra)
    job_list = _normalize_script_jobs(script_names)

    def run_process(folder, script, job_env, log_id, line_label):
        folder_abs = os.path.join(os.getcwd(), folder)
        script_path = os.path.join(folder_abs, script)
        if not os.path.exists(script_path):
            q.put(
                {
                    "folder": log_id,
                    "line": f"Script {line_label}/{script} not found.",
                }
            )
            return

        be = _job_subprocess_env(base_env, job_env)
        q.put({"folder": log_id, "line": f"Running {line_label}/{script}..."})
        proc = _popen_pipeline_script(folder_abs, script, be)
        # stream output
        for line in proc.stdout:
            q.put({"folder": log_id, "line": line.rstrip()})
        proc.stdout.close()
        proc.wait()
        q.put(
            {
                "folder": log_id,
                "line": f"Finished {line_label}/{script} (exit={proc.returncode}).",
            }
        )

    # Helper: chunk list into batches
    def chunked(seq, n):
        for i in range(0, len(seq), n):
            yield seq[i:i + n]

    threads = []
    # Process by batches
    for batch in chunked(list(job_list), batch_size):
        # start a batch
        batch_threads = []
        for f, s, e, lid, llab in batch:
            t = threading.Thread(
                target=run_process, args=(f, s, e, lid, llab), daemon=True
            )
            t.start()
            batch_threads.append(t)
            threads.append(t)

        # While this batch is running, keep yielding queue messages
        while any(t.is_alive() for t in batch_threads):
            try:
                msg = q.get(timeout=0.1)
                yield f"data: {json.dumps(msg)}\n\n"
            except queue.Empty:
                pass
        # drain remaining queue lines for this batch
        while True:
            try:
                msg = q.get_nowait()
                yield f"data: {json.dumps(msg)}\n\n"
            except queue.Empty:
                break

    # ensure all threads done
    for t in threads:
        t.join(timeout=0.1)

    yield "data: " + json.dumps({"folder": "all", "line": "Finished all processes (batched)."}) + "\n\n"


# -------------------- 3) Streaming Endpoints --------------------

@app.route("/stream-all-start")
def stream_all_start():
    raw_lim = (request.args.get("title_limit") or "").strip()
    env_extra = None
    if raw_lim:
        try:
            lim = max(0, int(raw_lim))
            if lim > 0:
                env_extra = {"PINTEREST_START_LIMIT": str(lim)}
        except ValueError:
            env_extra = None
    jobs = jobs_for_start1_all_except_s2()
    return Response(
        generate_log_parallel(jobs, env_extra=env_extra), mimetype="text/event-stream"
    )



@app.route("/stream-all-prompt")
def stream_all_prompt():
    return Response(
        generate_log_parallel(jobs_for_script("A.2-PROMPT.py")),
        mimetype="text/event-stream",
    )

@app.route("/stream-all-json")
def stream_all_json():
    scripts_to_run = _filter_jobs_missing_step(jobs_for_script("A.2-JSON.py"), "JSON")

    # باش نعطيك Done X/Y
    total = len(scripts_to_run)

    def gen():
        if total <= 0:
            yield "data: " + json.dumps({"folder": "all", "line": "JSON already complete for all projects. Nothing to run."}) + "\n\n"
            yield "data: " + json.dumps({"folder": "all", "line": "Finished all processes."}) + "\n\n"
            return
        done = 0
        for chunk in generate_log_in_batches(scripts_to_run, batch_size=5):
            # chunk = "data: {...}\n\n" أو "data: {...}\n\n"
            # كنحسبو "Finished ..." باش نعرفو شحال تسالا
            if '"line": "Finished ' in chunk:
                done += 1
            yield chunk

        yield "data: " + json.dumps({"folder": "all", "line": f"JSON Done {done}/{total}"}) + "\n\n"

    return Response(gen(), mimetype="text/event-stream")

def generate_log_sequential_with_delay(script_names, delay_s=120, env_extra=None):
    base_env = _subprocess_env(env_extra)
    jobs = _normalize_script_jobs(script_names)

    for idx, (folder, script, job_env, log_id, line_label) in enumerate(jobs, start=1):
        folder_abs = os.path.join(os.getcwd(), folder)
        script_path = os.path.join(folder_abs, script)

        if not os.path.exists(script_path):
            yield "data: " + json.dumps({
                "folder": log_id,
                "line": f"Script {line_label}/{script} not found."
            }) + "\n\n"
            continue

        yield "data: " + json.dumps({
            "folder": log_id,
            "line": f"Running {line_label}/{script}..."
        }) + "\n\n"

        be = _job_subprocess_env(base_env, job_env)
        proc = _popen_pipeline_script(folder_abs, script, be)
        running_processes.append(proc)

        for line in proc.stdout:
            yield "data: " + json.dumps({
                "folder": log_id,
                "line": line.rstrip()
            }) + "\n\n"

        proc.stdout.close()
        proc.wait()

        try:
            running_processes.remove(proc)
        except ValueError:
            pass

        yield "data: " + json.dumps({
            "folder": log_id,
            "line": f"Finished {line_label}/{script}"
        }) + "\n\n"

        if idx < len(jobs):
            next_log_id = jobs[idx][3]

            yield "data: " + json.dumps({
                "folder": next_log_id,
                "line": f"🚀 Starting script in {delay_s} seconds..."
            }) + "\n\n"

            time.sleep(delay_s)

    yield "data: " + json.dumps({
        "folder": "all",
        "line": "Finished all IMAGINE projects."
    }) + "\n\n"


@app.route("/stream-imagine-all")
def stream_imagine_all():
    return Response(
        stream_with_context(
            generate_log_sequential_with_delay(
                jobs_for_script("A.3-IMAGINE.py"),
                delay_s=120
            )
        ),
        mimetype="text/event-stream"
    )



@app.route("/stream-imagine-group1")
def stream_imagine_group1():
    return Response(
        generate_log_parallel(jobs_for_imagine_unit_group(IMAGINE_GROUP1)),
        mimetype="text/event-stream",
    )


@app.route("/stream-imagine-group2")
def stream_imagine_group2():
    return Response(
        generate_log_parallel(jobs_for_imagine_unit_group(IMAGINE_GROUP2)),
        mimetype="text/event-stream",
    )


@app.route("/stream-imagine-group3")
def stream_imagine_group3():
    return Response(
        generate_log_parallel(jobs_for_imagine_unit_group(IMAGINE_GROUP3)),
        mimetype="text/event-stream",
    )


@app.route("/stream-imagine-group4")
def stream_imagine_group4():
    return Response(
        generate_log_parallel(jobs_for_imagine_unit_group(IMAGINE_GROUP4)),
        mimetype="text/event-stream",
    )


@app.route("/stream-imagine-group5")
def stream_imagine_group5():
    return Response(
        generate_log_parallel(jobs_for_imagine_unit_group(IMAGINE_GROUP5)),
        mimetype="text/event-stream",
    )


@app.route("/stream-imagine-group6")
def stream_imagine_group6():
    return Response(
        generate_log_parallel(jobs_for_imagine_unit_group(IMAGINE_GROUP6)),
        mimetype="text/event-stream",
    )


@app.route("/stream-imagine-group7")
def stream_imagine_group7():
    return Response(
        generate_log_parallel(jobs_for_imagine_unit_group(IMAGINE_GROUP7)),
        mimetype="text/event-stream",
    )


@app.route("/stream-imagine-group8")
def stream_imagine_group8():
    return Response(
        generate_log_parallel(jobs_for_imagine_unit_group(IMAGINE_GROUP8)),
        mimetype="text/event-stream",
    )


@app.route("/stream-imagine-group9")
def stream_imagine_group9():
    return Response(
        generate_log_parallel(jobs_for_imagine_unit_group(IMAGINE_GROUP9)),
        mimetype="text/event-stream",
    )


@app.route("/stream-imagine-group10")
def stream_imagine_group10():
    return Response(
        generate_log_parallel(jobs_for_imagine_unit_group(IMAGINE_GROUP10)),
        mimetype="text/event-stream",
    )


@app.route("/stream-imagine-group11")
def stream_imagine_group11():
    return Response(
        generate_log_parallel(jobs_for_imagine_unit_group(IMAGINE_GROUP11)),
        mimetype="text/event-stream",
    )


@app.route("/stream-imagine-group12")
def stream_imagine_group12():
    return Response(
        generate_log_parallel(jobs_for_imagine_unit_group(IMAGINE_GROUP12)),
        mimetype="text/event-stream",
    )


@app.route("/stream-imagine-group13")
def stream_imagine_group13():
    return Response(
        generate_log_parallel(jobs_for_imagine_unit_group(IMAGINE_GROUP13)),
        mimetype="text/event-stream",
    )


@app.route("/stream-imagine-group14")
def stream_imagine_group14():
    return Response(
        generate_log_parallel(jobs_for_imagine_unit_group(IMAGINE_GROUP14)),
        mimetype="text/event-stream",
    )


@app.route("/stream-imagine-group15")
def stream_imagine_group15():
    return Response(
        generate_log_parallel(jobs_for_imagine_unit_group(IMAGINE_GROUP15)),
        mimetype="text/event-stream",
    )


@app.route("/stream-imagine-group16")
def stream_imagine_group16():
    return Response(
        generate_log_parallel(jobs_for_imagine_unit_group(IMAGINE_GROUP16)),
        mimetype="text/event-stream",
    )


@app.route("/stream-imagine-group17")
def stream_imagine_group17():
    return Response(
        generate_log_parallel(jobs_for_imagine_unit_group(IMAGINE_GROUP17)),
        mimetype="text/event-stream",
    )


# -------------------- UPDATED: Articles (pool 10 ب 10) --------------------
@app.route("/stream-all-article")
def stream_all_article():
    scripts_to_run = jobs_for_script("A.4-ARTICLES.py")
    return Response(
        generate_log_pool(scripts_to_run, max_concurrency=10),
        mimetype="text/event-stream"
    )


# -------------------- UPDATED: PIN DATA (pool 10 ب 10) + Skip Existing Rows flag --------------------
@app.route("/stream-all-pin-data")
def stream_all_pin_data():
    """
    كنشغل A.5-PIN DATA.py لكل مشروع، ومعاه env PIN_SKIP_EXISTING=1
    باش السكريبت ديال PIN DATA يتخطّى الصفوف العامرة وميرجعش يعاود لها.
    خاصك تزاد سنيبت فـ A.5-PIN DATA.py (أسفل) باش يفعل هاد السلوك.
    """
    scripts_to_run = jobs_for_script("A.5-PIN DATA.py")
    env_extra = {"PIN_SKIP_EXISTING": "1"}
    return Response(
        generate_log_pool(scripts_to_run, max_concurrency=10, env_extra=env_extra),
        mimetype="text/event-stream"
    )


@app.route("/stream-all-pin-image")
def stream_all_pin_image():
    BATCH_SIZE = 5
    _units = flat_run_units()
    total = len(_units)

    @stream_with_context
    def stream():
        done = 0
        yield f"data: 🚀 Starting Pin Image for {total} folders (batch={BATCH_SIZE})\n\n"

        for i in range(0, total, BATCH_SIZE):
            batch_u = _units[i : i + BATCH_SIZE]
            start = i + 1
            end = i + len(batch_u)

            yield f"data: 🚀 Batch {start}-{end} / {total}\n\n"

            scripts_to_run = [
                (u["folder"], "A.6-PIN IMAGES.py", u.get("env") or {}, u["log_id"], u["label"])
                for u in batch_u
            ]

            # run this batch
            for line in generate_log_parallel(scripts_to_run):
                # ✅ مهم: فلتر رسالة النهاية ديال generate_log_parallel باش مايتسدّش SSE وسط الخدمة
                if '"folder": "all"' in line and "Finished all processes." in line:
                    continue
                yield line

            done += len(batch_u)
            yield f"data: ✅ Done {done}/{total}\n\n"

        # ✅ رسالة نهاية واحدة فقط فالأخير
        yield "data: 🎉 ALL PIN IMAGE DONE\n\n"
        yield 'data: {"folder":"all","line":"Finished all processes."}\n\n'

    return Response(
        stream(),
        mimetype="text/event-stream",
        headers={
            "Cache-Control": "no-cache",
            "X-Accel-Buffering": "no"
        }
    )



@app.route("/stream-all-pin-image-html")
def stream_all_pin_image_html():
    """
    HTML/CSS-templated pin generation. Each project picks an .html template at random
    from ALL/<out_dir>/templates-html/ for each row in Recipes.xlsx and renders it
    with Playwright (1000x1500 JPEG). Runs all configured projects in parallel.
    """
    scripts_to_run = jobs_for_script("A.6b-PIN IMAGES HTML.py")

    @stream_with_context
    def stream():
        total = len(scripts_to_run)
        yield f"data: 🚀 Starting Pin Image (HTML) for {total} projects (all parallel)\n\n"

        for line in generate_log_parallel(scripts_to_run):
            if '"folder": "all"' in line and "Finished all processes." in line:
                continue
            yield line

        yield "data: 🎉 ALL PIN IMAGE HTML DONE\n\n"
        yield 'data: {"folder":"all","line":"Finished all processes."}\n\n'

    return Response(
        stream(),
        mimetype="text/event-stream",
        headers={
            "Cache-Control": "no-cache",
            "X-Accel-Buffering": "no",
        },
    )


@app.route("/stream-all-pin-bulk")
def stream_all_pin_bulk():
    return Response(
        generate_log_parallel(jobs_for_script("A.8-PIN BULK.py")),
        mimetype="text/event-stream",
    )


@app.route("/stream-all-auto-safe")
def stream_all_auto_safe():
    """
    Full automatic pipeline with strict step-by-step order.
    Image step is safety-throttled: 120s delay between projects to reduce ban risk.
    """
    @stream_with_context
    def stream():
        plan = [
            ("START", jobs_for_start1_all_except_s2(), {"kind": "parallel", "env_extra": None, "check_step": "START"}),
            ("JSON", jobs_for_script("A.2-JSON.py"), {"kind": "pool", "max_concurrency": 10, "check_step": "JSON"}),
            ("PROMPT", jobs_for_script("A.2-PROMPT.py"), {"kind": "parallel", "env_extra": None, "check_step": "PROMPT"}),
            ("IMAGINE ALL", jobs_for_script("A.3-IMAGINE.py"), {"kind": "parallel", "env_extra": None, "check_step": "IMAGINE"}),
            ("ARTICLE", jobs_for_script("A.4-ARTICLES.py"), {"kind": "pool", "max_concurrency": 10, "check_step": "ARTICLE"}),
            ("PIN DATA", jobs_for_script("A.5-PIN DATA.py"), {"kind": "pool", "max_concurrency": 10, "env_extra": {"PIN_SKIP_EXISTING": "1"}, "check_step": "PIN DATA"}),
            ("PIN IMAGE", jobs_for_script("A.6-PIN IMAGES.py"), {"kind": "sequential", "delay_s": 120, "check_step": "PIN IMAGE"}),
            ("WP UPLOAD", jobs_for_script("A.7-WP UPLOAD.py"), {"kind": "pool", "max_concurrency": 10, "check_step": "WP UPLOAD"}),
            ("PIN BULK", jobs_for_script("A.8-PIN BULK.py"), {"kind": "parallel", "env_extra": None, "check_step": "PIN DATA"}),
        ]

        yield "data: " + json.dumps(
            {"folder": "all", "line": "🚀 AUTO SAFE pipeline started (step-by-step)."}
        ) + "\n\n"
        yield "data: " + json.dumps(
            {"folder": "all", "line": "🛡️ PIN IMAGE safety mode: 120s delay between projects."}
        ) + "\n\n"

        for step_name, jobs, opts in plan:
            step_jobs = jobs
            check_step = opts.get("check_step")
            if check_step:
                step_jobs = _filter_jobs_missing_step(step_jobs, str(check_step))
            yield "data: " + json.dumps(
                {"folder": "all", "line": f"▶ Starting step: {step_name}"}
            ) + "\n\n"
            if not step_jobs:
                yield "data: " + json.dumps(
                    {"folder": "all", "line": f"⏭ Step {step_name}: already complete for all projects, skipping."}
                ) + "\n\n"
                continue

            kind = opts.get("kind")
            if kind == "pool":
                gen = generate_log_pool(
                    step_jobs,
                    max_concurrency=int(opts.get("max_concurrency", 10)),
                    env_extra=opts.get("env_extra"),
                )
            elif kind == "sequential":
                gen = generate_log_sequential(
                    step_jobs,
                    env_extra=opts.get("env_extra"),
                    delay_between_jobs_s=float(opts.get("delay_s", 0)),
                )
            else:
                gen = generate_log_parallel(step_jobs, env_extra=opts.get("env_extra"))

            for line in gen:
                if '"folder": "all"' in line and "Finished all processes." in line:
                    continue
                yield line

            yield "data: " + json.dumps(
                {"folder": "all", "line": f"✅ Finished step: {step_name}"}
            ) + "\n\n"

        yield "data: " + json.dumps(
            {"folder": "all", "line": "🎉 AUTO SAFE pipeline finished."}
        ) + "\n\n"
        yield 'data: {"folder":"all","line":"Finished all processes."}\n\n'

    return Response(
        stream(),
        mimetype="text/event-stream",
        headers={
            "Cache-Control": "no-cache",
            "X-Accel-Buffering": "no",
        },
    )


# -------------------- Stream Endpoint for START2 --------------------
@app.route("/stream_start2")
def stream_start2():
    return Response(
        generate_log_parallel(jobs_for_start2_only()), mimetype="text/event-stream"
    )


# -------------------- Stream Endpoint for Single Project Actions --------------------
@app.route("/stream-single")
def stream_single():
    """
    Re-trigger a specific action for one project.
    ?project=<project>&action=<action>
    """
    project = request.args.get("project")
    action = request.args.get("action")
    if not project or not action:
        return Response("Missing parameters", status=400)

    action_to_script = {
        "start": "A.1-START.py",
        "json": "A.2-JSON.py",
        "prompt": "A.2-PROMPT.py",
        "imagine": "A.3-IMAGINE.py",
        "article": "A.4-ARTICLES.py",
        "pin_data": "A.5-PIN DATA.py",
        "pin_image": "A.6-PIN IMAGES.py",
        "pin_image_html": "A.6b-PIN IMAGES HTML.py",
        "wp_upload": "A.7-WP UPLOAD.py",
        "pin_bulk": "A.8-PIN BULK.py",
        "cf_upload": "A.9-CF UPLOAD.py",
        "start2": "A.1-START.py"
    }
    script = action_to_script.get(action)
    if not script:
        return Response("Invalid action", status=400)

    u = _unit_by_label(project)
    if not u:
        return Response("Unknown project", status=404)
    env = dict(u.get("env") or {})
    if action == "start":
        raw_lim = (request.args.get("title_limit") or "").strip()
        if raw_lim:
            try:
                lim = max(0, int(raw_lim))
                if lim > 0:
                    env["PINTEREST_START_LIMIT"] = str(lim)
            except ValueError:
                pass
    return Response(
        generate_log_parallel(
            [
                (
                    u["folder"],
                    script,
                    env,
                    u["log_id"],
                    u["label"],
                )
            ]
        ),
        mimetype="text/event-stream",
    )


@app.route("/api/site-config")
def api_site_config():
    """
    Merged config for one dashboard log card (same as subprocess would see for PINTEREST_SITE_ID).
    Query: ?project=<label or log_id> — same value as the START / stream-single buttons.
    """
    project = (request.args.get("project") or "").strip()
    if not project or not is_allowed_project_label(project):
        return jsonify({"error": "Unknown project"}), 404
    u = _unit_by_label(project)
    if not u:
        return jsonify({"error": "Unknown project"}), 404
    folder = (u.get("folder") or "").strip() or "A1-Pinterest_01"
    with _site_config_lock:
        old_sid = os.environ.get("PINTEREST_SITE_ID")
        e = u.get("env") or {}
        try:
            if "PINTEREST_SITE_ID" in e:
                os.environ["PINTEREST_SITE_ID"] = str(e["PINTEREST_SITE_ID"])
            else:
                os.environ.pop("PINTEREST_SITE_ID", None)
            mod = _a1_config_module_for_pipeline_folder(folder)
            if mod is None or not hasattr(mod, "resolved_runtime_snapshot"):
                return jsonify({"error": "a1_config not found in " + folder}), 500
            snap = mod.resolved_runtime_snapshot()
            snap["unit"] = {
                "folder": folder,
                "label": u.get("label"),
                "log_id": u.get("log_id"),
                "subprocess_env": e,
            }
            return jsonify(snap)
        finally:
            if old_sid is not None:
                os.environ["PINTEREST_SITE_ID"] = old_sid
            else:
                os.environ.pop("PINTEREST_SITE_ID", None)


@app.route("/api/site-editor", methods=["GET", "POST"])
def api_site_editor():
    """
    GET: merged snapshot + raw sites.json row for this dashboard project (for Info / edit modal).
    POST JSON { "project": "<label>", "raw_site": { ... } } — writes that row to config/sites.json.
    """
    if request.method == "GET":
        project = (request.args.get("project") or "").strip()
        if not project or not is_allowed_project_label(project):
            return jsonify({"error": "Unknown project"}), 404
        u = _unit_by_label(project)
        if not u:
            return jsonify({"error": "Unknown project"}), 404
        folder = (u.get("folder") or "").strip() or "A1-Pinterest_01"
        idx = _site_index_by_project_label(project)
        raw_site = None
        if idx is not None:
            sites = _load_sites_file_app().get("sites") or []
            if 0 <= idx < len(sites) and isinstance(sites[idx], dict):
                raw_site = copy.deepcopy(sites[idx])
        with _site_config_lock:
            old_sid = os.environ.get("PINTEREST_SITE_ID")
            e = u.get("env") or {}
            try:
                if raw_site and str(raw_site.get("id", "")).strip():
                    os.environ["PINTEREST_SITE_ID"] = str(raw_site.get("id")).strip()
                elif "PINTEREST_SITE_ID" in e:
                    os.environ["PINTEREST_SITE_ID"] = str(e["PINTEREST_SITE_ID"])
                else:
                    os.environ.pop("PINTEREST_SITE_ID", None)
                mod = _a1_config_module_for_pipeline_folder(folder)
                if mod is None or not hasattr(mod, "resolved_runtime_snapshot"):
                    return jsonify({"error": "a1_config not found in " + folder}), 500
                snap = mod.resolved_runtime_snapshot()
            finally:
                if old_sid is not None:
                    os.environ["PINTEREST_SITE_ID"] = old_sid
                else:
                    os.environ.pop("PINTEREST_SITE_ID", None)
        return jsonify(
            {
                "project": project,
                "snapshot": snap,
                "raw_site": raw_site,
                "can_edit_sites_row": raw_site is not None,
            }
        )
    # POST
    data = request.get_json(force=True, silent=True)
    if not isinstance(data, dict):
        return jsonify({"error": "Send JSON: { project, raw_site }"}), 400
    project = (data.get("project") or "").strip()
    raw_site = data.get("raw_site")
    if not project or not is_allowed_project_label(project):
        return jsonify({"error": "Unknown project"}), 404
    if not isinstance(raw_site, dict):
        return jsonify({"error": "raw_site must be a JSON object"}), 400
    if not str(raw_site.get("id", "")).strip():
        return jsonify({"error": "raw_site.id is required"}), 400
    doc = _load_sites_file_app()
    if not isinstance(doc, dict):
        return jsonify({"error": "sites.json is invalid"}), 500
    sites = doc.get("sites")
    if not isinstance(sites, list):
        return jsonify({"error": "sites.json has no sites array"}), 500
    idx = _site_index_by_project_label(project)
    if idx is None or idx >= len(sites):
        return jsonify({"error": "Project row not found in sites.json"}), 404
    old_row = sites[idx] if isinstance(sites[idx], dict) else {}
    merged = copy.deepcopy(raw_site) if isinstance(raw_site, dict) else {}
    for k in _SITE_SECRET_REUSE_KEYS:
        v = (merged.get(k) if k in merged else None)
        t = (str(v).strip() if v is not None else "")
        if t:
            continue
        if k in old_row and (old_row.get(k) is not None) and str(old_row.get(k) or "").strip() != "":
            merged[k] = old_row[k]
    sites[idx] = merged
    doc["sites"] = sites
    doc.setdefault("pipeline_code_folder", "A1-Pinterest_01")
    try:
        _write_sites_doc(doc)
    except OSError as e:
        return jsonify({"error": str(e)}), 500
    return jsonify({"ok": True, "message": "config/sites.json updated for this project row."})


# -------------------- Themes API --------------------
@app.route("/api/themes")
def api_themes():
    """List all available global themes from themes/<slug>/theme.json."""
    themes = _list_themes_full()
    safe = []
    for t in themes:
        safe.append(
            {
                "slug": t.get("slug"),
                "display_name": t.get("display_name") or t.get("slug"),
                "description": t.get("description") or "",
                "version": t.get("version") or "",
                "author": t.get("author") or "",
                "cf_project_name": t.get("cf_project_name") or "",
                "valid": bool(t.get("valid")),
                "has_index": bool(t.get("_has_index")),
                "has_article": bool(t.get("_has_article")),
                "has_static": bool(t.get("_has_static")),
            }
        )
    return jsonify({"ok": True, "themes": safe, "count": len(safe)})


@app.route("/api/project-theme", methods=["GET", "POST"])
def api_project_theme():
    """
    GET ?project=<label>  -> { theme_slug, cf_project_name, cf_button_enabled, reasons[] }
    POST JSON { project, theme_slug, cloudflare_project_name? } -> writes to sites.json
    """
    if request.method == "GET":
        project = (request.args.get("project") or "").strip()
        if not project or not is_allowed_project_label(project):
            return jsonify({"error": "Unknown project"}), 404
        idx = _site_index_by_project_label(project)
        if idx is None:
            return jsonify({"error": "Project not found in sites.json"}), 404
        sites = _load_sites_file_app().get("sites") or []
        row = sites[idx] if 0 <= idx < len(sites) and isinstance(sites[idx], dict) else {}
        theme_slug = str(row.get("theme_slug") or "").strip()
        cf_project_name = str(row.get("cloudflare_project_name") or "").strip()

        shared = _load_shared_keys_app()
        global_token = str(shared.get("cloudflare_api_token") or "").strip()
        global_account = str(shared.get("cloudflare_account_id") or "").strip()
        per_token = str(row.get("cloudflare_api_token") or "").strip()
        per_account = str(row.get("cloudflare_account_id") or "").strip()

        effective_cf_project = cf_project_name
        if not effective_cf_project and theme_slug:
            meta = _read_theme_meta(theme_slug)
            effective_cf_project = str(meta.get("cf_project_name") or "").strip()

        effective_token = per_token or global_token
        effective_account = per_account or global_account

        reasons = []
        if not theme_slug:
            reasons.append("No theme selected for this project.")
        elif not _theme_exists(theme_slug):
            reasons.append(f"Selected theme '{theme_slug}' no longer exists in themes/.")
        if not effective_cf_project:
            reasons.append("No Cloudflare Pages project name set (on project or theme).")
        if not effective_token:
            reasons.append("No Cloudflare API token (set globally in shared_keys.json or override per project).")
        if not effective_account:
            reasons.append("No Cloudflare account_id (set globally in shared_keys.json or override per project).")

        return jsonify(
            {
                "ok": True,
                "project": project,
                "theme_slug": theme_slug,
                "cloudflare_project_name": cf_project_name,
                "effective_cf_project_name": effective_cf_project,
                "cf_button_enabled": len(reasons) == 0,
                "reasons": reasons,
                "has_global_cf_token": bool(global_token),
                "has_global_cf_account": bool(global_account),
                "has_project_cf_override": bool(per_token or per_account),
            }
        )

    # POST
    data = request.get_json(force=True, silent=True)
    if not isinstance(data, dict):
        return jsonify({"error": "Send JSON: { project, theme_slug }"}), 400
    project = (data.get("project") or "").strip()
    if not project or not is_allowed_project_label(project):
        return jsonify({"error": "Unknown project"}), 404
    theme_slug = (data.get("theme_slug") or "").strip()
    if theme_slug and not _theme_exists(theme_slug):
        return jsonify({"error": f"Theme '{theme_slug}' not found in themes/"}), 400
    cf_name = data.get("cloudflare_project_name")
    cf_name_set = "cloudflare_project_name" in data
    cf_token = data.get("cloudflare_api_token")
    cf_token_set = "cloudflare_api_token" in data
    cf_account = data.get("cloudflare_account_id")
    cf_account_set = "cloudflare_account_id" in data

    doc = _load_sites_file_app()
    sites = doc.get("sites") if isinstance(doc, dict) else None
    if not isinstance(sites, list):
        return jsonify({"error": "sites.json has no sites array"}), 500
    idx = _site_index_by_project_label(project)
    if idx is None or idx >= len(sites):
        return jsonify({"error": "Project row not found in sites.json"}), 404
    row = sites[idx] if isinstance(sites[idx], dict) else {}
    if theme_slug:
        row["theme_slug"] = theme_slug
    else:
        row.pop("theme_slug", None)
    if cf_name_set:
        v = str(cf_name or "").strip()
        if v:
            row["cloudflare_project_name"] = v
        else:
            row.pop("cloudflare_project_name", None)
    if cf_token_set:
        v = str(cf_token or "").strip()
        if v:
            row["cloudflare_api_token"] = v
        else:
            row.pop("cloudflare_api_token", None)
    if cf_account_set:
        v = str(cf_account or "").strip()
        if v:
            row["cloudflare_account_id"] = v
        else:
            row.pop("cloudflare_account_id", None)
    sites[idx] = row
    doc["sites"] = sites
    doc.setdefault("pipeline_code_folder", "A1-Pinterest_01")
    try:
        _write_sites_doc(doc)
    except OSError as e:
        return jsonify({"error": str(e)}), 500
    return jsonify({"ok": True, "message": "Theme / Cloudflare config saved to sites.json."})


def _load_shared_keys_app() -> dict:
    """Load config/shared_keys.json (best-effort)."""
    path = os.path.join(_APP_CONFIG, "shared_keys.json")
    if not os.path.isfile(path):
        return {}
    try:
        with open(path, "r", encoding="utf-8") as f:
            d = json.load(f)
        return d if isinstance(d, dict) else {}
    except (json.JSONDecodeError, OSError):
        return {}


def _resolve_cf_creds_for_project(project: str) -> dict:
    """Returns { token, account_id, project_name, errors[] } for the given project label."""
    out = {"token": "", "account_id": "", "project_name": "", "errors": []}
    if not project or not is_allowed_project_label(project):
        out["errors"].append("Unknown project")
        return out
    idx = _site_index_by_project_label(project)
    sites = _load_sites_file_app().get("sites") or []
    if idx is None or idx >= len(sites) or not isinstance(sites[idx], dict):
        out["errors"].append("Project not found in sites.json")
        return out
    row = sites[idx]
    shared = _load_shared_keys_app()
    out["token"] = (str(row.get("cloudflare_api_token") or "").strip()
                    or str(shared.get("cloudflare_api_token") or "").strip())
    out["account_id"] = (str(row.get("cloudflare_account_id") or "").strip()
                         or str(shared.get("cloudflare_account_id") or "").strip())
    name = str(row.get("cloudflare_project_name") or "").strip()
    if not name:
        slug = str(row.get("theme_slug") or "").strip()
        if slug:
            meta = _read_theme_meta(slug)
            name = str(meta.get("cf_project_name") or "").strip()
    out["project_name"] = name
    if not out["token"]:
        out["errors"].append("Missing Cloudflare API token")
    if not out["account_id"]:
        out["errors"].append("Missing Cloudflare account_id")
    if not out["project_name"]:
        out["errors"].append("Missing Cloudflare Pages project name")
    return out


def _cf_api(method: str, path: str, token: str, **kwargs) -> tuple:
    """
    Call the Cloudflare REST API. Returns (status_code, parsed_json_or_text).
    """
    import requests as _requests
    url = "https://api.cloudflare.com/client/v4" + path
    headers = kwargs.pop("headers", {}) or {}
    headers.setdefault("Authorization", f"Bearer {token}")
    headers.setdefault("Accept", "application/json")
    try:
        r = _requests.request(method, url, headers=headers, timeout=30, **kwargs)
    except Exception as e:
        return 0, {"error": f"network: {e}"}
    try:
        return r.status_code, r.json()
    except Exception:
        return r.status_code, {"raw": r.text}


@app.route("/api/cf-deployments")
def api_cf_deployments():
    """List Cloudflare Pages deployments for the project (max ~25 most recent)."""
    project = (request.args.get("project") or "").strip()
    creds = _resolve_cf_creds_for_project(project)
    if creds["errors"]:
        return jsonify({"ok": False, "error": "; ".join(creds["errors"])}), 400
    path = f"/accounts/{creds['account_id']}/pages/projects/{creds['project_name']}/deployments?per_page=25"
    status, body = _cf_api("GET", path, creds["token"])
    if status >= 400 or not isinstance(body, dict) or not body.get("success", False):
        msg = "Cloudflare API error"
        if isinstance(body, dict):
            errs = body.get("errors") or []
            if errs and isinstance(errs[0], dict):
                msg = errs[0].get("message") or msg
        return jsonify({"ok": False, "error": f"{msg} (HTTP {status})", "raw": body}), 502
    results = body.get("result") or []
    deployments = []
    for d in results:
        if not isinstance(d, dict):
            continue
        dep_id = str(d.get("id") or "")
        latest_stage = ""
        stages = d.get("latest_stage")
        if isinstance(stages, dict):
            latest_stage = str(stages.get("name") or "")
        deployments.append(
            {
                "id": dep_id,
                "short_id": dep_id[:8],
                "url": d.get("url") or "",
                "environment": d.get("environment") or "",
                "is_production": (d.get("environment") == "production"),
                "created_on": d.get("created_on") or "",
                "modified_on": d.get("modified_on") or "",
                "stage": latest_stage,
                "deployment_trigger": ((d.get("deployment_trigger") or {}).get("type") or ""),
            }
        )
    return jsonify({"ok": True, "project": project, "cf_project_name": creds["project_name"], "deployments": deployments})


@app.route("/api/cf-rollback", methods=["POST"])
def api_cf_rollback():
    """Promote a previous Pages deployment to production (Cloudflare rollback API)."""
    data = request.get_json(force=True, silent=True)
    if not isinstance(data, dict):
        return jsonify({"ok": False, "error": "Send JSON: { project, deployment_id }"}), 400
    project = (data.get("project") or "").strip()
    deployment_id = (data.get("deployment_id") or "").strip()
    if not deployment_id:
        return jsonify({"ok": False, "error": "deployment_id is required"}), 400
    creds = _resolve_cf_creds_for_project(project)
    if creds["errors"]:
        return jsonify({"ok": False, "error": "; ".join(creds["errors"])}), 400
    path = f"/accounts/{creds['account_id']}/pages/projects/{creds['project_name']}/deployments/{deployment_id}/rollback"
    status, body = _cf_api("POST", path, creds["token"])
    if status >= 400 or not isinstance(body, dict) or not body.get("success", False):
        msg = "Cloudflare API error"
        if isinstance(body, dict):
            errs = body.get("errors") or []
            if errs and isinstance(errs[0], dict):
                msg = errs[0].get("message") or msg
        return jsonify({"ok": False, "error": f"{msg} (HTTP {status})", "raw": body}), 502
    return jsonify({"ok": True, "message": "Rolled back to deployment.", "result": body.get("result")})


@app.route("/api/project-stats")
def api_project_stats():
    project = (request.args.get("project") or "").strip()
    if not project or not is_allowed_project_label(project):
        return jsonify({"ok": False, "error": "Unknown project"}), 404
    s = _project_column_stats(project)
    s["project"] = project
    return jsonify(s)


@app.route("/api/projects-stats")
def api_projects_stats():
    items = []
    for u in flat_run_units():
        label = str(u.get("label", "") or "").strip()
        log_id = str(u.get("log_id", "") or "").strip()
        if not label or not log_id:
            continue
        s = _project_column_stats(label)
        s["project"] = label
        s["log_id"] = log_id
        items.append(s)
    return jsonify({"ok": True, "items": items})


_STEP_CLEAR_ORDER = [
    "START",
    "JSON",
    "PROMPT",
    "IMAGINE",
    "ARTICLE",
    "PIN DATA",
    "PIN IMAGE",
    "WP UPLOAD",
    "PIN BULK",
]

# Columns each pipeline step fills in Recipes.xlsx (or related output).
_STEP_FILL_COLUMNS: Dict[str, List[str]] = {
    "START": ["Title", "Recipe", "Generated At"],
    "JSON": ["Json Recipe"],
    "PROMPT": ["Prompt", "Prompt Image Ingredients"],
    "IMAGINE": [
        "main_image", "image_1", "image_2", "image_3", "image_4",
        "statu", "error",
        "main_image_ingredients", "image_ing_1", "image_ing_2", "image_ing_3", "image_ing_4",
        "statu_ing",
    ],
    "ARTICLE": ["article"],
    "PIN DATA": [
        "recipe_title_pin",
        "pinterest_title",
        "pinterest_description",
        "pinterest_keywords",
        "rank_math_focus_keyword",
        "rank_math_description",
        "rank_math_pillar_content",
        "category",
        "categories",
    ],
    "PIN IMAGE": ["pinterest_image"],
    "PIN IMAGE HTML": ["pinterest_image"],
    "WP UPLOAD": ["status", "post_url"],
    "PIN BULK": [
        "Pin_01.xlsx → Pinterest Pin Link",
        "Pin_01.xlsx → Picture Url 1",
        "Pin_01.xlsx → Text",
        "Pin_01.xlsx → Pinterest Pin Title",
        "Pin_01.xlsx → Pinterest Board",
        "Pin_01.xlsx → Date",
        "Pin_01.xlsx → Time",
    ],
}

_STEP_CLEAR_COLUMNS = {
    "START": _STEP_FILL_COLUMNS["START"],
    "JSON": _STEP_FILL_COLUMNS["JSON"],
    "PROMPT": _STEP_FILL_COLUMNS["PROMPT"],
    "IMAGINE": _STEP_FILL_COLUMNS["IMAGINE"],
    "ARTICLE": _STEP_FILL_COLUMNS["ARTICLE"],
    "PIN DATA": _STEP_FILL_COLUMNS["PIN DATA"],
    "PIN IMAGE": _STEP_FILL_COLUMNS["PIN IMAGE"],
    "WP UPLOAD": _STEP_FILL_COLUMNS["WP UPLOAD"],
    # CLEAR PIN BULK resets pin-metadata columns on Recipes.xlsx (not Pin_01.xlsx).
    "PIN BULK": [
        "recipe_title_pin",
        "pinterest_title",
        "pinterest_description",
        "pinterest_keywords",
        "rank_math_focus_keyword",
        "rank_math_description",
        "rank_math_pillar_content",
        "category",
        "categories",
    ],
}

_STEP_ACTION_TOOLTIPS: Dict[str, Dict[str, str]] = {
    "CLEAR IMAGINE": {
        "run": "Deletes Recipes.xlsx rows where statu or statu_ing is empty or FAILED (failed Midjourney jobs only).",
        "clear": "Deletes Recipes.xlsx rows where statu or statu_ing is empty or FAILED (failed Midjourney jobs only).",
    },
    "AUTO SAFE": {
        "run": "Runs all steps in order (START → … → PIN BULK). Fills every Recipes.xlsx pipeline column; PIN BULK writes Pin_01.xlsx.",
        "clear": "",
    },
    "CLEAR ALL LOGS": {
        "run": "Clears log panel text only — does not change Recipes.xlsx or Pin_01.xlsx.",
        "clear": "Clears log panel text only — does not change Recipes.xlsx or Pin_01.xlsx.",
    },
    "CLEAR PROJECT LOG": {
        "run": "Clears this project's log panel only — does not change Recipes.xlsx or Pin_01.xlsx.",
        "clear": "Clears this project's log panel only — does not change Recipes.xlsx or Pin_01.xlsx.",
    },
    "DELETE ALL": {
        "run": (
            "Full pipeline reset for every project.\n\n"
            "Does:\n"
            "• Archives Recipes.xlsx, Pin_01.xlsx, and output_images/ to ALL/archive/<date-time>/\n"
            "• Clears pipeline columns in Recipes.xlsx (Title, Recipe, Generated At stay)\n"
            "• Empties output_images/ and removes Pin_01.xlsx\n\n"
            "Does NOT delete R2 cloud images or WordPress posts.\n"
            "Use when starting a fresh batch."
        ),
        "clear": (
            "Full pipeline reset for every project.\n"
            "Archives then clears Recipes pipeline data, output_images/, and Pin_01.xlsx."
        ),
    },
    "CLEANUP LOCAL": {
        "run": (
            "Frees local disk space — safe after PIN BULK (and CF UPLOAD if used).\n\n"
            "Removes for ALL projects:\n"
            "• Pin JPG files in ALL/<project>/output_images/ "
            "(copied first to ALL/archive/<date>/)\n"
            "• Old Cloudflare build folders in ALL/_cf_builds/ (keeps the latest build per site)\n\n"
            "Keeps:\n"
            "• Recipes.xlsx and Pin_01.xlsx\n"
            "• All R2 cloud images\n"
            "• WordPress posts and your live site\n\n"
            "Hover preview before click shows file counts."
        ),
        "clear": (
            "Archives and clears local pin JPGs (output_images/) and prunes old CF build folders."
        ),
    },
    "CLEANUP LOCAL PROJECT": {
        "run": (
            "Frees local disk for THIS project only — same rules as Cleanup local.\n\n"
            "Removes:\n"
            "• This project's output_images/ pin JPGs (archived first to ALL/archive/<date>/)\n"
            "• Old CF builds for this site (keeps the latest)\n\n"
            "Keeps:\n"
            "• Recipes.xlsx, Pin_01.xlsx, R2 images, WordPress\n\n"
            "Run after PIN BULK for this project."
        ),
        "clear": "Local cleanup for this project only.",
    },
    "CLEANUP R2": {
        "run": (
            "Deletes unused images from Cloudflare R2 — permanent.\n\n"
            "Removes:\n"
            "• Extra Midjourney splits/grids (image_2–4, unused grids)\n"
            "• Old pinterest_local / pinterest_remote copies not referenced in Excel\n\n"
            "Keeps:\n"
            "• Every URL still in Recipes.xlsx or Pin_01.xlsx "
            "(image_1, image_ing_1, Picture Url 1, etc.)\n\n"
            "You must type DELETE R2 to confirm.\n"
            "Best after: pins uploaded to Pinterest and site looks correct."
        ),
        "clear": (
            "Deletes unreferenced R2 objects. URLs in Recipes.xlsx / Pin_01.xlsx are kept."
        ),
    },
}


def _step_columns_meta_for_ui() -> Dict[str, Any]:
    """JSON-safe metadata for dashboard button tooltips."""
    steps = {}
    for step, cols in _STEP_FILL_COLUMNS.items():
        steps[step] = {
            "fills": list(cols),
            "clears": list(_STEP_CLEAR_COLUMNS.get(step, cols)),
        }
    return {
        "order": list(_STEP_CLEAR_ORDER),
        "steps": steps,
        "special": dict(_STEP_ACTION_TOOLTIPS),
    }


def _normalize_step_name(step: str) -> str:
    return str(step or "").strip().upper()


def _cascade_steps_from(step: str) -> list:
    s = _normalize_step_name(step)
    try:
        idx = _STEP_CLEAR_ORDER.index(s)
    except ValueError:
        return []
    return _STEP_CLEAR_ORDER[idx:]


def _step_name_for_column_py(name: str) -> str:
    lk = str(name or "").strip().lower()
    if lk in {"title", "recipe", "generated at"}:
        return "START"
    if lk in {"json recipe"}:
        return "JSON"
    if lk in {"prompt", "prompt image ingredients"}:
        return "PROMPT"
    if lk in {
        "main_image", "image_1", "image_2", "image_3", "image_4", "statu", "error",
        "main_image_ingredients", "image_ing_1", "image_ing_2", "image_ing_3", "image_ing_4", "statu_ing",
    }:
        return "IMAGINE"
    if lk in {"article"}:
        return "ARTICLE"
    if lk in {
        "recipe_title_pin", "pinterest_title", "pinterest_description", "pinterest_keywords",
        "rank_math_focus_keyword", "rank_math_description", "rank_math_pillar_content",
        "category", "categories",
    }:
        return "PIN DATA"
    if lk in {"pinterest_image"}:
        return "PIN IMAGE"
    if lk in {"status", "post_url"}:
        return "WP UPLOAD"
    return "OTHER"


def _project_step_complete(project_label: str, step_name: str) -> bool:
    """
    True if all columns belonging to step_name are fully filled for this project.
    """
    step = _normalize_step_name(step_name)
    s = _project_column_stats(project_label)
    if not isinstance(s, dict) or not s.get("ok"):
        return False
    cols = s.get("columns") or []
    if not isinstance(cols, list) or not cols:
        return False
    step_cols = [c for c in cols if _step_name_for_column_py((c or {}).get("name", "")) == step]
    if not step_cols:
        return False
    for c in step_cols:
        # Do not use error columns as completion blockers for step status.
        cname = str((c or {}).get("name", "") or "").strip().lower()
        if cname in {"error", "error_ing"}:
            continue
        filled = int((c or {}).get("filled", 0) or 0)
        total = int((c or {}).get("total", 0) or 0)
        if total > 0 and filled < total:
            return False
    return True


def _filter_jobs_missing_step(script_jobs: list, step_name: str) -> list:
    """
    Keep only jobs for projects where the given step is NOT complete yet.
    """
    out = []
    skipped = 0
    for item in _normalize_script_jobs(script_jobs):
        folder, script, env, log_id, line_label = item
        label = str(line_label or "")
        if _project_step_complete(label, step_name):
            skipped += 1
            continue
        out.append((folder, script, env, log_id, line_label))
    return out


def _clear_steps_in_project_excel(file_path: str, steps_to_clear: list) -> dict:
    wb = openpyxl.load_workbook(file_path)
    sh = wb.active
    max_col = int(sh.max_column or 0)
    max_row = int(sh.max_row or 0)
    if max_col <= 0 or max_row <= 1:
        return {"cleared_cells": 0, "columns_found": [], "columns_missing": []}

    header_values = next(
        sh.iter_rows(min_row=1, max_row=1, min_col=1, max_col=max_col, values_only=True),
        tuple(),
    )
    by_lower = {}
    for i in range(max_col):
        hv = header_values[i] if i < len(header_values) else None
        h = str(hv).strip() if hv is not None else ""
        if h:
            by_lower[h.lower()] = i + 1  # 1-based column index

    target_names = []
    for step in steps_to_clear:
        target_names.extend(_STEP_CLEAR_COLUMNS.get(step, []))
    # preserve order + dedupe
    seen = set()
    ordered_names = []
    for n in target_names:
        lk = str(n).strip().lower()
        if lk in seen:
            continue
        seen.add(lk)
        ordered_names.append(n)

    found_cols = []
    missing_cols = []
    found_idxs = []
    for name in ordered_names:
        lk = str(name).strip().lower()
        idx = by_lower.get(lk)
        if idx is None:
            missing_cols.append(name)
            continue
        found_cols.append(name)
        found_idxs.append(idx)

    cleared = 0
    if found_idxs:
        for r in range(2, max_row + 1):
            for cidx in found_idxs:
                cell = sh.cell(row=r, column=cidx)
                if cell.value is not None and str(cell.value) != "":
                    cleared += 1
                cell.value = None
        wb.save(file_path)
    wb.close()
    return {
        "cleared_cells": int(cleared),
        "columns_found": found_cols,
        "columns_missing": missing_cols,
    }


@app.route("/api/clear-step", methods=["POST"])
def api_clear_step():
    payload = request.get_json(silent=True) or {}
    step = _normalize_step_name(payload.get("step", ""))
    if not step:
        return jsonify({"ok": False, "error": "Missing step"}), 400
    steps_to_clear = _cascade_steps_from(step)
    if not steps_to_clear:
        return jsonify({"ok": False, "error": "Unknown step", "allowed_steps": _STEP_CLEAR_ORDER}), 400

    per_project = []
    total_cleared = 0
    for project in flat_ui_labels():
        out_dir = all_out_name_for_label(project)
        file_path = _project_excel_path_by_out_dir(out_dir)
        if not os.path.exists(file_path):
            per_project.append(
                {"project": project, "ok": False, "error": "file_not_found", "file_path": file_path}
            )
            continue
        try:
            result = _clear_steps_in_project_excel(file_path, steps_to_clear)
            total_cleared += int(result.get("cleared_cells", 0) or 0)
            per_project.append(
                {
                    "project": project,
                    "ok": True,
                    "file_path": file_path,
                    **result,
                }
            )
        except Exception as e:
            per_project.append(
                {"project": project, "ok": False, "error": str(e), "file_path": file_path}
            )

    _PROJECT_STATS_CACHE.clear()
    return jsonify(
        {
            "ok": True,
            "step": step,
            "cascade_steps": steps_to_clear,
            "total_cleared_cells": int(total_cleared),
            "projects": per_project,
        }
    )


@app.route("/files/all/<path:relpath>")
def serve_all_file(relpath):
    """Serve a file located inside the ALL/ directory.

    Used by the column-details popup so file-path values (e.g. generated
    images) become clickable links that open in a new tab.
    """
    rel = (relpath or "").replace("\\", "/").lstrip("/")
    if not rel or ".." in rel.split("/"):
        abort(404)
    all_root = os.path.abspath(_all_dir_path())
    target = os.path.abspath(os.path.join(all_root, rel))
    if os.path.commonpath([target, all_root]) != all_root:
        abort(404)
    if not os.path.isfile(target):
        abort(404)
    return send_from_directory(all_root, rel, as_attachment=False)


@app.route("/api/project-column-details")
def api_project_column_details():
    project = (request.args.get("project") or "").strip()
    column = (request.args.get("column") or "").strip()
    if not project or not is_allowed_project_label(project):
        return jsonify({"ok": False, "error": "Unknown project"}), 404
    if not column:
        return jsonify({"ok": False, "error": "Missing column"}), 400
    d = _project_column_details(project, column)
    d["project"] = project
    return jsonify(d)


@app.route("/manage_sites", methods=["GET", "POST"])
def manage_sites():
    if request.method == "POST":
        act = (request.form.get("action") or "save").strip()
        if act == "add":
            data = _load_sites_file_app()
            if not isinstance(data, dict):
                data = {}
            sites = data.get("sites")
            if not isinstance(sites, list):
                sites = []
            sites.append(_default_new_site(sites))
            data["sites"] = sites
            data["pipeline_code_folder"] = (str(data.get("pipeline_code_folder") or "A1-Pinterest_01").strip() or "A1-Pinterest_01")
            try:
                _write_sites_doc(data)
                _ensure_project_output_files_for_sites(data)
            except OSError as e:
                flash("Could not save sites.json: " + str(e), "error")
                return redirect(url_for("manage_sites"))
            flash("New project added. Edit fields below, then save all.", "success")
            return redirect(url_for("manage_sites"))
        if act == "delete":
            did = (request.form.get("site_id") or "").strip()
            if did:
                data = _load_sites_file_app()
                sites = [s for s in (data.get("sites") or []) if isinstance(s, dict) and str(s.get("id", "")) != did]
                data["sites"] = sites
                data.setdefault("pipeline_code_folder", "A1-Pinterest_01")
                try:
                    _write_sites_doc(data)
                    _ensure_project_output_files_for_sites(data)
                except OSError as e:
                    flash("Could not save sites.json: " + str(e), "error")
                    return redirect(url_for("manage_sites"))
                flash("Removed project " + did, "success")
            return redirect(url_for("manage_sites"))
        # full save
        n = int(request.form.get("site_count", 0) or 0)
        pl = (request.form.get("pipeline_code_folder") or "").strip() or "A1-Pinterest_01"
        old_list = _load_sites_file_app().get("sites") or []
        if not isinstance(old_list, list):
            old_list = []
        old_by_id: Dict[str, dict] = {
            str(s.get("id")): s for s in old_list if isinstance(s, dict) and s.get("id")
        }
        seen: set = set()
        sites_out: List[dict] = []
        for i in range(n):
            sub_id = (request.form.get(f"site_{i}_id") or "").strip()
            old = old_by_id.get(sub_id, {})
            try:
                one = _site_from_form(request.form, i, old)
            except ValueError as e:
                flash(str(e), "error")
                return redirect(url_for("manage_sites"))
            if not (one or {}).get("id"):
                continue
            if one["id"] in seen:
                flash("Duplicate site id: " + str(one["id"]), "error")
                return redirect(url_for("manage_sites"))
            seen.add(one["id"])
            sites_out.append(one)
        if not sites_out:
            flash("At least one site with an id is required.", "error")
            return redirect(url_for("manage_sites"))
        doc = {"pipeline_code_folder": pl, "sites": sites_out}
        try:
            _write_sites_doc(doc)
            _ensure_project_output_files_for_sites(doc)
        except OSError as e:
            flash("Could not write config/sites.json: " + str(e), "error")
            return redirect(url_for("manage_sites"))
        flash(
            "config/sites.json saved. Refresh the dashboard; restart the app if IMAGINE group buttons should match the new site list.",
            "success",
        )
        return redirect(url_for("manage_sites"))
    d = _load_sites_file_app()
    _ensure_project_output_files_for_sites(d)
    if not isinstance(d, dict) or "sites" not in d or not isinstance(d.get("sites"), list):
        d = {"pipeline_code_folder": "A1-Pinterest_01", "sites": []}
    d.setdefault("pipeline_code_folder", "A1-Pinterest_01")
    prompt_schema: Dict[str, list] = {}
    settings_schema: List[str] = []
    settings_groups: Dict[str, List[str]] = {}
    prompt_base_by_site_id: Dict[str, Dict[str, Any]] = {}
    settings_base_by_site_id: Dict[str, Dict[str, Any]] = {}
    r2_effective_by_site_id: Dict[str, Dict[str, str]] = {}
    shared_keys_doc = _load_shared_keys_app()
    shared_r2_defaults = {
        "r2_account_id": str(shared_keys_doc.get("r2_account_id") or "").strip(),
        "r2_bucket": str(shared_keys_doc.get("r2_bucket") or "").strip(),
        "r2_public_base_url": str(shared_keys_doc.get("r2_public_base_url") or "").strip(),
    }
    try:
        pipeline_folder = str(d.get("pipeline_code_folder") or "A1-Pinterest_01")
        mod = _a1_config_module_for_pipeline_folder(pipeline_folder)
        if mod is not None and hasattr(mod, "prompts_inline_field_schema"):
            raw_schema = mod.prompts_inline_field_schema()
            if isinstance(raw_schema, dict):
                prompt_schema = raw_schema
        if mod is not None and hasattr(mod, "resolved_runtime_snapshot"):
            with _site_config_lock:
                old_sid = os.environ.get("PINTEREST_SITE_ID")
                try:
                    for s in d.get("sites") or []:
                        if not isinstance(s, dict):
                            continue
                        sid = str(s.get("id", "") or "").strip()
                        if not sid:
                            continue
                        os.environ["PINTEREST_SITE_ID"] = sid
                        try:
                            snap = mod.resolved_runtime_snapshot()
                        except Exception:
                            continue
                        r2_eff = (snap or {}).get("r2_effective")
                        if isinstance(r2_eff, dict):
                            r2_effective_by_site_id[sid] = {
                                "r2_bucket": str(r2_eff.get("r2_bucket") or "").strip(),
                                "r2_public_base_url": str(r2_eff.get("r2_public_base_url") or "").strip(),
                            }
                        base = (snap or {}).get("prompts_excluding_row_inline_by_path")
                        if isinstance(base, dict):
                            prompt_base_by_site_id[sid] = base
                        raw_settings = (snap or {}).get("settings")
                        if isinstance(raw_settings, dict):
                            def _flatten_settings(din: Any, pref: str, out: Dict[str, Any]) -> None:
                                if not isinstance(din, dict):
                                    return
                                for kk, vv in din.items():
                                    sub = f"{pref}.{kk}" if pref else str(kk)
                                    if isinstance(vv, dict):
                                        # Keep simple mapping objects as one editable JSON field
                                        # (e.g. category_id_mapping) instead of exploding to many inputs.
                                        if vv and all(not isinstance(x, (dict, list)) for x in vv.values()):
                                            out[sub] = vv
                                        elif vv:
                                            _flatten_settings(vv, sub, out)
                                        else:
                                            out[sub] = vv
                                    else:
                                        out[sub] = vv
                            flat_settings: Dict[str, Any] = {}
                            _flatten_settings(raw_settings, "", flat_settings)
                            settings_base_by_site_id[sid] = flat_settings
                            for kpath in flat_settings.keys():
                                if kpath not in settings_schema:
                                    settings_schema.append(kpath)
                finally:
                    if old_sid is not None:
                        os.environ["PINTEREST_SITE_ID"] = old_sid
                    else:
                        os.environ.pop("PINTEREST_SITE_ID", None)
    except Exception:
        prompt_schema = {}

    def _get_nested(obj: dict, path: str):
        cur = obj
        for pp in (path or "").split("."):
            if not isinstance(cur, dict) or pp not in cur:
                return None
            cur = cur.get(pp)
        return cur

    sites_view: List[dict] = []
    for s in d.get("sites") or []:
        if not isinstance(s, dict):
            continue
        sv = dict(s)
        st = s.get("settings")
        sv["_settings_json"] = (
            json.dumps(st, ensure_ascii=False, indent=2) if isinstance(st, (dict, list)) else ""
        )
        pr = s.get("prompts")
        sv["_prompts_json"] = (
            json.dumps(pr, ensure_ascii=False, indent=2) if isinstance(pr, (dict, list)) else ""
        )
        prompt_values: Dict[str, str] = {}
        prompt_placeholders: Dict[str, str] = {}
        settings_values: Dict[str, str] = {}
        settings_placeholders: Dict[str, str] = {}
        sid = str(s.get("id", "") or "").strip()
        base_for_site = prompt_base_by_site_id.get(sid, {})
        settings_base_for_site = settings_base_by_site_id.get(sid, {})
        site_settings = s.get("settings") if isinstance(s.get("settings"), dict) else {}

        def _get_nested(obj: dict, path: str):
            cur = obj
            for pp in (path or "").split("."):
                if not isinstance(cur, dict) or pp not in cur:
                    return None
                cur = cur.get(pp)
            return cur

        for sp in settings_schema:
            vv = _get_nested(site_settings, sp) if isinstance(site_settings, dict) else None
            if vv is not None:
                if isinstance(vv, (dict, list)):
                    settings_values[sp.replace(".", "__")] = json.dumps(vv, ensure_ascii=False)
                else:
                    settings_values[sp.replace(".", "__")] = str(vv)
            bv = settings_base_for_site.get(sp)
            if bv is not None:
                if isinstance(bv, (dict, list)):
                    settings_placeholders[sp.replace(".", "__")] = json.dumps(bv, ensure_ascii=False)
                else:
                    settings_placeholders[sp.replace(".", "__")] = str(bv)
        if isinstance(pr, dict):
            for pn, fields in prompt_schema.items():
                if not isinstance(fields, list):
                    continue
                pnode = pr.get(pn)
                if not isinstance(pnode, dict):
                    continue
                for f in fields:
                    if not isinstance(f, dict):
                        continue
                    path = str(f.get("path") or "").strip()
                    if not path:
                        continue
                    vv = _get_nested(pnode, path)
                    if vv is None:
                        continue
                    key = f"{pn}__{path.replace('.', '__')}"
                    if isinstance(vv, (dict, list)):
                        prompt_values[key] = json.dumps(vv, ensure_ascii=False)
                    else:
                        prompt_values[key] = str(vv)
        for pn, fields in prompt_schema.items():
            if not isinstance(fields, list):
                continue
            flat_base = base_for_site.get(pn)
            if not isinstance(flat_base, dict):
                continue
            for f in fields:
                if not isinstance(f, dict):
                    continue
                path = str(f.get("path") or "").strip()
                if not path:
                    continue
                bval = flat_base.get(path)
                if bval is None:
                    continue
                key = f"{pn}__{path.replace('.', '__')}"
                if isinstance(bval, (dict, list)):
                    prompt_placeholders[key] = json.dumps(bval, ensure_ascii=False)
                else:
                    prompt_placeholders[key] = str(bval)
        sv["_prompt_values"] = prompt_values
        sv["_prompt_placeholders"] = prompt_placeholders
        sv["_settings_values"] = settings_values
        sv["_settings_placeholders"] = settings_placeholders
        sid = str(sv.get("id", "") or "").strip()
        sv["_r2_effective"] = r2_effective_by_site_id.get(sid, {})
        sites_view.append(sv)

    if settings_schema:
        settings_groups = {
            "General": [],
            "A2": [],
            "A3": [],
            "A4": [],
            "A5": [],
            "A6": [],
            "A8": [],
            "Other": [],
        }
        for sp in settings_schema:
            lk = str(sp or "").lower()
            if lk.startswith("a2_"):
                settings_groups["A2"].append(sp)
            elif lk.startswith("a3_"):
                settings_groups["A3"].append(sp)
            elif lk.startswith("a4_"):
                settings_groups["A4"].append(sp)
            elif lk.startswith("a5_"):
                settings_groups["A5"].append(sp)
            elif lk.startswith("a6_"):
                settings_groups["A6"].append(sp)
            elif lk.startswith("a8_"):
                settings_groups["A8"].append(sp)
            elif "." in lk:
                settings_groups["Other"].append(sp)
            else:
                settings_groups["General"].append(sp)
        settings_groups = {k: v for k, v in settings_groups.items() if v}
    return render_template(
        "manage_sites.html",
        data=d,
        sites_view=sites_view,
        site_count=len(sites_view),
        prompt_schema=prompt_schema,
        settings_schema=settings_schema,
        settings_groups=settings_groups,
        shared_r2_defaults=shared_r2_defaults,
    )


@app.route("/manage_starts")
def manage_starts():
    return render_template("manage_starts.html")


@app.route("/manage_recipes")
def manage_recipes():
    return render_template("manage_recipes.html")


# -------------------- WP UPLOAD (pool 10 ب 10) --------------------
@app.route("/stream-all-wp-upload")
def stream_all_wp_upload():
    return Response(
        generate_log_pool(jobs_for_script("A.7-WP UPLOAD.py"), max_concurrency=10),
        mimetype="text/event-stream"
    )


# -------------------- CF UPLOAD ALL (pool 4-by-4) --------------------
@app.route("/stream-all-cf-upload")
def stream_all_cf_upload():
    """
    Deploy every configured project to Cloudflare Pages. Each project runs
    A.9-CF UPLOAD.py in its own subprocess. The script itself short-circuits
    with a clear log line for projects that don't have a theme_slug or
    cloudflare_project_name configured.
    Pool is small (4) since each wrangler invocation hits the CF API hard.
    """
    return Response(
        generate_log_pool(jobs_for_script("A.9-CF UPLOAD.py"), max_concurrency=4),
        mimetype="text/event-stream"
    )

# -------------------- Endpoint to Stop Running Scripts --------------------
@app.route("/stop_scripts", methods=["POST"])
def stop_scripts():
    global running_processes
    for proc in running_processes[:]:
        try:
            proc.terminate()
            proc.wait(timeout=5)
        except Exception:
            try:
                proc.kill()
            except Exception:
                pass
    running_processes.clear()
    return jsonify({"status": "success", "message": "All running scripts have been stopped."})


# -------------------- 3) Delete 'ALL' — archive then clear --------------------

def _all_archive_root() -> str:
    return os.path.join(_APP_ROOT, "ALL", "archive")


def _new_archive_session_dir() -> str:
    stamp = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
    path = os.path.join(_all_archive_root(), stamp)
    os.makedirs(path, exist_ok=True)
    return path


def _project_output_dir(out_dir: str) -> str:
    return os.path.join(_APP_ROOT, "ALL", out_dir)


def _archive_project_outputs(out_dir: str, archive_session: str) -> Dict[str, str]:
    """Copy Recipes.xlsx, Pin_01.xlsx, output_images/ into ALL/archive/<stamp>/<out_dir>/."""
    project_dir = _project_output_dir(out_dir)
    arch_dir = os.path.join(archive_session, out_dir)
    os.makedirs(arch_dir, exist_ok=True)
    out: Dict[str, str] = {}

    recipes = os.path.join(project_dir, "Recipes.xlsx")
    legacy = os.path.join(project_dir, "images.xlsx")
    if os.path.isfile(recipes):
        shutil.copy2(recipes, os.path.join(arch_dir, "Recipes.xlsx"))
        out["Recipes.xlsx"] = "archived"
    elif os.path.isfile(legacy):
        shutil.copy2(legacy, os.path.join(arch_dir, "images.xlsx"))
        out["images.xlsx"] = "archived"
    else:
        out["Recipes.xlsx"] = "missing"

    pin_path = os.path.join(project_dir, "Pin_01.xlsx")
    if os.path.isfile(pin_path):
        shutil.copy2(pin_path, os.path.join(arch_dir, "Pin_01.xlsx"))
        out["Pin_01.xlsx"] = "archived"
    else:
        out["Pin_01.xlsx"] = "missing"

    imgs_src = os.path.join(project_dir, "output_images")
    imgs_dst = os.path.join(arch_dir, "output_images")
    if os.path.isdir(imgs_src):
        if os.path.exists(imgs_dst):
            shutil.rmtree(imgs_dst)
        shutil.copytree(imgs_src, imgs_dst)
        n = sum(len(files) for _r, _d, files in os.walk(imgs_dst))
        out["output_images"] = f"archived ({n} files)"
    else:
        out["output_images"] = "missing"

    return out


def _clear_recipes_pipeline_rows(file_path: str) -> int:
    """Clear pipeline columns in Recipes.xlsx; keep Title, Recipe, Generated At."""
    protected_start_cols = {"title", "recipe", "generated at"}
    wb = openpyxl.load_workbook(file_path)
    sh = wb.active
    max_row = int(sh.max_row or 1)
    max_col = int(sh.max_column or 1)
    header_values = next(
        sh.iter_rows(min_row=1, max_row=1, min_col=1, max_col=max_col, values_only=True),
        tuple(),
    )
    clear_col_idxs = []
    for c in range(1, max_col + 1):
        hv = header_values[c - 1] if c - 1 < len(header_values) else None
        h = str(hv).strip().lower() if hv is not None else ""
        if h in protected_start_cols:
            continue
        clear_col_idxs.append(c)
    cleared_rows = 0
    if max_row >= 2:
        for r in range(2, max_row + 1):
            for c in clear_col_idxs:
                sh.cell(row=r, column=c, value=None)
            cleared_rows += 1
    wb.save(file_path)
    wb.close()
    return cleared_rows


def _clear_project_after_archive(out_dir: str) -> Dict[str, str]:
    project_dir = _project_output_dir(out_dir)
    out: Dict[str, str] = {}
    file_path = _project_excel_path_by_out_dir(out_dir)
    if os.path.isfile(file_path):
        try:
            rows = _clear_recipes_pipeline_rows(file_path)
            out["Recipes.xlsx"] = f"cleared ({rows} rows)"
        except Exception as e:
            out["Recipes.xlsx"] = f"error: {e}"
    else:
        out["Recipes.xlsx"] = "missing"

    imgs = os.path.join(project_dir, "output_images")
    try:
        if os.path.isdir(imgs):
            shutil.rmtree(imgs)
        os.makedirs(imgs, exist_ok=True)
        out["output_images"] = "cleared"
    except Exception as e:
        out["output_images"] = f"error: {e}"

    pin_path = os.path.join(project_dir, "Pin_01.xlsx")
    try:
        if os.path.isfile(pin_path):
            os.remove(pin_path)
            out["Pin_01.xlsx"] = "removed"
        else:
            out["Pin_01.xlsx"] = "missing"
    except Exception as e:
        out["Pin_01.xlsx"] = f"error: {e}"

    return out


@app.route("/delete-all-folder", methods=["POST"])
def delete_all_folder():
    """
    Archive then clear all projects:
    - ALL/archive/<YYYY-MM-DD_HH-MM-SS>/<out_dir>/Recipes.xlsx, Pin_01.xlsx, output_images/
    - Clear pipeline columns in Recipes.xlsx (keep Title, Recipe, Generated At)
    - Empty output_images/ and remove Pin_01.xlsx from each project folder
    """
    archive_session = _new_archive_session_dir()
    archive_rel = os.path.relpath(archive_session, _APP_ROOT).replace("\\", "/")
    results = []
    total_cleared = 0
    any_found = False

    for label in flat_ui_labels():
        out_dir = all_out_name_for_label(label)
        project_dir = _project_output_dir(out_dir)
        has_any = (
            os.path.isfile(os.path.join(project_dir, "Recipes.xlsx"))
            or os.path.isfile(os.path.join(project_dir, "images.xlsx"))
            or os.path.isfile(os.path.join(project_dir, "Pin_01.xlsx"))
            or os.path.isdir(os.path.join(project_dir, "output_images"))
        )
        if not has_any:
            results.append({
                "label": label,
                "out_dir": out_dir,
                "status": "missing",
                "archive": {},
                "clear": {},
            })
            continue
        any_found = True
        try:
            archived = _archive_project_outputs(out_dir, archive_session)
            cleared = _clear_project_after_archive(out_dir)
            rows_part = cleared.get("Recipes.xlsx", "")
            if "cleared (" in rows_part:
                try:
                    total_cleared += int(rows_part.split("cleared (")[1].split(" rows")[0])
                except (ValueError, IndexError):
                    pass
            results.append({
                "label": label,
                "out_dir": out_dir,
                "status": "ok",
                "archive": archived,
                "clear": cleared,
            })
        except Exception as e:
            results.append({
                "label": label,
                "out_dir": out_dir,
                "status": f"error: {e}",
                "archive": {},
                "clear": {},
            })

    _PROJECT_STATS_CACHE.clear()

    if not any_found:
        return (
            "<h1>No project output found to archive.</h1>"
            f"<p>Expected under <code>ALL/&lt;out_dir&gt;/</code></p>"
            "<a href='/'>Back</a>"
        )

    lines = [
        "<h1>Archive + clear completed</h1>",
        f"<p><strong>Archive folder:</strong> <code>{archive_rel}</code></p>",
        f"<p>Total recipe rows cleared (pipeline columns only): <strong>{total_cleared}</strong></p>",
        "<ul>",
    ]
    for row in results:
        lines.append(f"<li><strong>{row['label']}</strong> (<code>{row['out_dir']}</code>) — {row['status']}")
        if row.get("archive"):
            lines.append("<ul>")
            for k, v in row["archive"].items():
                lines.append(f"<li>Archive {k}: {v}</li>")
            lines.append("</ul>")
        if row.get("clear"):
            lines.append("<ul>")
            for k, v in row["clear"].items():
                lines.append(f"<li>Clear {k}: {v}</li>")
            lines.append("</ul>")
        lines.append("</li>")
    lines.append("</ul>")
    lines.append("<a href='/'>Back</a>")
    return "".join(lines)


# -------------------- Cleanup (local + R2) --------------------

_CF_BUILDS_ROOT = os.path.join(_APP_ROOT, "ALL", "_cf_builds")
_R2_CLEANUP_PREFIXES = (
    "midjourney_grids/",
    "midjourney_splits/",
    "pinterest_local/",
    "pinterest_remote/",
    "pinterest_images/",
)
_IMAGE_URL_HEADERS = {
    "main_image",
    "image_1",
    "image_2",
    "image_3",
    "image_4",
    "main_image_ingredients",
    "image_ing_1",
    "image_ing_2",
    "image_ing_3",
    "image_ing_4",
    "pinterest_image",
    "picture url 1",
}


def _format_bytes(n: int) -> str:
    size = float(max(0, int(n)))
    for unit in ("B", "KB", "MB", "GB"):
        if size < 1024 or unit == "GB":
            if unit == "B":
                return f"{int(size)} {unit}"
            return f"{size:.1f} {unit}"
        size /= 1024.0
    return f"{int(n)} B"


def _dir_size_and_count(path: str) -> tuple:
    if not os.path.isdir(path):
        return 0, 0
    total = 0
    count = 0
    for _root, _dirs, files in os.walk(path):
        for nm in files:
            fp = os.path.join(_root, nm)
            try:
                total += os.path.getsize(fp)
                count += 1
            except OSError:
                pass
    return count, total


def _archive_output_images_only(out_dir: str, archive_session: str) -> Dict[str, str]:
    project_dir = _project_output_dir(out_dir)
    arch_dir = os.path.join(archive_session, out_dir)
    os.makedirs(arch_dir, exist_ok=True)
    out: Dict[str, str] = {}
    imgs_src = os.path.join(project_dir, "output_images")
    imgs_dst = os.path.join(arch_dir, "output_images")
    if os.path.isdir(imgs_src):
        if os.path.exists(imgs_dst):
            shutil.rmtree(imgs_dst)
        shutil.copytree(imgs_src, imgs_dst)
        n, sz = _dir_size_and_count(imgs_dst)
        out["output_images"] = f"archived ({n} files, {_format_bytes(sz)})"
    else:
        out["output_images"] = "missing"
    return out


def _clear_output_images_dir(out_dir: str) -> Dict[str, str]:
    project_dir = _project_output_dir(out_dir)
    imgs = os.path.join(project_dir, "output_images")
    try:
        if os.path.isdir(imgs):
            shutil.rmtree(imgs)
        os.makedirs(imgs, exist_ok=True)
        return {"output_images": "cleared"}
    except Exception as e:
        return {"output_images": f"error: {e}"}


def _cleanup_cf_builds(*, keep_latest_per_site: int = 1, site_id_filter: Optional[str] = None) -> Dict[str, Any]:
    """Remove old Cloudflare static build folders; keep the newest N per site."""
    keep = max(0, int(keep_latest_per_site))
    removed_dirs = 0
    freed_bytes = 0
    kept_dirs = 0
    if not os.path.isdir(_CF_BUILDS_ROOT):
        return {
            "removed_dirs": 0,
            "kept_dirs": 0,
            "freed_bytes": 0,
            "freed_human": _format_bytes(0),
        }
    for site_nm in os.listdir(_CF_BUILDS_ROOT):
        if site_id_filter and site_nm != site_id_filter:
            continue
        site_path = os.path.join(_CF_BUILDS_ROOT, site_nm)
        if not os.path.isdir(site_path):
            continue
        builds = []
        for build_nm in os.listdir(site_path):
            build_path = os.path.join(site_path, build_nm)
            if not os.path.isdir(build_path):
                continue
            try:
                mtime = os.path.getmtime(build_path)
            except OSError:
                mtime = 0
            builds.append((mtime, build_path))
        builds.sort(key=lambda x: x[0], reverse=True)
        for i, (_mtime, build_path) in enumerate(builds):
            if i < keep:
                kept_dirs += 1
                continue
            n, sz = _dir_size_and_count(build_path)
            try:
                shutil.rmtree(build_path)
                removed_dirs += 1
                freed_bytes += sz
            except OSError:
                pass
    return {
        "removed_dirs": removed_dirs,
        "kept_dirs": kept_dirs,
        "freed_bytes": freed_bytes,
        "freed_human": _format_bytes(freed_bytes),
    }


def _collect_http_urls_from_workbook(path: str) -> set:
    urls: set = set()
    if not os.path.isfile(path):
        return urls
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        sh = wb.active
        max_col = int(sh.max_column or 1)
        max_row = int(sh.max_row or 1)
        header_values = next(
            sh.iter_rows(min_row=1, max_row=1, min_col=1, max_col=max_col, values_only=True),
            tuple(),
        )
        cols: List[int] = []
        for c in range(1, max_col + 1):
            hv = header_values[c - 1] if c - 1 < len(header_values) else None
            h = str(hv).strip().lower() if hv is not None else ""
            if h in _IMAGE_URL_HEADERS:
                cols.append(c)
        if max_row >= 2:
            for r in range(2, max_row + 1):
                for c in cols:
                    v = sh.cell(row=r, column=c).value
                    if v is None:
                        continue
                    s = str(v).strip()
                    if s.lower().startswith("http"):
                        urls.add(s)
    finally:
        wb.close()
    return urls


def _collect_referenced_r2_urls_for_out_dir(out_dir: str) -> set:
    urls: set = set()
    project_dir = _project_output_dir(out_dir)
    recipes = os.path.join(project_dir, "Recipes.xlsx")
    legacy = os.path.join(project_dir, "images.xlsx")
    pin_path = os.path.join(project_dir, "Pin_01.xlsx")
    if os.path.isfile(recipes):
        urls |= _collect_http_urls_from_workbook(recipes)
    elif os.path.isfile(legacy):
        urls |= _collect_http_urls_from_workbook(legacy)
    if os.path.isfile(pin_path):
        urls |= _collect_http_urls_from_workbook(pin_path)
    return urls


def _r2_object_key_from_url(url: str, public_bases: List[str]) -> Optional[str]:
    u = str(url or "").strip()
    if not u.lower().startswith("http"):
        return None
    u_clean = u.split("?")[0].split("#")[0]
    for base in public_bases:
        b = str(base or "").strip().rstrip("/")
        if not b:
            continue
        if u_clean.startswith(b + "/"):
            from urllib.parse import unquote

            return unquote(u_clean[len(b) + 1 :])
    from urllib.parse import unquote, urlparse

    path = unquote((urlparse(u_clean).path or "")).lstrip("/")
    if path.startswith(("midjourney_", "pinterest_")):
        return path
    return None


def _run_with_site_env(site_id: str, out_dir: str, fn: Callable[[], Any]) -> Any:
    with _site_config_lock:
        old_sid = os.environ.get("PINTEREST_SITE_ID")
        old_out = os.environ.get("PINTEREST_OUT_DIR")
        try:
            if site_id:
                os.environ["PINTEREST_SITE_ID"] = site_id
            else:
                os.environ.pop("PINTEREST_SITE_ID", None)
            if out_dir:
                os.environ["PINTEREST_OUT_DIR"] = out_dir
            else:
                os.environ.pop("PINTEREST_OUT_DIR", None)
            return fn()
        finally:
            if old_sid is not None:
                os.environ["PINTEREST_SITE_ID"] = old_sid
            else:
                os.environ.pop("PINTEREST_SITE_ID", None)
            if old_out is not None:
                os.environ["PINTEREST_OUT_DIR"] = old_out
            else:
                os.environ.pop("PINTEREST_OUT_DIR", None)


def _project_site_env(label: str) -> tuple:
    u = _unit_by_label(label)
    if not u:
        return "", ""
    e = u.get("env") or {}
    return (
        str(e.get("PINTEREST_SITE_ID", "") or "").strip(),
        str(e.get("PINTEREST_OUT_DIR", "") or "").strip(),
    )


def _cleanup_target_labels(project: str = "") -> List[str]:
    p = str(project or "").strip()
    if p:
        if not is_allowed_project_label(p):
            return []
        return [p]
    return flat_ui_labels()


def _preview_local_cleanup(project: str = "") -> Dict[str, Any]:
    labels = _cleanup_target_labels(project)
    output_files = 0
    output_bytes = 0
    cf_prune = 0
    per_project: List[Dict[str, Any]] = []
    site_ids: set = set()
    for label in labels:
        out_dir = all_out_name_for_label(label)
        sid, _ = _project_site_env(label)
        if sid:
            site_ids.add(sid)
        imgs = os.path.join(_project_output_dir(out_dir), "output_images")
        n, sz = _dir_size_and_count(imgs)
        output_files += n
        output_bytes += sz
        per_project.append({"label": label, "out_dir": out_dir, "output_images_files": n, "output_images_bytes": sz})
    cf_freed = 0
    if os.path.isdir(_CF_BUILDS_ROOT):
        for site_nm in os.listdir(_CF_BUILDS_ROOT):
            if site_ids and site_nm not in site_ids:
                continue
            site_path = os.path.join(_CF_BUILDS_ROOT, site_nm)
            if not os.path.isdir(site_path):
                continue
            builds = []
            for build_nm in os.listdir(site_path):
                build_path = os.path.join(site_path, build_nm)
                if os.path.isdir(build_path):
                    try:
                        builds.append((os.path.getmtime(build_path), build_path))
                    except OSError:
                        builds.append((0, build_path))
            builds.sort(key=lambda x: x[0], reverse=True)
            for i, (_mtime, build_path) in enumerate(builds):
                if i < 1:
                    continue
                cf_prune += 1
                _n, sz = _dir_size_and_count(build_path)
                cf_freed += sz
    return {
        "projects": per_project,
        "output_images_files": output_files,
        "output_images_bytes": output_bytes,
        "output_images_bytes_human": _format_bytes(output_bytes),
        "cf_builds_prune": cf_prune,
        "cf_builds_freed_estimate": _format_bytes(cf_freed),
    }


def _execute_local_cleanup(project: str = "", *, archive: bool = True) -> Dict[str, Any]:
    labels = _cleanup_target_labels(project)
    if not labels:
        return {"ok": False, "error": "Unknown project"}
    archive_session = _new_archive_session_dir() if archive else ""
    archive_rel = (
        os.path.relpath(archive_session, _APP_ROOT).replace("\\", "/") if archive_session else ""
    )
    results: List[Dict[str, Any]] = []
    site_ids: set = set()
    for label in labels:
        out_dir = all_out_name_for_label(label)
        sid, _ = _project_site_env(label)
        if sid:
            site_ids.add(sid)
        row: Dict[str, Any] = {"label": label, "out_dir": out_dir}
        try:
            if archive:
                row["archive"] = _archive_output_images_only(out_dir, archive_session)
            row["clear"] = _clear_output_images_dir(out_dir)
            row["status"] = "ok"
        except Exception as e:
            row["status"] = f"error: {e}"
        results.append(row)
    cf_stats = {"removed_dirs": 0, "kept_dirs": 0, "freed_human": _format_bytes(0)}
    if os.path.isdir(_CF_BUILDS_ROOT):
        total_removed = 0
        total_kept = 0
        total_freed = 0
        if site_ids:
            for sid in sorted(site_ids):
                st = _cleanup_cf_builds(keep_latest_per_site=1, site_id_filter=sid)
                total_removed += int(st.get("removed_dirs") or 0)
                total_kept += int(st.get("kept_dirs") or 0)
                total_freed += int(st.get("freed_bytes") or 0)
        else:
            st = _cleanup_cf_builds(keep_latest_per_site=1)
            total_removed = int(st.get("removed_dirs") or 0)
            total_kept = int(st.get("kept_dirs") or 0)
            total_freed = int(st.get("freed_bytes") or 0)
        cf_stats = {
            "removed_dirs": total_removed,
            "kept_dirs": total_kept,
            "freed_human": _format_bytes(total_freed),
        }
    _PROJECT_STATS_CACHE.clear()
    return {
        "ok": True,
        "archive_rel": archive_rel,
        "projects": results,
        "cf_builds": cf_stats,
    }


def _group_projects_by_r2_bucket(labels: List[str]) -> Dict[str, Dict[str, Any]]:
    grouped: Dict[str, Dict[str, Any]] = {}
    mod = _load_pipeline_a1_config()
    if mod is None:
        return grouped
    for label in labels:
        sid, out_dir = _project_site_env(label)
        if not sid:
            continue

        def _load_cfg(_sid=sid, _out=out_dir):
            return mod.get_r2_config(mod.load_keys())

        try:
            cfg = _run_with_site_env(sid, out_dir, _load_cfg)
        except Exception:
            continue
        bucket = str((cfg or {}).get("bucket") or "").strip()
        if not bucket:
            continue
        entry = grouped.setdefault(
            bucket,
            {
                "bucket": bucket,
                "public_base_url": str((cfg or {}).get("public_base_url") or "").strip(),
                "labels": [],
                "site_ids": [],
            },
        )
        entry["labels"].append(label)
        entry["site_ids"].append(sid)
        if not entry.get("public_base_url"):
            entry["public_base_url"] = str((cfg or {}).get("public_base_url") or "").strip()
    return grouped


def _preview_r2_cleanup(project: str = "") -> Dict[str, Any]:
    labels = _cleanup_target_labels(project)
    if not labels:
        return {"ok": False, "error": "Unknown project", "buckets": []}
    grouped = _group_projects_by_r2_bucket(labels)
    mod = _load_pipeline_a1_config()
    buckets_out: List[Dict[str, Any]] = []
    for bucket, meta in grouped.items():
        referenced_keys: set = set()
        public_bases: List[str] = []
        pub = str(meta.get("public_base_url") or "").strip()
        if pub:
            public_bases.append(pub)
        for label in meta.get("labels") or []:
            out_dir = all_out_name_for_label(label)
            for url in _collect_referenced_r2_urls_for_out_dir(out_dir):
                key = _r2_object_key_from_url(url, public_bases)
                if key:
                    referenced_keys.add(key)
        would_delete = 0
        would_keep = 0
        sample_delete: List[str] = []
        if mod is not None:
            sid = (meta.get("site_ids") or [""])[0]
            out_dir = all_out_name_for_label((meta.get("labels") or [""])[0])

            def _scan(_sid=sid, _out=out_dir, _bucket=bucket):
                nonlocal would_delete, would_keep
                client = mod.make_r2_client(mod.load_keys())
                for prefix in _R2_CLEANUP_PREFIXES:
                    token = None
                    while True:
                        kwargs = {"Bucket": _bucket, "Prefix": prefix, "MaxKeys": 1000}
                        if token:
                            kwargs["ContinuationToken"] = token
                        resp = client.list_objects_v2(**kwargs)
                        for obj in resp.get("Contents") or []:
                            key = str(obj.get("Key") or "")
                            if not key:
                                continue
                            if key in referenced_keys:
                                would_keep += 1
                            else:
                                would_delete += 1
                                if len(sample_delete) < 8:
                                    sample_delete.append(key)
                        if not resp.get("IsTruncated"):
                            break
                        token = resp.get("NextContinuationToken")

            try:
                _run_with_site_env(sid, out_dir, _scan)
            except Exception as e:
                buckets_out.append(
                    {
                        "bucket": bucket,
                        "labels": meta.get("labels") or [],
                        "error": str(e),
                        "referenced_keys": len(referenced_keys),
                        "would_delete": 0,
                        "would_keep": 0,
                        "sample_delete": [],
                    }
                )
                continue
        buckets_out.append(
            {
                "bucket": bucket,
                "labels": meta.get("labels") or [],
                "referenced_keys": len(referenced_keys),
                "would_delete": would_delete,
                "would_keep": would_keep,
                "sample_delete": sample_delete,
            }
        )
    return {"ok": True, "buckets": buckets_out}


def _execute_r2_cleanup(project: str = "", *, dry_run: bool = False) -> Dict[str, Any]:
    labels = _cleanup_target_labels(project)
    if not labels:
        return {"ok": False, "error": "Unknown project"}
    grouped = _group_projects_by_r2_bucket(labels)
    mod = _load_pipeline_a1_config()
    if mod is None:
        return {"ok": False, "error": "a1_config not found"}
    buckets_out: List[Dict[str, Any]] = []
    for bucket, meta in grouped.items():
        referenced_keys: set = set()
        public_bases: List[str] = []
        pub = str(meta.get("public_base_url") or "").strip()
        if pub:
            public_bases.append(pub)
        for label in meta.get("labels") or []:
            out_dir = all_out_name_for_label(label)
            for url in _collect_referenced_r2_urls_for_out_dir(out_dir):
                key = _r2_object_key_from_url(url, public_bases)
                if key:
                    referenced_keys.add(key)
        deleted = 0
        kept = 0
        errors: List[str] = []
        sid = (meta.get("site_ids") or [""])[0]
        out_dir = all_out_name_for_label((meta.get("labels") or [""])[0])

        def _delete(_sid=sid, _out=out_dir, _bucket=bucket):
            nonlocal deleted, kept
            client = mod.make_r2_client(mod.load_keys())
            for prefix in _R2_CLEANUP_PREFIXES:
                token = None
                while True:
                    kwargs = {"Bucket": _bucket, "Prefix": prefix, "MaxKeys": 1000}
                    if token:
                        kwargs["ContinuationToken"] = token
                    resp = client.list_objects_v2(**kwargs)
                    keys_batch: List[str] = []
                    for obj in resp.get("Contents") or []:
                        key = str(obj.get("Key") or "")
                        if not key:
                            continue
                        if key in referenced_keys:
                            kept += 1
                            continue
                        if dry_run:
                            deleted += 1
                        else:
                            keys_batch.append(key)
                    if keys_batch and not dry_run:
                        try:
                            client.delete_objects(
                                Bucket=_bucket,
                                Delete={"Objects": [{"Key": k} for k in keys_batch], "Quiet": True},
                            )
                            deleted += len(keys_batch)
                        except Exception as e:
                            errors.append(str(e))
                    if not resp.get("IsTruncated"):
                        break
                    token = resp.get("NextContinuationToken")

        try:
            _run_with_site_env(sid, out_dir, _delete)
        except Exception as e:
            buckets_out.append(
                {
                    "bucket": bucket,
                    "labels": meta.get("labels") or [],
                    "error": str(e),
                    "deleted": 0,
                    "kept": 0,
                }
            )
            continue
        buckets_out.append(
            {
                "bucket": bucket,
                "labels": meta.get("labels") or [],
                "referenced_keys": len(referenced_keys),
                "deleted": deleted,
                "kept": kept,
                "errors": errors,
                "dry_run": dry_run,
            }
        )
    return {"ok": True, "buckets": buckets_out, "dry_run": dry_run}


@app.route("/api/cleanup/preview")
def api_cleanup_preview():
    project = str(request.args.get("project") or "").strip()
    return jsonify(
        {
            "ok": True,
            "project": project or "all",
            "local": _preview_local_cleanup(project),
            "r2": _preview_r2_cleanup(project),
        }
    )


@app.route("/api/cleanup/local", methods=["POST"])
def api_cleanup_local():
    data = request.get_json(silent=True) or {}
    project = str(data.get("project") or request.form.get("project") or "").strip()
    result = _execute_local_cleanup(project, archive=True)
    if not result.get("ok"):
        return jsonify(result), 400
    return jsonify(result)


@app.route("/api/cleanup/r2-unused", methods=["POST"])
def api_cleanup_r2_unused():
    data = request.get_json(silent=True) or {}
    project = str(data.get("project") or request.form.get("project") or "").strip()
    confirm = str(data.get("confirm") or request.form.get("confirm") or "").strip()
    dry_run = str(data.get("dry_run") or request.form.get("dry_run") or "").lower() in (
        "1",
        "true",
        "yes",
    )
    if not dry_run and confirm != "DELETE R2":
        return jsonify({"ok": False, "error": "Confirmation required: send confirm=DELETE R2"}), 400
    result = _execute_r2_cleanup(project, dry_run=dry_run)
    if not result.get("ok"):
        return jsonify(result), 400
    return jsonify(result)


# ------------------------------------------------------------------
# 4) Bulk Titles Upload (OpenAI line-by-line, NO JSON from OpenAI)
# ------------------------------------------------------------------

def _load_pipeline_a1_config():
    folder = "A1-Pinterest_01"
    try:
        d = _load_sites_file_app()
        folder = str(d.get("pipeline_code_folder") or folder)
    except Exception:
        pass
    return _a1_config_module_for_pipeline_folder(folder)


def _load_pipeline_prompts(name: str) -> dict:
    mod = _load_pipeline_a1_config()
    if mod is not None and hasattr(mod, "load_prompts"):
        return mod.load_prompts(name)
    path = os.path.join(_APP_CONFIG, "prompts", f"{name}.json")
    if os.path.isfile(path):
        with open(path, "r", encoding="utf-8") as f:
            data = json.load(f)
        return data if isinstance(data, dict) else {}
    return {}


def filter_only_recipes_with_openai(raw_titles):
    """
    Keep ONLY the lines that are actual cooking recipe titles.
    - Do NOT add, rewrite, or translate any titles.
    - Output must be a subset of the INPUT lines (verbatim), one per line.
    - Remove empty lines, URLs, hashtags, @mentions, emails, and generic non-recipe lines.
    - Deduplicate while preserving original order.
    """
    # Quick pre-filter (regex) to remove obvious junk before sending to OpenAI
    import re
    junk_patterns = [
        r'https?://\S+',
        r'www\.\S+',
        r'\b(?:privacy|terms|policy|cookies?|login|signin|register|subscribe|newsletter|about|contact|homepage|copyright)\b',
        r'#\w+',
        r'@\w+',
        r'\b(?:ingredients?|instructions?|directions?|introduction|recipe overview|conclusion|notes?)\b',
        r'^\s*[\-\*\•\·]+\s*$',
        r'^\s*\d+\s*$',
        r'^\s*$'
    ]
    rx = re.compile("|".join(junk_patterns), re.IGNORECASE)

    cleaned = []
    seen = set()
    for line in raw_titles:
        s = line.strip()
        if not s:
            continue
        if rx.search(s):
            continue
        # Very short tokens (likely noise) -> skip
        if len(s) < 4:
            continue
        # Deduplicate (case-insensitive) while preserving original casing
        key = s.lower()
        if key in seen:
            continue
        seen.add(key)
        cleaned.append(s)

    if not cleaned:
        return []

    pp = _load_pipeline_prompts("app_title_filter")
    mod = _load_pipeline_a1_config()
    if mod is None:
        raise RuntimeError("a1_config not found — cannot load app_title_filter prompts")
    system_msg, user_payload = mod.format_app_title_filter(cleaned, "filter_only_recipes", prompts=pp)
    sec = pp.get("filter_only_recipes") or {}

    try:
        messages = [
            {"role": "system", "content": system_msg},
            {"role": "user", "content": user_payload},
        ]
        resp = chat_completion_with_retry(
            messages,
            model=str(sec.get("model") or "gpt-4o-mini"),
            temperature=float(sec.get("temperature", 0)),
        )
        assistant_msg = resp["choices"][0]["message"]["content"].strip()

        # Post-validate: keep only lines that existed in cleaned (verbatim match ignoring trailing spaces)
        valid_set = {c.strip(): True for c in cleaned}
        out = []
        for ln in assistant_msg.splitlines():
            ln2 = ln.strip()
            if not ln2:
                continue
            if ln2 in valid_set:
                out.append(ln2)
        # Final dedupe preserving order
        final = []
        seen2 = set()
        for s in out:
            if s.lower() not in seen2:
                seen2.add(s.lower())
                final.append(s)
        return final
    except Exception as e:
        print("OpenAI filter_only_recipes_with_openai error:", e)
        # Fallback: return the pre-cleaned list
        return cleaned
def filter_titles_with_openai(raw_titles):
    """
    Filter ONLY. Do not rewrite or generate new titles.
    - Input: list of lines (raw_titles)
    - Output: subset of those lines that are valid cooking recipe titles.
    Rules:
      * Keep original text exactly (no rephrasing, no capitalization changes).
      * Preserve the original order of the kept lines.
      * Do NOT add any new lines.
      * Remove lines that are clearly not recipes: URLs, hashtags, social handles, generic section headers
        (e.g., Introduction, Ingredients, Instructions, Conclusion), promotional/legal text (subscribe, privacy,
        login, terms, cookie), non-food content, or anything unrelated to a dish/recipe name.
      * Do NOT deduplicate similar titles; if the same title appears multiple times, keep all occurrences.
    """
    import re

    # Local quick pre-filter for obvious junk
    cleaned = []
    for line in raw_titles:
        t = line.strip()
        if not t:
            continue
        # Drop anything that looks like a link or handle
        if re.search(r'(https?://|www\.|\.com\b|\.net\b|\.org\b|@\w+|#\w+)', t, re.IGNORECASE):
            continue
        cleaned.append(t)

    if not cleaned:
        return []

    pp = _load_pipeline_prompts("app_title_filter")
    mod = _load_pipeline_a1_config()
    if mod is None:
        raise RuntimeError("a1_config not found — cannot load app_title_filter prompts")
    system_content, user_content = mod.format_app_title_filter(cleaned, "filter_titles", prompts=pp)
    sec = pp.get("filter_titles") or {}

    try:
        resp = chat_completion_with_retry(
            messages=[
                {"role": "system", "content": system_content},
                {"role": "user", "content": user_content},
            ],
            model=str(sec.get("model") or "gpt-4o-mini"),
            temperature=float(sec.get("temperature", 0)),
        )
        text = resp["choices"][0]["message"]["content"]
    except Exception as e:
        print("Error calling OpenAI API or parsing response:", e)
        # fallback: just do a conservative local filter with heuristics
        heuristics = []
        for t in cleaned:
            if re.search(r'(\brecipe\b|\bchicken\b|\bbeef\b|\bpasta\b|\bcake\b|\bbrownies\b|\bsalad\b|\bsoup\b|\bcasserole\b|\btacos\b|\bpancakes\b|\bair\s*fryer\b|\binstant\s*pot\b|\bcrock\s*pot\b)', t, re.IGNORECASE):
                heuristics.append(t)
        return heuristics if heuristics else cleaned

    # Parse lines; keep order; remove any accidental extras like bullet prefixes
    kept = []
    for line in text.splitlines():
        s = line.strip(" -• ")
        if not s:
            continue
        # Guard again against links
        if re.search(r'(https?://|www\.|\.com\b|\.net\b|\.org\b)', s, re.IGNORECASE):
            continue
        # Only keep if it existed in the cleaned input (exact match)
        if s in cleaned:
            kept.append(s)

    return kept
@app.route("/upload_titles", methods=["POST"])
def upload_titles():
    """
    2-step process:
      1) preview_only=1 -> use filter_titles_with_openai(...) but do NOT save to XLSX; return them for preview
      2) preview_only=0 -> expects final_titles (the user-edited list) to write to XLSX
    """
    preview_only = request.form.get("preview_only", "0")

    # ------------------ Step 1: PREVIEW ------------------
    if preview_only == "1":
        titles_text = request.form.get("titles", "")
        if not titles_text.strip():
            return jsonify({"status": "error", "message": "No titles provided."}), 400

        raw_titles = [line.strip() for line in titles_text.splitlines() if line.strip()]
        if not raw_titles:
            return jsonify({"status": "error", "message": "No valid titles."}), 400

        filtered_titles = filter_titles_with_openai(raw_titles)
        if not filtered_titles:
            return jsonify({"status": "error", "message": "OpenAI returned no valid recipe titles."}), 400

        return jsonify({
            "status": "preview",
            "message": "Preview of filtered titles (not written to XLSX).",
            "filtered_titles": filtered_titles
        })

    # ------------------ Step 2: CONFIRM & WRITE XLSX ------------------
    else:
        final_titles_json = request.form.get("final_titles", "")
        if not final_titles_json.strip():
            return jsonify({"status": "error", "message": "No final titles provided."}), 400

        try:
            final_titles = json.loads(final_titles_json)
        except:
            return jsonify({"status": "error", "message": "Invalid JSON in final_titles."}), 400

        if not isinstance(final_titles, list) or not final_titles:
            return jsonify({"status": "error", "message": "No valid final titles list."}), 400

        starts_folder = os.path.join(os.getcwd(), "STARTS")
        if not os.path.exists(starts_folder):
            return jsonify({"status": "error", "message": "STARTS folder not found."}), 400

        xlsx_files = [
            f for f in os.listdir(starts_folder)
            if f.lower().endswith(".xlsx") and not f.startswith("~$")
        ]
        if not xlsx_files:
            return jsonify({"status": "error", "message": "No XLSX files in STARTS."}), 400

        num_files = len(xlsx_files)

        def split_list(lst, n):
            k, m = divmod(len(lst), n)
            return [lst[i] for i in range(len(lst))],  # simplified; implement as needed

        # توزيع عادل على الملفات
        chunks = [[] for _ in range(num_files)]
        for i, title in enumerate(final_titles):
            chunks[i % num_files].append(title)

        file_counts = {}

        for file_name, chunk in zip(xlsx_files, chunks):
            file_path = os.path.join(starts_folder, file_name)
            try:
                wb = openpyxl.load_workbook(file_path)
                sheet = wb.active

                # Clear old data (row 2..end)
                for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row):
                    for cell in row:
                        cell.value = None

                # Insert the chunk
                for row_num, title in enumerate(chunk, start=2):
                    sheet.cell(row=row_num, column=1, value=title)

                wb.save(file_path)
                file_counts[file_name] = len(chunk)
            except Exception as e:
                return jsonify({"status": "error", "message": f"Error in {file_name}: {e}"}), 500

        return jsonify({
            "status": "success",
            "message": "Final titles uploaded to XLSX files.",
            "total": len(final_titles),
            "counts": file_counts
        })


# ------------------------------------------------------------------
# 5) Clear FAILED Rows  (UPDATED: supports 'statu' AND 'statu_ing')
# ------------------------------------------------------------------
@app.route("/clear_failed", methods=["POST"])
def clear_failed():
    """
    كيمسح أي صف فـ Recipes.xlsx (أو images.xlsx القديم) إذا:
      - 'statu' = FAILED أو خاوي
      - أو 'statu_ing' = FAILED أو خاوي
    إذا كاين غير واحد فيهم، خدام. إذا جوج ما كاينينش كيرجع رسالة مناسبة.
    """
    results = {}
    for project in flat_ui_labels():
        out_folder = all_out_name_for_label(project)
        file_path = _project_excel_path_by_out_dir(out_folder)

        if not os.path.exists(file_path):
            results[project] = "File not found"
            continue

        try:
            wb = openpyxl.load_workbook(file_path)
            sheet = wb.active

            # Locate columns (case-insensitive)
            status_col = None
            status_ing_col = None
            for col in range(1, sheet.max_column + 1):
                v = sheet.cell(row=1, column=col).value
                if not v:
                    continue
                name = str(v).strip().lower()
                if name == "statu":
                    status_col = col
                elif name == "statu_ing":
                    status_ing_col = col

            if not status_col and not status_ing_col:
                results[project] = "No 'statu' or 'statu_ing' column"
                continue

            def needs_delete(row_idx: int) -> bool:
                """
                True إذا خاص الصف يتحيد بناءً على واحد من الأعمدة المتوفّرين:
                - 'statu' أو 'statu_ing' : FAILED أو فارغ
                """

                def cell_val(col_idx):
                    if not col_idx:
                        return None
                    raw = sheet.cell(row=row_idx, column=col_idx).value
                    return str(raw).strip() if raw is not None else ""

                v_statu = cell_val(status_col)
                v_statu_ing = cell_val(status_ing_col)

                cond_statu = (v_statu == "" or (isinstance(v_statu, str) and v_statu.upper() == "FAILED"))
                cond_statu_ing = (
                            v_statu_ing == "" or (isinstance(v_statu_ing, str) and v_statu_ing.upper() == "FAILED"))

                # خدم بالـ OR: إذا شي واحد فيهم مطبق، تحيد الصف
                # إلا بغيتها AND بدّل لـ (cond_statu and cond_statu_ing)
                return cond_statu or cond_statu_ing

            removed_count = 0
            for row in range(sheet.max_row, 1, -1):
                if needs_delete(row):
                    sheet.delete_rows(row)
                    removed_count += 1

            wb.save(file_path)
            results[project] = f"Removed {removed_count} rows"
        except Exception as e:
            results[project] = f"Error: {e}"

    return jsonify({"status": "success", "results": results})


# ------------------------------------------------------------------
# 6) Manage Images (UI)
# ------------------------------------------------------------------
@app.route("/manage_images")
def manage_images():
    html = """
    <!DOCTYPE html>
    <html lang="en">
    <head>
      <meta charset="utf-8" />
      <title>Manage Images | Materio-Like UI</title>
      <!-- Bootstrap 5 + Boxicons -->
      <link href="https://cdn.jsdelivr.net/npm/bootstrap@5.0.2/dist/css/bootstrap.min.css" rel="stylesheet">
      <link href="https://unpkg.com/boxicons@2.1.4/css/boxicons.min.css" rel="stylesheet">
      <style>
        body { background-color: #f8f9fd; }
        .materio-sidebar { width: 250px; min-height: 100vh; background-color: #fff; border-right: 1px solid #eceef1; position: fixed; left: 0; top: 0; padding: 20px; }
        .materio-sidebar .sidebar-title { font-size: 1.3rem; font-weight: bold; margin-bottom: 1rem; display: flex; align-items: center; }
        .materio-sidebar .sidebar-title i { font-size: 1.5rem; margin-right: 10px; color: #7367F0; }
        .materio-sidebar ul { list-style: none; padding: 0; }
        .materio-sidebar ul li { margin: 10px 0; }
        .materio-sidebar ul li a { text-decoration: none; color: #626262; font-weight: 500; display: block; padding: 8px 10px; border-radius: 6px; }
        .materio-sidebar ul li a:hover { background-color: #7367F0; color: #fff; }
        .navbar-materio { margin-left: 270px; background-color: #fff; border-bottom: 1px solid #eceef1; height: 60px; display: flex; align-items: center; justify-content: space-between; padding: 0 20px; position: fixed; width: calc(100% - 270px); top: 0; z-index: 1000; }
        .navbar-materio h5 { margin: 0; }
        .materio-content { margin-left: 270px; padding: 20px; margin-top: 80px; }
        .card-image-table img { max-width: 60px; border-radius: 4px; }
        .table thead th { vertical-align: bottom; }
        .badge-row { font-size: 0.9rem; }
      </style>
    </head>
    <body>
      <div class="materio-sidebar">
        <div class="sidebar-title">
          <i class='bx bxs-component'></i>
          <span>AUTOMATION</span>
        </div>
        <ul>
          <li><a href="/"><i class='bx bx-home-alt'></i> Dashboard</a></li>
          <li><a href="/manage_images"><i class='bx bx-image'></i> Manage Images</a></li>
          <li><a href="/manage_articles"><i class='bx bx-file'></i> Manage Articles</a></li>
          <li><a href="#" onclick="history.back()"><i class='bx bx-left-arrow-alt'></i> Go Back</a></li>
        </ul>
      </div>

      <div class="navbar-materio">
        <h5>Manage Images</h5>
        <div>
          <i class='bx bx-bell' style="font-size: 20px; margin-right: 20px;'></i>
          <i class='bx bx-user-circle' style="font-size: 24px;"></i>
        </div>
      </div>

      <div class="materio-content">
        <div class="row row-cols-1 row-cols-md-2 g-4">
    """

    for project in flat_ui_labels():
        file_path = _project_excel_path_by_out_dir(all_out_name_for_label(project))
        if not os.path.exists(file_path):
            html += f"""
            <div class="col">
              <div class="card h-100 mb-4">
                <div class="card-body">
                  <h4 class="card-title">{project}</h4>
                  <p class="card-text text-danger">Recipes.xlsx not found.</p>
                </div>
              </div>
            </div>
            """
            continue

        wb = openpyxl.load_workbook(file_path)
        sheet = wb.active

        headers = {}
        for col in range(1, sheet.max_column + 1):
            val = sheet.cell(row=1, column=col).value
            if val and val.strip().lower() in ["image_1", "image_2", "image_3", "image_4"]:
                headers[val.strip().lower()] = col

        html += f"""
          <div class="col">
            <div class="card h-100 card-image-table">
              <div class="card-header bg-light">
                <strong>{project}</strong>
              </div>
              <div class="card-body p-0 table-responsive">
                <table class="table table-striped table-hover mb-0 align-middle">
                  <thead class="table-light">
                    <tr>
                      <th style="width: 60px;">Row</th>
                      <th>image_1</th>
                      <th>image_2</th>
                      <th>image_3</th>
                      <th>image_4</th>
                      <th style="width: 80px;">Action</th>
                    </tr>
                  </thead>
                  <tbody>
        """

        for row_idx in range(2, sheet.max_row + 1):
            image_tags = []
            for img_col_name in ["image_1", "image_2", "image_3", "image_4"]:
                col_index = headers.get(img_col_name)
                if col_index:
                    cell_val = sheet.cell(row=row_idx, column=col_index).value
                    if cell_val:
                        image_tags.append(
                            f"<img src='{cell_val}?r={row_idx}_{img_col_name}' alt='{cell_val}' />"
                        )
                    else:
                        image_tags.append("")
                else:
                    image_tags.append("")

            delete_form = f"""
            <form class="delete-form" action="/delete_image_row" method="post" style="display:inline-block;" onsubmit="return deleteImageRow(event, this);">
              <input type="hidden" name="project" value="{project}" />
              <input type="hidden" name="row_number" value="{row_idx}" />
              <button class="btn btn-sm btn-danger" type="submit">Del</button>
            </form>
            """

            html += f"""
            <tr>
              <td><span class="badge bg-secondary badge-row">{row_idx}</span></td>
              <td>{image_tags[0]}</td>
              <td>{image_tags[1]}</td>
              <td>{image_tags[2]}</td>
              <td>{image_tags[3]}</td>
              <td>{delete_form}</td>
            </tr>
            """

        html += f"""
              </tbody>
                </table>
              </div>
              <div class="card-footer"></div>
            </div>
          </div>
        """

    html += """
        </div> <!-- .row -->
      </div> <!-- .materio-content -->

      <script>
        function deleteImageRow(event, form) {
          event.preventDefault();
          fetch(form.action, {
            method: "POST",
            body: new FormData(form)
          })
          .then(response => response.text())
          .then(data => {
            let row = form.closest("tr");
            if(row) { row.remove(); }
          })
          .catch(error => { alert("Error deleting row: " + error); });
          return false;
        }
      </script>

      <script src="https://cdn.jsdelivr.net/npm/bootstrap@5.0.2/dist/js/bootstrap.bundle.min.js"></script>
    </body>
    </html>
    """
    return html


@app.route("/manage_articles")
def manage_articles():
    cols = [
        "Main Keyword", "Recipe", "image_1", "image_2", "image_3", "image_4",
        "article", "pinterest_title", "pinterest_description",
        "pinterest_image_short_title", "category", "pinterest_image"
    ]

    html = """
    <!DOCTYPE html>
    <html lang="en">
    <head>
      <meta charset="utf-8" />
      <title>Manage Articles | Materio-Like UI</title>
      <link href="https://cdn.jsdelivr.net/npm/bootstrap@5.0.2/dist/css/bootstrap.min.css" rel="stylesheet">
      <link href="https://unpkg.com/boxicons@2.1.4/css/boxicons.min.css" rel="stylesheet">
      <style>
        body { background-color: #f8f9fd; }
        .materio-sidebar { width: 250px; min-height: 100vh; background: #fff; border-right: 1px solid #eceef1; position: fixed; padding: 20px; }
        .navbar-materio { margin-left:270px; background:#fff; border-bottom:1px solid #eceef1; height:60px; display:flex; align-items:center; justify-content:space-between; padding:0 20px; position:fixed; width:calc(100% - 270px); top:0; }
        .materio-content { margin-left:270px; padding:20px; margin-top:80px; }
        table td, table th { white-space: nowrap; overflow: hidden; text-overflow: ellipsis; max-width: 150px; }
      </style>
    </head>
    <body>
      <div class="materio-sidebar">
        <div class="sidebar-title"><i class='bx bxs-component'></i> <strong>AUTOMATION</strong></div>
        <ul class="list-unstyled">
          <li><a href="/"><i class='bx bx-home-alt'></i> Dashboard</a></li>
          <li><a href="/manage_images"><i class='bx bx-image'></i> Manage Images</a></li>
          <li><a href="/manage_articles"><i class='bx bx-file'></i> Manage Articles</a></li>
        </ul>
      </div>
      <div class="navbar-materio">
        <h5>Manage Articles</h5>
        <div><i class='bx bx-user-circle' style="font-size:24px;"></i></div>
      </div>
      <div class="materio-content">
        <div class="mb-3">
          <form action="/delete_all_article_blanks" method="post">
            <button class="btn btn-danger">Remove All Blank Rows from All Projects</button>
          </form>
        </div>
        <div class="row row-cols-1 g-4">
    """

    for project in flat_ui_labels():
        path = os.path.join(
            os.getcwd(), "ALL", all_out_name_for_label(project), "ARTICLE.xlsx"
        )
        if not os.path.exists(path):
            html += f"""
            <div class="col">
              <div class="card h-100">
                <div class="card-body">
                  <h4 class="card-title">{project}</h4>
                  <p class="text-danger">ARTICLE.xlsx not found.</p>
                </div>
              </div>
            </div>
            """
            continue

        wb = openpyxl.load_workbook(path, data_only=True)
        sheet = wb.active

        header_map = {}
        for c in range(1, sheet.max_column + 1):
            val = sheet.cell(row=1, column=c).value
            if val and val.strip() in cols:
                header_map[val.strip()] = c

        html += f"""
          <div class="col">
            <div class="card h-100">
              <div class="card-header bg-light"><strong>{project}</strong></div>
              <div class="card-body p-0 table-responsive">
                <table class="table table-hover mb-0">
                  <thead class="table-light"><tr>
        """
        for col in cols:
            html += f"<th title='{col}'>{col}</th>"
        html += "</tr></thead><tbody>"

        for r in range(2, sheet.max_row + 1):
            vals = []
            empty_found = False
            for col in cols:
                col_idx = header_map.get(col)
                v = sheet.cell(row=r, column=col_idx).value if col_idx else None
                if v is None or (isinstance(v, str) and not v.strip()):
                    empty_found = True
                    vals.append("")
                else:
                    vals.append(v)
            row_class = "table-danger" if empty_found else ""
            html += f"<tr class='{row_class}'>"
            for v in vals:
                html += f"<td title='{v}'>{v}</td>"
            html += "</tr>"

        html += f"""
                  </tbody>
                </table>
              </div>
              <div class="card-footer">
                <form action="/delete_article_blanks" method="post" style="display:inline-block;">
                  <input type="hidden" name="project" value="{project}" />
                  <button class="btn btn-sm btn-danger">Remove Blank Rows</button>
                </form>
              </div>
            </div>
          </div>
        """

    html += """
        </div>
      </div>
      <script src="https://cdn.jsdelivr.net/npm/bootstrap@5.0.2/dist/js/bootstrap.bundle.min.js"></script>
    </body>
    </html>
    """
    return html


@app.route("/delete_article_blanks", methods=["POST"])
def delete_article_blanks():
    project = request.form.get("project", "")
    cols = [
        "Main Keyword", "Recipe", "image_1", "image_2", "image_3", "image_4",
        "article", "pinterest_title", "pinterest_description",
        "pinterest_image_short_title", "category", "pinterest_image"
    ]
    if not is_allowed_project_label(project):
        return f"<h1>Invalid project: {project}</h1><a href='/manage_articles'>Back</a>"
    path = os.path.join(
        os.getcwd(), "ALL", all_out_name_for_label(project), "ARTICLE.xlsx"
    )
    if not os.path.exists(path):
        return f"<h1>ARTICLE.xlsx not found for {project}</h1><a href='/manage_articles'>Back</a>"

    wb = openpyxl.load_workbook(path)
    sheet = wb.active
    header_map = {}
    for c in range(1, sheet.max_column + 1):
        val = sheet.cell(row=1, column=c).value
        if val and val.strip() in cols:
            header_map[val.strip()] = c

    removed = 0
    for row_idx in range(sheet.max_row, 1, -1):
        if any(
                (sheet.cell(row=row_idx, column=col_idx).value is None or
                 (isinstance(sheet.cell(row=row_idx, column=col_idx).value, str) and not sheet.cell(row=row_idx,
                                                                                                    column=col_idx).value.strip()))
                for col_idx in header_map.values()
        ):
            sheet.delete_rows(row_idx)
            removed += 1
    wb.save(path)
    return f"<h1>Removed {removed} blank rows from {project}</h1><a href='/manage_articles'>Back</a>"


@app.route("/delete_all_article_blanks", methods=["POST"])
def delete_all_article_blanks():
    cols = [
        "Main Keyword", "Recipe", "image_1", "image_2", "image_3", "image_4",
        "article", "pinterest_title", "pinterest_description",
        "pinterest_image_short_title", "category", "pinterest_image"
    ]
    summary = []
    for project in flat_ui_labels():
        path = os.path.join(
            os.getcwd(), "ALL", all_out_name_for_label(project), "ARTICLE.xlsx"
        )
        if not os.path.exists(path):
            summary.append(f"{project}: file not found")
            continue
        wb = openpyxl.load_workbook(path)
        sheet = wb.active
        header_map = {}
        for c in range(1, sheet.max_column + 1):
            val = sheet.cell(row=1, column=c).value
            if val and val.strip() in cols:
                header_map[val.strip()] = c
        removed = 0
        for row_idx in range(sheet.max_row, 1, -1):
            if any(
                    (sheet.cell(row=row_idx, column=col_idx).value is None or
                     (isinstance(sheet.cell(row=row_idx, column=col_idx).value, str) and not sheet.cell(row=row_idx,
                                                                                                        column=col_idx).value.strip()))
                    for col_idx in header_map.values()
            ):
                sheet.delete_rows(row_idx)
                removed += 1
        wb.save(path)
        summary.append(f"{project}: {removed} removed")
    result = "<h1>Bulk Blank Removal Summary</h1><ul>" + "".join(
        f"<li>{s}</li>" for s in summary) + "</ul><a href='/manage_articles'>Back</a>"
    return result


# -------------------- 7) Delete Image Row --------------------
@app.route("/delete_image_row", methods=["POST"])
def delete_image_row():
    project = request.form.get("project", "")
    row_number_str = request.form.get("row_number", "")

    if not project or not row_number_str:
        return "<h1>Missing project/row_number</h1><a href='/manage_images'>Back</a>"

    if not is_allowed_project_label(project):
        return f"<h1>Invalid project: {project}</h1><a href='/manage_images'>Back</a>"

    try:
        row_number = int(row_number_str)
    except:
        return "<h1>Invalid row_number</h1><a href='/manage_images'>Back</a>"

    file_path = _project_excel_path_by_out_dir(all_out_name_for_label(project))
    if not os.path.exists(file_path):
        return f"<h1>Recipes.xlsx not found for {project}</h1><a href='/manage_images'>Back</a>"

    wb = openpyxl.load_workbook(file_path)
    sheet = wb.active

    if 2 <= row_number <= sheet.max_row:
        sheet.delete_rows(row_number, 1)
        wb.save(file_path)
        return f"<h1>Row {row_number} deleted from {project}'s Recipes.xlsx</h1>"
    else:
        return f"<h1>Row {row_number} out of range</h1>"


# -------------------- 8) Main Dashboard Page --------------------
@app.route("/")
def index():
    _ensure_project_output_files_for_sites()
    starts_folder = os.path.join(os.getcwd(), "STARTS")
    if os.path.exists(starts_folder):
        xlsx_files = [f for f in os.listdir(starts_folder) if f.lower().endswith(".xlsx") and not f.startswith("~$")]
        num_xlsx = len(xlsx_files)
    else:
        num_xlsx = 0

    _units = flat_run_units()
    project_folders_json = _json_for_inline_script([u["log_id"] for u in _units])
    project_units_json = _json_for_inline_script(
        [{"log_id": u["log_id"], "label": u["label"]} for u in _units]
    )
    step_columns_meta_json = _json_for_inline_script(_step_columns_meta_for_ui())

    log_boxes = ""
    for u in _units:
        lid = u["log_id"]
        title = u["label"]
        lidj = json.dumps(lid)
        titlej = json.dumps(title)
        log_boxes += f"""
          <div class="col-lg-6 mb-4">
            <div class="card h-100">
              <div class="card-header d-flex justify-content-between align-items-start flex-wrap gap-1">
                <div class="me-2 flex-grow-1">
                  <h6 class="mb-0 text-secondary">{title} Log</h6>
                  <div id="stats_{lid}" class="small text-muted project-stats-line">Loading stats...</div>
                </div>
                <button type="button" class="btn btn-sm btn-outline-info" title="Edit this project in sites.json (tabbed: WordPress, API, Start, a2, pipeline, …)" onclick='showSiteConfigInfo({titlej})'>Info</button>
              </div>
              <div class="card-body overflow-auto" style="height:200px;" id="log_{lid}"></div>
              <div class="card-footer">
                <div class="theme-row d-flex align-items-center flex-wrap gap-2 mb-2" data-project={titlej}>
                  <label class="small text-muted mb-0 me-1">Theme:</label>
                  <select class="form-select form-select-sm theme-picker" style="max-width:200px;" data-project={titlej} data-log-id={lidj}>
                    <option value="">— none —</option>
                  </select>
                  <input type="text" class="form-control form-control-sm cf-project-input" placeholder="CF Pages project name (optional override)" style="max-width:260px;" data-project={titlej} />
                  <button type="button" class="btn btn-sm btn-outline-secondary theme-save-btn" data-project={titlej} title="Save theme + CF project name to sites.json">Save</button>
                  <span class="theme-status small text-muted" data-project={titlej}></span>
                </div>
                <button class="btn btn-sm btn-primary project-action" data-action="start" data-step-key="START" onclick='startProjectAction({lidj}, {titlej}, "start", this)'>START</button>
                <button class="btn btn-sm btn-secondary project-action" data-action="json" data-step-key="JSON" onclick='startProjectAction({lidj}, {titlej}, "json", this)'>JSON</button>
                <button class="btn btn-sm btn-warning project-action" data-action="prompt" data-step-key="PROMPT" onclick='startProjectAction({lidj}, {titlej}, "prompt", this)'>PROMPT</button>
                <button class="btn btn-sm btn-info project-action" data-action="imagine" data-step-key="IMAGINE" onclick='startProjectAction({lidj}, {titlej}, "imagine", this)'>IMAGINE</button>
                <button class="btn btn-sm btn-success project-action" data-action="article" data-step-key="ARTICLE" onclick='startProjectAction({lidj}, {titlej}, "article", this)'>ARTICLE</button>
                <button class="btn btn-sm btn-dark project-action" data-action="pin_data" data-step-key="PIN DATA" onclick='startProjectAction({lidj}, {titlej}, "pin_data", this)'>PIN DATA</button>
                <button class="btn btn-sm btn-dark project-action" data-action="pin_image" data-step-key="PIN IMAGE" onclick='startProjectAction({lidj}, {titlej}, "pin_image", this)'>PIN IMAGE</button>
                <button class="btn btn-sm btn-dark project-action" data-action="pin_image_html" data-step-key="PIN IMAGE HTML" onclick='startProjectAction({lidj}, {titlej}, "pin_image_html", this)'>PIN IMAGE HTML</button>
                <button class="btn btn-sm btn-dark project-action" data-action="wp_upload" data-step-key="WP UPLOAD" onclick='startProjectAction({lidj}, {titlej}, "wp_upload", this)'>WP UPLOAD</button>
                <button class="btn btn-sm btn-dark project-action" data-action="pin_bulk" data-step-key="PIN BULK" onclick='startProjectAction({lidj}, {titlej}, "pin_bulk", this)'>PIN BULK</button>
                <button class="btn btn-sm cf-upload-btn project-action" data-action="cf_upload" data-step-key="CF UPLOAD" data-project={titlej} disabled title="Select a theme and set Cloudflare Pages project name to enable" onclick='startProjectAction({lidj}, {titlej}, "cf_upload", this)' style="background:#f48120;color:#fff;border:1px solid #d96e10;">CF UPLOAD</button>
                <button class="btn btn-sm btn-outline-warning cf-versions-btn" data-project={titlej} data-log-id={lidj} disabled title="Show recent Cloudflare Pages deployments" onclick='showCfVersions({titlej}, {lidj})'>Versions</button>
                <button class="btn btn-sm btn-outline-warning" data-step-key="CLEANUP LOCAL PROJECT" onclick='confirmCleanupLocal({titlej})'>Cleanup</button>
                <button class="btn btn-sm btn-outline-danger project-clear-log" data-step-key="CLEAR PROJECT LOG" onclick='clearProjectLog({lidj})'>CLEAR LOG</button>
              </div>
            </div>
          </div>
        """

    return f"""
    <!DOCTYPE html>
    <html lang="en">
    <head>
      <meta charset="utf-8" />
      <title>Automation Pinterest</title>
      <link href="https://cdn.jsdelivr.net/npm/bootstrap@5.0.2/dist/css/bootstrap.min.css" rel="stylesheet">
      <link href="https://unpkg.com/boxicons@2.1.4/css/boxicons.min.css" rel="stylesheet">
      <style>
        body {{
          background-color: #f8f9fd;
        }}
        .materio-sidebar {{
          width: 250px;
          min-height: 100vh;
          background-color: #fff;
          border-right: 1px solid #eceef1;
          position: fixed;
          left: 0;
          top: 0;
          padding: 20px;
        }}
        .materio-sidebar .sidebar-title {{
          font-size: 1.3rem;
          font-weight: bold;
          margin-bottom: 1rem;
          display: flex;
          align-items: center;
        }}
        .materio-sidebar .sidebar-title i {{
          font-size: 1.5rem;
          margin-right: 10px;
          color: #7367F0;
        }}
        .materio-sidebar ul {{
          list-style: none;
          padding: 0;
        }}
        .materio-sidebar ul li {{
          margin: 10px 0;
        }}
        .materio-sidebar ul li a {{
          text-decoration: none;
          color: #626262;
          font-weight: 500;
          display: block;
          padding: 8px 10px;
          border-radius: 6px;
        }}
        .materio-sidebar ul li a:hover {{
          background-color: #7367F0;
          color: #fff;
        }}
        .navbar-materio {{
          margin-left: 270px;
          background-color: #fff;
          border-bottom: 1px solid #eceef1;
          height: 60px;
          display: flex;
          align-items: center;
          justify-content: space-between;
          padding: 0 20px;
          position: fixed;
          width: calc(100% - 270px);
          top: 0;
          z-index: 1000;
        }}
        .navbar-materio h5 {{ margin: 0; }}
        .materio-content {{ margin-left: 270px; padding: 20px; margin-top: 80px; }}
        .number-button {{
          display: inline-block;
          position: relative;
          margin: 5px;
          padding: 10px 20px 10px 50px;
          background-color: #7367F0;
          color: #fff;
          border: none;
          border-radius: 25px;
          cursor: pointer;
          font-size: 14px;
          transition: background-color 0.3s;
          text-decoration: none;
        }}
        .number-button::before {{
          content: attr(data-number);
          position: absolute;
          left: 10px;
          top: 50%;
          transform: translateY(-50%);
          width: 30px;
          height: 30px;
          background-color: #fff;
          color: #7367F0;
          border-radius: 50%;
          display: flex;
          align-items: center;
          justify-content: center;
          font-weight: bold;
        }}
        .number-button:hover {{ background-color: #5b51db; }}
        .number-button.active {{ background-color: yellow; color: #000; }}
        .number-button.finished {{ background-color: #28c76f; }}
        .number-button.partial {{ background-color: #ff9f43; color: #fff; }}
        @keyframes runningGlowPulse {{
          0% {{ box-shadow: 0 0 0 0 rgba(255, 193, 7, 0.65); filter: brightness(1); }}
          50% {{ box-shadow: 0 0 0 8px rgba(255, 193, 7, 0.10); filter: brightness(1.08); }}
          100% {{ box-shadow: 0 0 0 0 rgba(255, 193, 7, 0.00); filter: brightness(1); }}
        }}
        .running-glow {{
          animation: runningGlowPulse 1.2s ease-in-out infinite;
          border-color: #ffc107 !important;
        }}
        .tooltip.step-col-tooltip .tooltip-inner {{
          max-width: min(480px, 92vw);
          text-align: left;
          white-space: normal;
        }}
        .card-footer .project-action {{
          margin: 0 0.15rem 0.25rem 0;
        }}
        #previewTable td {{ padding: 5px; vertical-align: middle; }}
        #previewTable button {{ margin-left: 10px; }}
        #siteConfigModal .modal-dialog {{ max-width: min(1140px, 96vw); margin-left: auto; margin-right: auto; }}
        #siteConfigModal .nav-tabs .nav-link {{ font-weight: 600; font-size: 0.9rem; padding: 0.45rem 0.7rem; white-space: nowrap; }}
        #siteConfigModal .site-editor-top-tabs {{ border-bottom: 0; flex-wrap: nowrap; }}
        #siteConfigModal .site-editor-form-tab-content {{ min-height: 12rem; }}
        #siteConfigModal #siteEditorFormPane label {{ font-size: 0.8rem; margin-bottom: 0.1rem; }}
        .project-stats-line {{
          max-width: 100%;
          margin-top: 2px;
          white-space: normal;
          overflow-wrap: anywhere;
          word-break: break-word;
          line-height: 1.25;
          display: flex;
          flex-wrap: wrap;
          gap: 6px;
          align-items: center;
        }}
        .stat-chip {{
          display: inline-flex;
          align-items: center;
          gap: 6px;
          padding: 2px 8px;
          border-radius: 999px;
          font-size: 11px;
          font-weight: 600;
          border: 1px solid transparent;
        }}
        .stat-chip i {{ font-size: 14px; line-height: 1; }}
        .stat-chip-good {{
          background: #28c76f;
          color: #fff;
        }}
        .stat-chip-bad {{
          background: #ea5455;
          color: #fff;
        }}
        .stat-chip-neutral {{
          background: #f1f3f5;
          color: #5c6670;
          border-color: #d6dbe1;
        }}
        .stat-chip-name {{
          max-width: 180px;
          overflow: hidden;
          text-overflow: ellipsis;
          white-space: nowrap;
        }}
        .stat-chip-open {{ color: inherit !important; }}
        .stat-chip-open:hover {{ opacity: 0.85; }}
        .stat-step-group {{
          width: 100%;
          margin-top: 2px;
          padding-top: 2px;
          border-top: 1px dashed #e3e6eb;
        }}
        .stat-step-title {{
          font-size: 10px;
          font-weight: 700;
          color: #6c757d;
          margin-right: 6px;
          text-transform: uppercase;
        }}
      </style>
      <script>
        var source = null;
        var projectFolders = {project_folders_json};
        var projectUnits = {project_units_json};
        var stepColumnsMeta = {step_columns_meta_json};

        function _stepMeta(stepKey) {{
          if (!stepColumnsMeta || !stepColumnsMeta.steps) return null;
          return stepColumnsMeta.steps[stepKey] || null;
        }}

        function _columnsForCascadeClear(stepKey) {{
          var order = (stepColumnsMeta && stepColumnsMeta.order) ? stepColumnsMeta.order.slice() : _clearCascadeFrom(stepKey);
          var idx = order.indexOf(String(stepKey || "").trim().toUpperCase());
          if (idx < 0) return {{ steps: [], columns: [] }};
          var steps = order.slice(idx);
          var seen = {{}};
          var cols = [];
          steps.forEach(function(st) {{
            var m = _stepMeta(st);
            var arr = (m && m.clears) ? m.clears : [];
            arr.forEach(function(c) {{
              var k = String(c || "");
              if (!k || seen[k]) return;
              seen[k] = true;
              cols.push(k);
            }});
          }});
          return {{ steps: steps, columns: cols }};
        }}

        function _runStepTooltip(stepKey) {{
          var sk = String(stepKey || "").trim().toUpperCase();
          var special = stepColumnsMeta && stepColumnsMeta.special && stepColumnsMeta.special[sk];
          if (special && special.run) return special.run;
          var m = _stepMeta(sk);
          if (!m || !m.fills || !m.fills.length) return "Run " + sk;
          return "Fills: " + m.fills.join(", ");
        }}

        function _clearStepTooltip(stepKey) {{
          var sk = String(stepKey || "").trim().toUpperCase();
          var special = stepColumnsMeta && stepColumnsMeta.special && stepColumnsMeta.special[sk];
          if (special && special.clear) return special.clear;
          var casc = _columnsForCascadeClear(sk);
          if (!casc.columns.length) return "Clear " + sk + " and dependent steps";
          return "Clears Recipes.xlsx (cascade " + casc.steps.join(" → ") + "): " + casc.columns.join(", ");
        }}

        function _attachStepTooltip(btn, text) {{
          if (!btn || text == null) return;
          var t = String(text);
          var html = t.replace(/\\n/g, "<br>");
          btn.setAttribute("title", t);
          btn.setAttribute("data-bs-toggle", "tooltip");
          btn.setAttribute("data-bs-placement", "top");
          btn.setAttribute("data-bs-html", "true");
          if (typeof bootstrap !== "undefined" && bootstrap.Tooltip) {{
            try {{
              var old = bootstrap.Tooltip.getInstance(btn);
              if (old) old.dispose();
              new bootstrap.Tooltip(btn, {{
                customClass: "step-col-tooltip",
                trigger: "hover focus",
                html: true,
                title: html
              }});
            }} catch (e) {{}}
          }}
        }}

        function applyStepButtonTooltips() {{
          // Per-project card footer (all buttons with data-step-key)
          document.querySelectorAll(".card-footer [data-step-key]").forEach(function(btn) {{
            _attachStepTooltip(btn, _runStepTooltip(btn.getAttribute("data-step-key")));
          }});
          document.querySelectorAll(".card-footer button[onclick*='clearProjectLog']").forEach(function(btn) {{
            if (btn.getAttribute("data-step-key")) return;
            btn.setAttribute("data-step-key", "CLEAR PROJECT LOG");
            _attachStepTooltip(btn, _runStepTooltip("CLEAR PROJECT LOG"));
          }});

          // Global Actions — run buttons
          document.querySelectorAll("[data-step-key]").forEach(function(btn) {{
            if (btn.closest(".card-footer")) return;
            var sk = btn.getAttribute("data-step-key");
            if (sk) _attachStepTooltip(btn, _runStepTooltip(sk));
          }});
          document.querySelectorAll(".number-button[data-step]").forEach(function(btn) {{
            if (btn.getAttribute("data-step-key")) return;
            var num = btn.getAttribute("data-number");
            if (num === "11b") {{
              btn.setAttribute("data-step-key", "PIN IMAGE HTML");
              _attachStepTooltip(btn, _runStepTooltip("PIN IMAGE HTML"));
              return;
            }}
            if (num === "AUTO") {{
              btn.setAttribute("data-step-key", "AUTO SAFE");
              _attachStepTooltip(btn, _runStepTooltip("AUTO SAFE"));
              return;
            }}
            if (btn.getAttribute("onclick") && btn.getAttribute("onclick").indexOf("clearAllLogs") >= 0) {{
              btn.setAttribute("data-step-key", "CLEAR ALL LOGS");
              _attachStepTooltip(btn, _runStepTooltip("CLEAR ALL LOGS"));
              return;
            }}
            if (btn.getAttribute("onclick") && btn.getAttribute("onclick").indexOf("clearFailed") >= 0) {{
              btn.setAttribute("data-step-key", "CLEAR IMAGINE");
              _attachStepTooltip(btn, _runStepTooltip("CLEAR IMAGINE"));
              return;
            }}
            var step = btn.getAttribute("data-step");
            if (step) {{
              btn.setAttribute("data-step-key", step);
              _attachStepTooltip(btn, _runStepTooltip(step));
            }}
          }});
        }}
        var numXlsx = {num_xlsx};
        var _projectLogHistory = {{}};
        var STATS_REFRESH_MS = 30000;
        var _statsLastRefreshMsByLog = {{}};
        var _projectRunStartedMs = {{}};
        var _statsTickerId = null;
        var _statsAutoRefreshId = null;

        function _logStorageKey(logId) {{
          return "pinterest_log_" + String(logId || "");
        }}

        function _persistLog(logId) {{
          try {{
            localStorage.setItem(_logStorageKey(logId), _projectLogHistory[logId] || "");
          }} catch (e) {{}}
        }}

        function _appendProjectLog(logId, line, isFinished) {{
          if (!logId) return;
          var htmlLine = isFinished
            ? "<span style='color: green;'>" + String(line || "") + "</span><br>"
            : String(line || "") + "<br>";
          _projectLogHistory[logId] = String(_projectLogHistory[logId] || "") + htmlLine;
          var logDiv = document.getElementById("log_" + logId);
          if (logDiv) {{
            logDiv.innerHTML = _projectLogHistory[logId];
            logDiv.scrollTop = logDiv.scrollHeight;
          }}
          _persistLog(logId);
        }}

        function clearProjectLog(logId) {{
          if (!logId) return;
          _projectLogHistory[logId] = "";
          var logDiv = document.getElementById("log_" + logId);
          if (logDiv) logDiv.innerHTML = "";
          try {{ localStorage.removeItem(_logStorageKey(logId)); }} catch (e) {{}}
        }}

        function clearAllLogs() {{
          (projectFolders || []).forEach(function(logId) {{
            clearProjectLog(logId);
          }});
        }}

        function confirmDeleteAll() {{
          var ok = window.confirm(
            "⚠️ DANGEROUS ACTION — Delete ALL\\n\\n" +
            "For EVERY project this will:\\n" +
            "1) Archive Recipes.xlsx, Pin_01.xlsx, and output_images/ to ALL/archive/<date-time>/\\n" +
            "2) Clear pipeline columns in Recipes.xlsx (Title, Recipe, Generated At stay)\\n" +
            "3) Empty output_images/ folders\\n" +
            "4) Remove Pin_01.xlsx\\n\\n" +
            "This cannot be undone from the dashboard. Continue?"
          );
          if (!ok) return;
          var form = document.getElementById("deleteAllForm");
          if (form) form.submit();
        }}

        function _cleanupPreviewUrl(projectLabel) {{
          var q = projectLabel ? ("?project=" + encodeURIComponent(projectLabel)) : "";
          return "/api/cleanup/preview" + q;
        }}

        async function confirmCleanupLocal(projectLabel) {{
          var scope = projectLabel ? ("project " + projectLabel) : "ALL projects";
          try {{
            var prev = await fetch(_cleanupPreviewUrl(projectLabel)).then(function(r) {{ return r.json(); }});
            if (!prev || !prev.ok) {{
              alert("Preview failed.");
              return;
            }}
            var loc = prev.local || {{}};
            var msg = "Cleanup LOCAL for " + scope + ":\\n\\n";
            msg += "• output_images: " + (loc.output_images_files || 0) + " files (" + (loc.output_images_bytes_human || "0 B") + ")\\n";
            msg += "• Old CF builds to remove: " + (loc.cf_builds_prune || 0) + " (~" + (loc.cf_builds_freed_estimate || "0 B") + ")\\n\\n";
            msg += "Keeps Recipes.xlsx, Pin_01.xlsx, and all R2 images.\\n";
            msg += "Pin JPGs are archived first under ALL/archive/<date>/.\\n\\nContinue?";
            if (!window.confirm(msg)) return;
            var res = await fetch("/api/cleanup/local", {{
              method: "POST",
              headers: {{"Content-Type": "application/json"}},
              body: JSON.stringify({{project: projectLabel || ""}})
            }}).then(function(r) {{ return r.json(); }});
            if (!res.ok) {{
              alert("Cleanup failed: " + (res.error || "unknown"));
              return;
            }}
            var lines = ["Cleanup local done."];
            if (res.archive_rel) lines.push("Archive: " + res.archive_rel);
            if (res.cf_builds) {{
              lines.push("CF builds removed: " + (res.cf_builds.removed_dirs || 0) + " (~" + (res.cf_builds.freed_human || "0 B") + ")");
            }}
            alert(lines.join("\\n"));
          }} catch (e) {{
            alert("Cleanup error: " + e);
          }}
        }}

        async function confirmCleanupR2(projectLabel) {{
          var scope = projectLabel ? ("project " + projectLabel) : "ALL projects";
          try {{
            var prev = await fetch(_cleanupPreviewUrl(projectLabel)).then(function(r) {{ return r.json(); }});
            if (!prev || !prev.ok) {{
              alert("Preview failed.");
              return;
            }}
            var r2 = prev.r2 || {{}};
            var buckets = r2.buckets || [];
            if (!buckets.length) {{
              alert("No R2 buckets configured for " + scope + ".");
              return;
            }}
            var msg = "Cleanup R2 UNUSED for " + scope + ":\\n\\n";
            buckets.forEach(function(b) {{
              if (b.error) {{
                msg += "• Bucket " + b.bucket + ": ERROR — " + b.error + "\\n";
                return;
              }}
              msg += "• Bucket " + b.bucket + " (" + (b.labels || []).join(", ") + "):\\n";
              msg += "  keep " + (b.would_keep || 0) + ", delete " + (b.would_delete || 0) + " unreferenced\\n";
            }});
            msg += "\\nURLs still in Recipes.xlsx / Pin_01.xlsx are kept.\\n";
            msg += "WordPress / live sites must not need deleted images.\\n\\nContinue?";
            if (!window.confirm(msg)) return;
            var typed = window.prompt('Type DELETE R2 to confirm permanent R2 deletion');
            if (typed !== "DELETE R2") return;
            var res = await fetch("/api/cleanup/r2-unused", {{
              method: "POST",
              headers: {{"Content-Type": "application/json"}},
              body: JSON.stringify({{project: projectLabel || "", confirm: "DELETE R2"}})
            }}).then(function(r) {{ return r.json(); }});
            if (!res.ok) {{
              alert("R2 cleanup failed: " + (res.error || "unknown"));
              return;
            }}
            var lines = ["R2 cleanup done."];
            (res.buckets || []).forEach(function(b) {{
              if (b.error) {{
                lines.push(b.bucket + ": ERROR — " + b.error);
              }} else {{
                lines.push(b.bucket + ": deleted " + (b.deleted || 0) + ", kept " + (b.kept || 0));
              }}
            }});
            alert(lines.join("\\n"));
          }} catch (e) {{
            alert("R2 cleanup error: " + e);
          }}
        }}

        function _restoreProjectLogsFromStorage() {{
          (projectFolders || []).forEach(function(logId) {{
            var v = "";
            try {{ v = localStorage.getItem(_logStorageKey(logId)) || ""; }} catch (e) {{ v = ""; }}
            _projectLogHistory[logId] = v;
            var logDiv = document.getElementById("log_" + logId);
            if (logDiv && v) {{
              logDiv.innerHTML = v;
              logDiv.scrollTop = logDiv.scrollHeight;
            }}
          }});
        }}

        function disableActionButtons() {{
          var buttons = document.querySelectorAll(".number-button");
          buttons.forEach(function(btn) {{ btn.disabled = true; }});
        }}

        function enableActionButtons() {{
          var buttons = document.querySelectorAll(".number-button");
          buttons.forEach(function(btn) {{ btn.disabled = false; }});
        }}

        function _escapeHtml(s) {{
          return String(s == null ? "" : s)
            .replaceAll("&", "&amp;")
            .replaceAll("<", "&lt;")
            .replaceAll(">", "&gt;");
        }}

        function _fmtDuration(ms) {{
          var total = Math.max(0, Math.floor(Number(ms || 0) / 1000));
          var h = Math.floor(total / 3600);
          var m = Math.floor((total % 3600) / 60);
          var s = total % 60;
          function p2(n) {{ return String(n).padStart(2, "0"); }}
          if (h > 0) return String(h) + ":" + p2(m) + ":" + p2(s);
          return String(m) + ":" + p2(s);
        }}

        function _humanAgo(ms) {{
          if (!ms) return "never";
          var sec = Math.max(0, Math.floor((Date.now() - ms) / 1000));
          if (sec < 60) return String(sec) + "s ago";
          var min = Math.floor(sec / 60);
          var rem = sec % 60;
          return String(min) + "m " + String(rem) + "s ago";
        }}

        function requestBrowserNotifyPermission() {{
          if (!("Notification" in window)) return;
          try {{
            if (Notification.permission === "default") {{
              Notification.requestPermission().catch(function() {{}});
            }}
          }} catch (e) {{
            console.warn("Notification permission request failed:", e);
          }}
        }}

        function notifyTaskFinished(title, body) {{
          if (!("Notification" in window)) return;
          if (Notification.permission !== "granted") return;
          try {{
            var n = new Notification(String(title || "Task finished"), {{
              body: String(body || ""),
              tag: "pinterest-automation-finish"
            }});
            setTimeout(function() {{
              try {{ n.close(); }} catch (e) {{}}
            }}, 9000);
          }} catch (e) {{
            console.warn("Notification failed:", e);
          }}
        }}

        function _isProjectRunning(logId) {{
          return !!_projectRunStartedMs[logId];
        }}

        function _statsMetaHtml(logId) {{
          var last = _humanAgo(_statsLastRefreshMsByLog[logId] || 0);
          var running = _isProjectRunning(logId);
          var runText = running
            ? ("Running for " + _fmtDuration(Date.now() - _projectRunStartedMs[logId]))
            : "Idle";
          return "<div class='stats-meta-line text-muted'>"
            + "<span class='me-2'><i class='bx bx-time-five'></i> " + _escapeHtml(runText) + "</span>"
            + "<span><i class='bx bx-refresh'></i> Refreshed " + _escapeHtml(last) + " (auto " + String(Math.floor(STATS_REFRESH_MS / 1000)) + "s)</span>"
            + "</div>";
        }}

        function refreshProjectStatsMetaOnly(logId) {{
          var el = document.getElementById("stats_" + logId);
          if (!el) return;
          var meta = _statsMetaHtml(logId);
          var old = el.querySelector(".stats-meta-line");
          if (old) {{
            old.outerHTML = meta;
          }} else {{
            el.innerHTML = el.innerHTML + meta;
          }}
        }}

        function _setProjectRunning(logId, running) {{
          if (!logId) return;
          if (running) {{
            if (!_projectRunStartedMs[logId]) _projectRunStartedMs[logId] = Date.now();
          }} else {{
            delete _projectRunStartedMs[logId];
          }}
          refreshProjectStatsMetaOnly(logId);
        }}

        function _scriptToAction(scriptName) {{
          var s = String(scriptName || "").trim();
          var map = {{
            "A.1-START.py": "start",
            "A.2-JSON.py": "json",
            "A.2-PROMPT.py": "prompt",
            "A.3-IMAGINE.py": "imagine",
            "A.4-ARTICLES.py": "article",
            "A.5-PIN DATA.py": "pin_data",
            "A.6-PIN IMAGES.py": "pin_image",
            "A.6b-PIN IMAGES HTML.py": "pin_image_html",
            "A.7-WP UPLOAD.py": "wp_upload",
            "A.8-PIN BULK.py": "pin_bulk"
          }};
          return map[s] || "";
        }}

        function _setBulkStepGlow(stepName, running) {{
          var step = String(stepName || "");
          if (!step) return;
          var btns = document.querySelectorAll('.number-button[data-step="' + step + '"]');
          btns.forEach(function(btn) {{
            if (running) btn.classList.add("running-glow");
            else btn.classList.remove("running-glow");
          }});
        }}

        function _clearAllBulkGlow() {{
          var btns = document.querySelectorAll(".number-button.running-glow");
          btns.forEach(function(btn) {{ btn.classList.remove("running-glow"); }});
        }}

        function _bulkStepFromEndpoint(endpoint) {{
          var ep = String(endpoint || "");
          var map = {{
            "/stream-all-start": "START",
            "/stream-all-json": "JSON",
            "/stream-all-prompt": "PROMPT",
            "/stream-imagine-all": "IMAGINE",
            "/stream-all-article": "ARTICLE",
            "/stream-all-pin-data": "PIN DATA",
            "/stream-all-pin-image": "PIN IMAGE",
            "/stream-all-pin-image-html": "PIN IMAGE",
            "/stream-all-wp-upload": "WP UPLOAD",
            "/stream-all-pin-bulk": "PIN DATA",
            "/stream-all-auto-safe": "START"
          }};
          return map[ep] || "";
        }}

        function _bulkStepFromLogLine(line) {{
          var txt = String(line || "");
          var m = txt.match(/\/(A\.\d-[^ ]+\.py)/);
          if (!m) return "";
          var action = _scriptToAction(m[1]);
          var byAction = {{
            "start": "START",
            "json": "JSON",
            "prompt": "PROMPT",
            "imagine": "IMAGINE",
            "article": "ARTICLE",
            "pin_data": "PIN DATA",
            "pin_image": "PIN IMAGE",
            "pin_image_html": "PIN IMAGE",
            "wp_upload": "WP UPLOAD",
            "pin_bulk": "PIN DATA"
          }};
          return byAction[action] || "";
        }}

        function showColumnDetails(projectLabel, columnName) {{
          var modalEl = document.getElementById("columnDetailsModal");
          var titleEl = document.getElementById("columnDetailsTitle");
          var bodyEl = document.getElementById("columnDetailsBody");
          if (!modalEl || !titleEl || !bodyEl) return;
          titleEl.textContent = projectLabel + " — " + columnName;
          bodyEl.innerHTML = "<div class='text-muted'>Loading...</div>";
          if (typeof bootstrap !== "undefined") {{
            new bootstrap.Modal(modalEl).show();
          }}
          fetch("/api/project-column-details?project=" + encodeURIComponent(projectLabel) + "&column=" + encodeURIComponent(columnName))
            .then(function(r) {{ return r.json().then(function(j) {{ if (!r.ok) throw new Error((j && (j.error || j.message)) || ("HTTP " + r.status)); return j; }}); }})
            .then(function(data) {{
              if (!data || !data.ok) {{
                bodyEl.innerHTML = "<div class='alert alert-danger py-2 mb-0'>Could not load details.</div>";
                return;
              }}
              var rows = Array.isArray(data.rows) ? data.rows : [];
              if (!rows.length) {{
                bodyEl.innerHTML = "<div class='alert alert-warning py-2 mb-0'>No rows found.</div>";
                return;
              }}
              var html = "";
              if (!data.column_exists) {{
                html += "<div class='alert alert-warning py-2'>Column does not exist yet in sheet. Showing Title list with empty values.</div>";
              }}
              html += "<div class='table-responsive'><table class='table table-sm table-striped align-middle mb-0'>";
              html += "<thead><tr><th style='width:70px'>#</th><th>Title</th><th style='width:140px'>Status</th><th>Result</th></tr></thead><tbody>";
              rows.forEach(function(r) {{
                var badge = r.filled
                  ? "<span class='badge bg-success'>Done</span>"
                  : "<span class='badge bg-danger'>Missing</span>";
                var rawVal = String(r.value || "");
                var url = String(r.value_url || "");
                var valueCell = "";
                if (url) {{
                  var lower = rawVal.toLowerCase();
                  var isImg = /\.(jpe?g|png|gif|webp|bmp|svg)$/i.test(lower);
                  var safeUrl = _escapeHtml(url);
                  var safeText = _escapeHtml(rawVal);
                  if (isImg) {{
                    valueCell = ""
                      + "<a href='" + safeUrl + "' target='_blank' rel='noopener noreferrer' "
                      + "class='d-inline-flex align-items-center gap-2 text-decoration-none'>"
                      + "<img src='" + safeUrl + "' alt='' "
                      + "style='max-height:48px;max-width:64px;object-fit:cover;border:1px solid #e5e7eb;border-radius:4px'>"
                      + "<span class='text-break' style='word-break:break-all'>" + safeText + "</span>"
                      + "</a>";
                  }} else {{
                    valueCell = "<a href='" + safeUrl + "' target='_blank' rel='noopener noreferrer' "
                      + "class='text-break' style='word-break:break-all'>" + safeText + "</a>";
                  }}
                }} else {{
                  valueCell = _escapeHtml(rawVal);
                }}
                html += "<tr>"
                  + "<td>" + _escapeHtml(String(r.row || "")) + "</td>"
                  + "<td>" + _escapeHtml(String(r.title || "")) + "</td>"
                  + "<td>" + badge + "</td>"
                  + "<td>" + valueCell + "</td>"
                  + "</tr>";
              }});
              html += "</tbody></table></div>";
              bodyEl.innerHTML = html;
            }})
            .catch(function(e) {{
              bodyEl.innerHTML = "<div class='alert alert-danger py-2 mb-0'>Error: " + _escapeHtml(e && e.message ? e.message : e) + "</div>";
            }});
        }}

        function showColumnDetailsByEncoded(projectEnc, columnEnc) {{
          var p = decodeURIComponent(String(projectEnc || ""));
          var c = decodeURIComponent(String(columnEnc || ""));
          showColumnDetails(p, c);
        }}

        function _statStepForColumn(name) {{
          var lk = String(name || "").trim().toLowerCase();
          if (["title", "recipe", "generated at"].includes(lk)) return "START";
          if (["json recipe"].includes(lk)) return "JSON";
          if (["prompt", "prompt image ingredients"].includes(lk)) return "PROMPT";
          if (["main_image", "image_1", "image_2", "image_3", "image_4", "statu", "error", "main_image_ingredients", "image_ing_1", "image_ing_2", "image_ing_3", "image_ing_4", "statu_ing"].includes(lk)) return "IMAGINE";
          if (["article"].includes(lk)) return "ARTICLE";
          if (["recipe_title_pin", "pinterest_title", "pinterest_description", "pinterest_keywords", "rank_math_focus_keyword", "rank_math_description", "rank_math_pillar_content", "category", "categories"].includes(lk)) return "PIN DATA";
          if (["pinterest_image"].includes(lk)) return "PIN IMAGE";
          if (["status", "post_url"].includes(lk)) return "WP UPLOAD";
          return "OTHER";
        }}

        function _isReverseStatusColumn(name) {{
          var lk = String(name || "").trim().toLowerCase();
          // Reverse semantics: empty means success, filled means problem.
          return lk === "error" || lk === "error_ing";
        }}

        function _renderStatChip(projectLabel, c) {{
          var filled = Number(c.filled || 0);
          var total = Number(c.total || 0);
          var name = String(c.name || "");
          var reverse = _isReverseStatusColumn(name);
          var effFilled = reverse ? Math.max(0, total - filled) : filled;
          var projEnc = encodeURIComponent(String(projectLabel || ""));
          var nameEnc = encodeURIComponent(name);
          var klass = "stat-chip-neutral";
          var icon = "bx-minus-circle";
          if (total > 0) {{
            if (effFilled >= total) {{
              klass = "stat-chip-good";
              icon = "bx-check-circle";
            }} else {{
              klass = "stat-chip-bad";
              icon = "bx-x-circle";
            }}
          }}
          var shownCount = reverse
            ? (String(Math.max(0, total - filled)) + "/" + String(total))
            : (String(filled) + "/" + String(total));
          return ""
            + "<span class='stat-chip " + klass + "' data-st-pe='" + projEnc + "' data-st-ce='" + nameEnc
            + "' title='" + _escapeHtml(name + ": " + shownCount).replace(/'/g, "&#39;") + "'>"
            + "<i class='bx " + icon + "'></i>"
            + "<span class='stat-chip-count'>" + _escapeHtml(shownCount) + "</span>"
            + "<span class='stat-chip-name'>" + _escapeHtml(name) + "</span>"
            + "<i class='bx bx-expand-alt stat-chip-open' role='button' tabindex='0' title='Show details'></i>"
            + "</span>";
        }}

        function _renderGroupedStats(projectLabel, cols) {{
          if (!Array.isArray(cols) || !cols.length) return "No columns";
          var order = ["START", "JSON", "PROMPT", "IMAGINE", "ARTICLE", "PIN DATA", "PIN IMAGE", "WP UPLOAD", "OTHER"];
          var groups = {{}};
          cols.forEach(function(c) {{
            var step = _statStepForColumn(c && c.name);
            if (!groups[step]) groups[step] = [];
            groups[step].push(c);
          }});
          var html = [];
          order.forEach(function(step) {{
            var arr = groups[step] || [];
            if (!arr.length) return;
            var chips = arr.map(function(c) {{ return _renderStatChip(projectLabel, c); }}).join("");
            html.push("<div class='stat-step-group'><span class='stat-step-title'>" + _escapeHtml(step) + ":</span>" + chips + "</div>");
          }});
          return html.join("");
        }}

        function _stepProgressForProject(statsItem, step) {{
          if (!statsItem || !statsItem.ok) return {{ filled: 0, total: 0 }};
          var cols = Array.isArray(statsItem.columns) ? statsItem.columns : [];
          if (!cols.length) return {{ filled: 0, total: 0 }};
          var grouped = _columnsByStep(cols);
          var arr = grouped[step] || [];
          if (!arr.length) return {{ filled: 0, total: 0 }};
          var filled = 0;
          var total = 0;
          arr.forEach(function(c) {{
            if (_isReverseStatusColumn(c && c.name)) {{
              // Do not count error columns in step completion color.
              return;
            }}
            var cf = Number(c && c.filled || 0);
            var ct = Number(c && c.total || 0);
            filled += cf;
            total += ct;
          }});
          return {{ filled: filled, total: total }};
        }}

        function _setStepButtonState(step, state) {{
          var btns = document.querySelectorAll('.number-button[data-step="' + step + '"]');
          btns.forEach(function(btn) {{
            if (btn.classList.contains("active")) return;
            btn.classList.remove("finished", "partial");
            if (state === "done") {{
              btn.classList.add("finished");
            }} else if (state === "partial") {{
              btn.classList.add("partial");
            }}
          }});
        }}

        function refreshActionButtonStates() {{
          fetch("/api/projects-stats")
            .then(function(r) {{
              return r.json().then(function(j) {{
                if (!r.ok) throw new Error((j && (j.error || j.message)) || ("HTTP " + r.status));
                return j;
              }});
            }})
            .then(function(payload) {{
              var items = Array.isArray(payload && payload.items) ? payload.items : [];
              var steps = ["START", "JSON", "PROMPT", "IMAGINE", "ARTICLE", "PIN DATA", "PIN IMAGE", "WP UPLOAD"];
              steps.forEach(function(step) {{
                var sumFilled = 0;
                var sumTotal = 0;
                var hasPartial = false;
                items.forEach(function(it) {{
                  var p = _stepProgressForProject(it, step);
                  sumFilled += p.filled;
                  sumTotal += p.total;
                  if (p.filled > 0) hasPartial = true;
                }});
                if (sumTotal > 0 && sumFilled >= sumTotal) {{
                  _setStepButtonState(step, "done");
                }} else if (hasPartial) {{
                  _setStepButtonState(step, "partial");
                }} else {{
                  _setStepButtonState(step, "none");
                }}
              }});
            }})
            .catch(function(err) {{
              console.warn("Could not refresh action button states:", err);
            }});
        }}

        function _clearCascadeFrom(step) {{
          var order = ["START", "JSON", "PROMPT", "IMAGINE", "ARTICLE", "PIN DATA", "PIN IMAGE", "WP UPLOAD", "PIN BULK"];
          var s = String(step || "").trim().toUpperCase();
          var idx = order.indexOf(s);
          if (idx < 0) return [];
          return order.slice(idx);
        }}

        function clearStepAction(step) {{
          var s = String(step || "").trim().toUpperCase();
          if (!s) return;
          var cascade = _clearCascadeFrom(s);
          if (!cascade.length) {{
            alert("Unknown step: " + s);
            return;
          }}
          var casc = _columnsForCascadeClear(s);
          var colLine = casc.columns.length ? ("Columns cleared:\\n  " + casc.columns.join(", ")) : "";
          var msg = [
            "WARNING: You are about to clear step data.",
            "",
            "Selected step: " + s,
            "This will also clear dependent next steps:",
            "  " + cascade.join("  ->  "),
            colLine,
            "",
            "This action will run for ALL projects and cannot be undone.",
            "Do you want to continue?"
          ].filter(function(x) {{ return x !== ""; }}).join("\\n");
          if (!window.confirm(msg)) return;

          fetch("/api/clear-step", {{
            method: "POST",
            headers: {{ "Content-Type": "application/json" }},
            body: JSON.stringify({{ step: s }})
          }})
          .then(function(r) {{
            return r.json().then(function(j) {{
              if (!r.ok) throw new Error((j && (j.error || j.message)) || ("HTTP " + r.status));
              return j;
            }});
          }})
          .then(function(data) {{
            var rows = Array.isArray(data.projects) ? data.projects : [];
            var okCount = rows.filter(function(x) {{ return !!x.ok; }}).length;
            var failCount = rows.length - okCount;
            alert(
              "Clear completed.\\n\\n" +
              "Step: " + String(data.step || s) + "\\n" +
              "Cascade: " + (Array.isArray(data.cascade_steps) ? data.cascade_steps.join(" -> ") : cascade.join(" -> ")) + "\\n" +
              "Cleared cells: " + String(data.total_cleared_cells || 0) + "\\n" +
              "Projects OK: " + String(okCount) + "\\n" +
              "Projects failed: " + String(failCount)
            );
            refreshAllProjectStats();
            refreshActionButtonStates();
          }})
          .catch(function(err) {{
            alert("Step clear failed: " + (err && err.message ? err.message : err));
          }});
        }}

        function refreshProjectStats(logId, projectLabel) {{
          var el = document.getElementById("stats_" + logId);
          if (!el) return;
          function renderCols(cols) {{
            if (!Array.isArray(cols) || !cols.length) {{
              el.innerHTML = "No columns" + _statsMetaHtml(logId);
              return;
            }}
            el.innerHTML = _renderGroupedStats(projectLabel, cols) + _statsMetaHtml(logId);
          }}
          fetch("/api/project-stats?project=" + encodeURIComponent(projectLabel))
            .then(function(r) {{ return r.json().then(function(j) {{ if (!r.ok) throw new Error((j && (j.error || j.message)) || ("HTTP " + r.status)); return j; }}); }})
            .then(function(data) {{
              if (!data || !data.ok) {{
                el.innerHTML = "Stats unavailable" + _statsMetaHtml(logId);
                return;
              }}
              _statsLastRefreshMsByLog[logId] = Date.now();
              renderCols(Array.isArray(data.columns) ? data.columns : []);
            }})
            .catch(function() {{
              el.innerHTML = "Stats unavailable" + _statsMetaHtml(logId);
            }});
        }}

        function refreshAllProjectStats() {{
          // Progressive load: fetch each project stats separately so UI becomes responsive faster.
          var units = Array.isArray(projectUnits) ? projectUnits.slice() : [];
          units.forEach(function(u, idx) {{
            setTimeout(function() {{
              refreshProjectStats(u.log_id, u.label);
            }}, idx * 120);
          }});
        }}

        var _endpointPrerequisites = {{
          "/stream-all-json": ["START"],
          "/stream-all-prompt": ["JSON"],
          "/stream-imagine-all": ["PROMPT"],
          "/stream-imagine-group1": ["PROMPT"],
          "/stream-imagine-group2": ["PROMPT"],
          "/stream-imagine-group3": ["PROMPT"],
          "/stream-imagine-group4": ["PROMPT"],
          "/stream-imagine-group5": ["PROMPT"],
          "/stream-imagine-group6": ["PROMPT"],
          "/stream-imagine-group7": ["PROMPT"],
          "/stream-imagine-group8": ["PROMPT"],
          "/stream-imagine-group9": ["PROMPT"],
          "/stream-imagine-group10": ["PROMPT"],
          "/stream-imagine-group11": ["PROMPT"],
          "/stream-imagine-group12": ["PROMPT"],
          "/stream-imagine-group13": ["PROMPT"],
          "/stream-imagine-group14": ["PROMPT"],
          "/stream-imagine-group15": ["PROMPT"],
          "/stream-imagine-group16": ["PROMPT"],
          "/stream-imagine-group17": ["PROMPT"],
          "/stream-all-article": ["IMAGINE"],
          "/stream-all-pin-data": ["ARTICLE"],
          "/stream-all-pin-image": ["PIN DATA"],
          "/stream-all-pin-image-html": ["PIN DATA"],
          "/stream-all-wp-upload": ["ARTICLE", "PIN DATA", "PIN IMAGE"],
          "/stream-all-pin-bulk": ["PIN DATA"],
          "/stream-all-cf-upload": ["ARTICLE"]
        }};

        function _columnsByStep(cols) {{
          var grouped = {{}};
          (Array.isArray(cols) ? cols : []).forEach(function(c) {{
            var step = _statStepForColumn(c && c.name);
            if (!grouped[step]) grouped[step] = [];
            grouped[step].push(c || {{}});
          }});
          return grouped;
        }}

        function _isStepCompletedForProject(statsItem, step) {{
          if (!statsItem || !statsItem.ok) return false;
          var cols = Array.isArray(statsItem.columns) ? statsItem.columns : [];
          if (!cols.length) return false;
          var maxTotal = 0;
          cols.forEach(function(c) {{
            var t = Number(c && c.total || 0);
            if (t > maxTotal) maxTotal = t;
          }});
          // Empty project (no titles) should not block the pipeline.
          if (maxTotal <= 0) return true;

          var grouped = _columnsByStep(cols);
          var stepCols = grouped[step] || [];
          if (!stepCols.length) return false;
          for (var i = 0; i < stepCols.length; i++) {{
            var one = stepCols[i] || {{}};
            var filled = Number(one.filled || 0);
            var total = Number(one.total || 0);
            if (total > 0 && filled < total) return false;
          }}
          return true;
        }}

        function _buildDependencyWarningMessage(endpoint, requiredSteps, blockedProjects) {{
          var lines = [];
          lines.push("Warning: this action needs previous step(s) completed first.");
          lines.push("");
          lines.push("Action: " + endpoint);
          lines.push("Required before run: " + requiredSteps.join(" -> "));
          lines.push("");
          lines.push("Blocked projects (" + blockedProjects.length + "):");
          var preview = blockedProjects.slice(0, 8);
          preview.forEach(function(bp) {{
            lines.push("- " + bp.project + " (missing: " + bp.missing.join(", ") + ")");
          }});
          if (blockedProjects.length > preview.length) {{
            lines.push("... and " + String(blockedProjects.length - preview.length) + " more projects.");
          }}
          lines.push("");
          lines.push("Continue anyway?");
          return lines.join("\\n");
        }}

        function _checkEndpointDependencies(endpoint) {{
          // Disabled on purpose: do not perform pre-run verification calls.
          // Runs must start immediately for all action buttons.
          return Promise.resolve({{ allowed: true }});
        }}

        document.addEventListener("click", function(ev) {{
          var opener = ev.target.closest(".stat-chip-open");
          if (!opener) return;
          var chip = opener.closest(".stat-chip");
          if (!chip) return;
          var pe = chip.getAttribute("data-st-pe");
          var ce = chip.getAttribute("data-st-ce");
          if (pe !== null && ce !== null && String(pe).length) {{
            showColumnDetailsByEncoded(pe, ce);
          }}
        }});

        function startLog(endpoint, btn) {{
          if(source !== null) {{ return; }}
          disableActionButtons();
          _clearAllBulkGlow();
          var initialStep = _bulkStepFromEndpoint(endpoint);
          if (initialStep) _setBulkStepGlow(initialStep, true);
          btn.classList.add("running-glow");

          source = new EventSource(endpoint);
          source.onmessage = function(e) {{
            try {{
              let data = JSON.parse(e.data);
              let folder = data.folder;
              let line = data.line;
              if (folder && folder !== "all") {{
                _appendProjectLog(folder, line, line.includes("Finished"));
                var lineStep = _bulkStepFromLogLine(line);
                if (line && line.includes("Running ")) {{
                  if (lineStep) _setBulkStepGlow(lineStep, true);
                }}
                if (line && line.includes("Finished")) {{
                  var finStep = _bulkStepFromLogLine(line);
                  if (finStep) _setBulkStepGlow(finStep, false);
                }}
                if (line && line.includes("Running ")) _setProjectRunning(folder, true);
                if(line.includes("Finished")) {{
                  _setProjectRunning(folder, false);
                  var pu = (projectUnits || []).find(function(x) {{ return x.log_id === folder; }});
                  if (pu) refreshProjectStats(pu.log_id, pu.label);
                }}
              }}
    if (
      folder === "all" &&
      (
        line.includes("Finished all processes.") ||
        line.includes("Finished all IMAGINE ALL groups.")
      )
    ) {{
      btn.classList.remove("active");
      btn.classList.add("finished");
      enableActionButtons();
      source.close();
      source = null;
      _clearAllBulkGlow();
      refreshActionButtonStates();
      notifyTaskFinished("Bulk task finished", "The selected dashboard step completed.");
    }}

  }} catch (err) {{
    console.error("SSE parse error:", err);
  }}
}};
          source.onerror = function(err) {{
            console.error("EventSource error:", err);
            enableActionButtons();
            source.close();
            source = null;
            _clearAllBulkGlow();
          }};
          btn.classList.add("active");

          // No pre-run dependency verification: start immediately.
        }}

        function askStartLimit() {{
          var raw = prompt("How many titles to use now? (empty = all)", "");
          if (raw === null) return null;
          var s = String(raw || "").trim();
          if (!s) return "";
          var n = Number(s);
          if (!Number.isFinite(n) || n <= 0) {{
            alert("Please enter a positive number, or leave empty for all.");
            return null;
          }}
          return String(Math.floor(n));
        }}

        function startAllStart(btn) {{
          var lim = askStartLimit();
          if (lim === null) return;
          var ep = "/stream-all-start";
          if (lim) ep += "?title_limit=" + encodeURIComponent(lim);
          startLog(ep, btn);
        }}

        function stopLog() {{
          if (source) {{
            source.close();
            source = null;
            enableActionButtons();
          }}
          fetch("/stop_scripts", {{ method: "POST" }})
            .then(response => response.json())
            .then(data => {{ console.log("Stopped scripts:", data.message); }})
            .catch(err => {{ console.error("Error stopping scripts:", err); }});
        }}

        function clearFailed() {{
          fetch("/clear_failed", {{ method: "POST" }})
          .then(r => r.json())
          .then(data => {{
            if(data.status === "success") {{
              alert("Clear FAILED done: " + JSON.stringify(data.results));
            }} else {{
              alert("Error: " + data.message);
            }}
          }})
          .catch(err => alert("Error: " + err));
        }}

        document.addEventListener("DOMContentLoaded", function() {{
          applyStepButtonTooltips();
        }});

        var projectStreams = {{}};
        function startProjectAction(logId, projectLabel, action, btn) {{
          if(projectStreams[logId]) {{ return; }}
          var card = btn.closest('.card');
          var buttons = card.querySelectorAll('button.project-action');
          var endpointByAction = {{
            json: "/stream-all-json",
            prompt: "/stream-all-prompt",
            imagine: "/stream-imagine-all",
            article: "/stream-all-article",
            pin_data: "/stream-all-pin-data",
            pin_image: "/stream-all-pin-image",
            pin_image_html: "/stream-all-pin-image-html",
            wp_upload: "/stream-all-wp-upload",
            pin_bulk: "/stream-all-pin-bulk",
            cf_upload: "/stream-cf-upload"
          }};
          var depEndpoint = endpointByAction[action] || "";
          buttons.forEach(function(b) {{ b.disabled = true; }});
          btn.classList.add("running-glow");
          var ep = "/stream-single?project=" + encodeURIComponent(projectLabel) + "&action=" + encodeURIComponent(action);
          if (action === "start") {{
            var lim = askStartLimit();
            if (lim === null) {{
              buttons.forEach(function(b) {{ b.disabled = false; }});
              return;
            }}
            if (lim) ep += "&title_limit=" + encodeURIComponent(lim);
          }}
          var source = new EventSource(ep);
          projectStreams[logId] = source;
          _setProjectRunning(logId, true);
          source.onmessage = function(e) {{
            try {{
              let data = JSON.parse(e.data);
              let logDiv = document.getElementById("log_" + data.folder);
              if (!logDiv) logDiv = document.getElementById("log_" + logId);
              var targetLogId = data.folder || logId;
              _appendProjectLog(targetLogId, data.line, data.line.includes("Finished"));
              if(data.line.includes("Finished")) {{
                _setProjectRunning(logId, false);
                source.close();
                delete projectStreams[logId];
                buttons.forEach(function(b) {{ b.disabled = false; }});
                btn.classList.remove("running-glow");
                refreshProjectStats(logId, projectLabel);
                refreshActionButtonStates();
                  notifyTaskFinished("Project task finished", projectLabel + " - " + action + " completed.");
              }}
            }} catch(err) {{
              console.error("Project SSE parse error:", err);
            }}
          }};
          source.onerror = function(err) {{
            console.error("Project EventSource error:", err);
            _setProjectRunning(logId, false);
            source.close();
            delete projectStreams[logId];
            buttons.forEach(function(b) {{ b.disabled = false; }});
            btn.classList.remove("running-glow");
          }};
          // No pre-run dependency verification: start immediately.
        }}
        document.addEventListener("DOMContentLoaded", function() {{
          _restoreProjectLogsFromStorage();
        }});

        var _siteEditorProject = "";
        var _siteEditorBase = null;
        var _siteEditorRawDirty = false;
        var PROMPT_FORM_KEYS = ["a1_start", "a2_json", "a2_prompt", "a4_articles", "a5_pin_data", "a6b_pin_image_html", "a8_pin_bulk", "app_title_filter"];
        var _siteEditorPromptFieldSchema = null;
        function _promptGroupContainerId(pr) {{
          if (pr === "a1_start") return "ed_pr_sub_start";
          if (pr === "a2_json" || pr === "a2_prompt") return "ed_pr_sub_a2";
          if (pr === "a4_articles" || pr === "a5_pin_data" || pr === "a6b_pin_image_html" || pr === "a8_pin_bulk") return "ed_pr_sub_pipe";
          if (pr === "app_title_filter") return "ed_pr_sub_other";
          return "ed_pr_sub_other";
        }}
        function _promptFieldDomId(pr, path) {{ return "edpf_" + pr + "__" + String(path).replace(/\\./g, "__"); }}
        function _fieldIsNonEmpty(f, el) {{
          if (!el) return false;
          if (f.kind === "bool") return !!el.checked;
          if (f.kind === "number") return String(el.value).trim() !== "";
          return String(el.value || "").trim() !== "";
        }}
        function renderPromptsFromSchema(snap) {{
          var sch = (snap && snap.prompts_inline_field_schema) || null;
          window._siteEditorPromptFieldSchema = sch;
          var ids = ["ed_pr_sub_start", "ed_pr_sub_a2", "ed_pr_sub_pipe", "ed_pr_sub_other"];
          ids.forEach(function(id) {{ var n = document.getElementById(id); if (n) n.innerHTML = ""; }});
          var oLbl = document.getElementById("ed_pr_sub_other_lbl");
          if (oLbl) oLbl.classList.add("d-none");
          if (!sch || typeof sch !== "object") {{
            var n0 = document.getElementById("ed_pr_sub_start");
            if (n0) n0.innerHTML = '<p class="text-warning small">No prompt field schema. Add <code>config/prompts/*.json</code> in the repo.</p>';
            return;
          }}
          var order = Object.keys(sch);
          order.sort();
          for (var oi = 0; oi < order.length; oi++) {{
            var pr = order[oi];
            var gid = _promptGroupContainerId(pr);
            var c = document.getElementById(gid);
            if (gid === "ed_pr_sub_other" && oLbl) oLbl.classList.remove("d-none");
            if (!c) continue;
            var block = document.createElement("div");
            block.className = "mb-3 border-bottom pb-2";
            var title = document.createElement("div");
            title.className = "fw-bold text-secondary small";
            title.textContent = pr;
            block.appendChild(title);
            var layerP = document.createElement("p");
            layerP.className = "form-text small mb-2";
            layerP.id = "ed_src_pr_" + pr;
            layerP.style.fontSize = "0.68rem";
            block.appendChild(layerP);
            var flist = sch[pr] || [];
            var categorizeHdrShown = false;
            for (var fi = 0; fi < flist.length; fi++) {{
              var f = flist[fi];
              if (!f || !f.path) continue;
              if (pr === "a5_pin_data" && f.group === "categorize" && !categorizeHdrShown) {{
                categorizeHdrShown = true;
                var catHdr = document.createElement("div");
                catHdr.className = "alert alert-secondary py-2 px-2 small mb-2";
                catHdr.innerHTML = "<strong>Recipe → WordPress category</strong> — title in, category name out. Edit <code>categorize.system</code> / <code>categorize.user</code> below. Category names come from <code>category_id_mapping</code> in Settings (<code>{{category_list}}</code> placeholder).";
                block.appendChild(catHdr);
              }}
              var pid = _promptFieldDomId(pr, f.path);
              var wrap = document.createElement("div");
              wrap.className = "mb-2";
              var lab = document.createElement("label");
              lab.className = "form-label small text-muted d-block mb-0";
              lab.setAttribute("for", pid);
              lab.textContent = (f.label || f.path) + "  (" + (f.kind || "text") + ")";
              wrap.appendChild(lab);
              var el;
              var knd = f.kind || "text";
              if (knd === "bool") {{
                el = document.createElement("input");
                el.type = "checkbox";
                el.className = "form-check-input";
                el.id = pid;
              }} else if (knd === "number") {{
                el = document.createElement("input");
                el.type = "number";
                el.step = "any";
                el.className = "form-control form-control-sm";
                el.id = pid;
              }} else if (knd === "textarea" || knd === "json" || knd === "lines") {{
                el = document.createElement("textarea");
                el.className = "form-control form-control-sm" + (knd === "json" || knd === "lines" ? " font-monospace" : "");
                el.id = pid;
                el.spellcheck = false;
                if (knd === "json") el.rows = 5;
                else if (knd === "lines") el.rows = 4;
                else el.rows = 3;
              }} else {{
                el = document.createElement("input");
                el.type = "text";
                el.className = "form-control form-control-sm";
                el.id = pid;
                el.spellcheck = false;
              }}
              wrap.appendChild(el);
              var effHint = document.createElement("p");
              effHint.className = "form-text text-muted mb-0 d-none";
              effHint.id = "edeff_" + pr + "__" + String(f.path).replace(/\\./g, "__");
              effHint.style.fontSize = "0.68rem";
              effHint.setAttribute("aria-hidden", "true");
              wrap.appendChild(effHint);
              block.appendChild(wrap);
            }}
            c.appendChild(block);
          }}
        }}
        function _getNested(obj, path) {{
          if (!obj || !path) return undefined;
          var c = obj;
          var p = path.split(".");
          for (var i = 0; i < p.length; i++) {{ if (c == null) return undefined; c = c[p[i]]; }}
          return c;
        }}
        function _setNested(obj, path, val) {{
          if (!obj || !path) return;
          var p = path.split(".");
          var c = obj;
          for (var i = 0; i < p.length - 1; i++) {{
            var k = p[i];
            if (!c[k] || typeof c[k] !== "object" || c[k] === null) c[k] = {{}};
            c = c[k];
          }}
          c[p[p.length - 1]] = val;
        }}
        function _removeNestedPath(obj, path) {{
          if (!obj || !path) return;
          var p = path.split(".");
          function walk(o, d) {{
            if (o == null || typeof o !== "object") return;
            if (d === p.length - 1) {{ delete o[p[d]]; return; }}
            if (!(p[d] in o)) return;
            var ch = o[p[d]];
            walk(ch, d + 1);
            if (ch && typeof ch === "object" && !Array.isArray(ch) && !Object.keys(ch).length) delete o[p[d]];
          }}
          walk(obj, 0);
        }}
        function _pruneEmpty(obj) {{
          if (obj == null || typeof obj !== "object" || Array.isArray(obj)) return;
          for (var k in obj) {{ if (Object.prototype.hasOwnProperty.call(obj, k)) _pruneEmpty(obj[k]); }}
          for (var k2 in obj) {{
            if (Object.prototype.hasOwnProperty.call(obj, k2)) {{
              var v = obj[k2];
              if (v && typeof v === "object" && !Array.isArray(v) && !Object.keys(v).length) delete obj[k2];
            }}
          }}
        }}
        function _isDeepEmptyObject(o) {{
          if (o == null) return true;
          if (typeof o !== "object" || Array.isArray(o)) return false;
          return Object.keys(o).length === 0;
        }}
        function _fillPromptFieldsFromRow(s) {{
          if (!s || typeof s !== "object") return;
          var sch = window._siteEditorPromptFieldSchema;
          if (!sch) return;
          var pm = s.prompts || {{}};
          Object.keys(sch).forEach(function(pr) {{
            (sch[pr] || []).forEach(function(f) {{
              var el = document.getElementById(_promptFieldDomId(pr, f.path));
              if (!el) return;
              var node = pm[pr];
              var v = _getNested(node, f.path);
              var knd = f.kind || "text";
              if (knd === "bool") {{ el.checked = !!v; return; }}
              if (knd === "number") {{
                el.value = (v != null && v !== "" && !isNaN(Number(v))) ? String(v) : "";
                return;
              }}
              if (knd === "lines" && v && Array.isArray(v)) {{ el.value = v.join("\\n"); return; }}
              if (knd === "json") {{
                el.value = (v == null) ? "" : JSON.stringify(v, null, 2);
                return;
              }}
              el.value = (v == null) ? "" : String(v);
            }});
          }});
        }}
        function _readPromptsFromFormInto(bp, b) {{
          var sch = window._siteEditorPromptFieldSchema;
          if (!sch) return;
          Object.keys(sch).forEach(function(name) {{
            var fields = sch[name] || [];
            if (!fields.length) return;
            var anySet = fields.some(function(f) {{
              return _fieldIsNonEmpty(f, document.getElementById(_promptFieldDomId(name, f.path)));
            }});
            if (!anySet) {{ delete bp[name]; return; }}
            var m = (b.prompts && b.prompts[name] && typeof b.prompts[name] === "object")
              ? JSON.parse(JSON.stringify(b.prompts[name])) : {{}};
            fields.forEach(function(f) {{
              var e2 = document.getElementById(_promptFieldDomId(name, f.path));
              if (!e2) return;
              var knd = f.kind || "text";
              if (!_fieldIsNonEmpty(f, e2)) {{
                if (f.path.indexOf(".") === -1) delete m[f.path];
                else {{ _removeNestedPath(m, f.path); _pruneEmpty(m); }}
                return;
              }}
              if (knd === "bool") {{ _setNested(m, f.path, !!e2.checked); return; }}
              if (knd === "number") {{
                var n = Number(e2.value);
                if (isNaN(n)) throw new Error("prompts." + name + " " + f.path + ": not a number");
                _setNested(m, f.path, n);
                return;
              }}
              if (knd === "lines") {{
                var lines = String(e2.value).split(/\\r?\\n/).map(function(x) {{ return x.trim(); }}).filter(Boolean);
                _setNested(m, f.path, lines);
                return;
              }}
              if (knd === "json") {{
                try {{ _setNested(m, f.path, JSON.parse(e2.value.trim())); }}
                catch (err) {{ throw new Error("prompts." + name + " " + f.path + ": " + (err.message || err)); }}
                return;
              }}
              _setNested(m, f.path, String(e2.value));
            }});
            _pruneEmpty(m);
            if (_isDeepEmptyObject(m) || (typeof m === "object" && !Object.keys(m).length)) delete bp[name];
            else bp[name] = m;
          }});
        }}
        var PROVENANCE_FIELD_IDS = [
          "wordpress_url", "wordpress_user", "wordpress_app_password", "openai_api_key", "openai_model",
          "useapi_token", "useapi_midjourney_channel", "r2_account_id", "r2_access_key_id", "r2_secret_access_key",
          "r2_bucket", "r2_public_base_url"
        ];
        function _applyProvenanceToForm(snap) {{
          var prov = (snap && snap.keys_provenance) || {{}};
          PROVENANCE_FIELD_IDS.forEach(function(k) {{
            var e = document.getElementById("ed_src_" + k);
            if (e) e.textContent = prov[k] ? ("Key: " + k + " — " + prov[k]) : "";
          }});
          var ph = (snap && snap.prompts_form_hints) || {{}};
          PROMPT_FORM_KEYS.forEach(function(name) {{
            var e = document.getElementById("ed_src_pr_" + name);
            if (e) e.textContent = ph[name] ? ("Layers: " + ph[name]) : "";
          }});
        }}
        function _promptPathHasRowOverride(s, pr, path) {{
          if (!s || !pr || !path) return false;
          var node = s.prompts && s.prompts[pr];
          if (node == null || typeof node !== "object") return false;
          return _getNested(node, path) !== undefined;
        }}
        function _strEffForUi(v) {{
          if (v === null || v === undefined) return "";
          if (typeof v === "object") return JSON.stringify(v, null, 2);
          if (typeof v === "boolean") return v ? "true" : "false";
          return String(v);
        }}
        function _applyEffectivePromptPlaceholders(snap, s) {{
          var eff = (snap && snap.prompts_effective_by_path) || null;
          var excl = (snap && snap.prompts_excluding_row_inline_by_path) || null;
          if (!eff || typeof eff !== "object") return;
          if (!excl) excl = {{}};
          var sch = window._siteEditorPromptFieldSchema;
          if (!sch) return;
          Object.keys(sch).forEach(function(pr) {{
            var pmap = (eff[pr] && typeof eff[pr] === "object") ? eff[pr] : {{}};
            var xmap = (excl[pr] && typeof excl[pr] === "object") ? excl[pr] : {{}};
            (sch[pr] || []).forEach(function(f) {{
              if (!f || !f.path) return;
              var el = document.getElementById(_promptFieldDomId(pr, f.path));
              if (!el) return;
              var hintId = "edeff_" + pr + "__" + String(f.path).replace(/\\./g, "__");
              var hint = document.getElementById(hintId);
              if (!Object.prototype.hasOwnProperty.call(pmap, f.path)) {{
                if (f.kind === "bool") el.removeAttribute("title");
                else {{ el.placeholder = ""; el.removeAttribute("title"); }}
                if (hint) {{ hint.textContent = ""; hint.classList.add("d-none"); }}
                return;
              }}
              var pe = pmap[f.path];
              var xb = Object.prototype.hasOwnProperty.call(xmap, f.path) ? xmap[f.path] : undefined;
              var full = _strEffForUi(pe);
              var fullBase = (xb === undefined) ? "" : _strEffForUi(xb);
              var knd = f.kind || "text";
              var hasO = _promptPathHasRowOverride(s, pr, f.path);
              if (knd === "bool") {{
                if (hint) {{
                  if (!hasO) {{ hint.textContent = "No per-row value — effective merged: " + (pe ? "true" : "false") + "."; hint.classList.remove("d-none"); }}
                  else {{ hint.textContent = (xb === undefined) ? "Row overrides. No file baseline for this key (inline-only or nested)." : ("Row sets explicit checkbox. From file layers (no row inline): " + (xb ? "true" : "false") + "."); hint.classList.remove("d-none"); }}
                }}
                el.setAttribute("title", hasO
                  ? ((xb === undefined) ? "Effective merged (includes row): " + full : ("From file layers (no row inline `prompts`): " + fullBase + " — effective with row: " + full))
                  : ("Effective merged: " + full));
                return;
              }}
              if (!hasO) {{
                el.placeholder = (full.length > 220) ? (full.slice(0, 217) + "…") : full;
                el.setAttribute("title", "Merged when the row has no key for " + f.path + " (hover for full if truncated).\\n" + full);
                if (hint) {{ hint.textContent = (full.length > 160) ? ("Effective merged: " + full.slice(0, 157) + "…") : ("Effective merged: " + full); hint.classList.remove("d-none"); }}
              }} else {{
                el.placeholder = "";
                el.setAttribute("title", (xb === undefined) ? "Row overrides. Effective at this path: " + full
                  : "Row inline overrides. From file layers (no `prompts` in row): " + fullBase + "\\n--\\nWith row, effective: " + full);
                if (hint) {{ var hb = (fullBase.length > 100) ? fullBase.slice(0, 97) + "…" : fullBase; hint.textContent = (xb === undefined) ? "Row overrides. No separate file-baseline for this key." : ("Row overrides. File-layer baseline: " + hb); hint.classList.remove("d-none"); }}
              }}
            }});
          }});
        }}
        function siteEditorApplyKeepSecretsFromBase(base, o) {{
          if (!base || !o || typeof o !== "object") return;
          var keys = [
            "wordpress_app_password", "openai_api_key", "useapi_token", "useapi_midjourney_channel",
            "r2_account_id", "r2_access_key_id", "r2_secret_access_key", "r2_bucket", "r2_public_base_url"
          ];
          keys.forEach(function(k) {{
            var t = o[k] != null ? String(o[k]).trim() : "";
            if (t) return;
            if (base[k] != null && String(base[k]).trim() !== "") o[k] = base[k];
          }});
        }}
        function _getV(id) {{ var e = document.getElementById(id); return e ? (e.value || "").trim() : ""; }}
        function _setV(id, v) {{ var e = document.getElementById(id); if (e) e.value = (v != null && v !== undefined) ? String(v) : ""; }}
        function _rowStr(s, k) {{
          if (!s || typeof s !== "object" || !k) return "";
          var x = s[k];
          if (x == null) return "";
          return String(x).trim();
        }}
        function _setPh(id, text) {{
          var e = document.getElementById(id);
          if (!e) return;
          e.placeholder = text != null ? String(text) : "";
          if (text) e.setAttribute("title", "Effective merged (what the run uses): " + String(text).slice(0, 400));
          else e.removeAttribute("title");
        }}
        function _fillOrPlaceholder(s, kf, rowKey, elId) {{
          var r = _rowStr(s, rowKey);
          if (r) {{ _setV(elId, r); _setPh(elId, ""); return; }}
          _setV(elId, "");
          var m = (kf && kf[rowKey] != null) ? kf[rowKey] : null;
          var hint = (m != null && String(m).trim() !== "") ? String(m) : "";
          if (!hint) hint = "(not on this row — use merged keys: open the API & R2 tab for placeholders / Source lines.)";
          _setPh(elId, hint);
        }}
        function _fillSecretOrPh(s, kf, rowKey, elId) {{
          var r = _rowStr(s, rowKey);
          if (r) {{ _setV(elId, r); _setPh(elId, "leave blank=keep; replace to change in row"); return; }}
          _setV(elId, "");
          var m = (kf && kf[rowKey] != null) ? kf[rowKey] : null;
          var hint = (m != null && String(m).trim() !== "")
            ? String(m)
            : "(not on row — merged from shared / project / env. See the API & R2 tab. Leave empty to not store in row.)";
          _setPh(elId, hint);
        }}
        function siteEditorFillForm(s, snap) {{
          if (!s || typeof s !== "object") return;
          var kf = (snap && snap.keys && typeof snap.keys === "object") ? snap.keys : {{}};
          _setV("ed_id", s.id);
          _setV("ed_display_name", s.display_name);
          _setV("ed_out_dir", s.out_dir);
          _setV("ed_start_file", s.start_file);
          _setV("ed_templates_dir", s.templates_dir);
          _setV("ed_log_id", s.log_id);
          _setV("ed_prompts_dir", s.prompts_dir);
          _fillOrPlaceholder(s, kf, "wordpress_url", "ed_wordpress_url");
          _fillOrPlaceholder(s, kf, "wordpress_user", "ed_wordpress_user");
          _setV("ed_wordpress_app_password", "");
          _setPh("ed_wordpress_app_password", _rowStr(s, "wordpress_app_password")
            ? "leave blank=keep; replace to change in row"
            : ((kf.wordpress_app_password != null && String(kf.wordpress_app_password).trim() !== "")
                ? String(kf.wordpress_app_password)
                : "blank=keep; effective from merged keys (if set)"));
          _fillSecretOrPh(s, kf, "openai_api_key", "ed_openai_api_key");
          _fillOrPlaceholder(s, kf, "openai_model", "ed_openai_model");
          _fillSecretOrPh(s, kf, "useapi_token", "ed_useapi_token");
          _fillOrPlaceholder(s, kf, "useapi_midjourney_channel", "ed_useapi_midjourney_channel");
          _fillOrPlaceholder(s, kf, "r2_account_id", "ed_r2_account_id");
          _fillSecretOrPh(s, kf, "r2_access_key_id", "ed_r2_access_key_id");
          _fillSecretOrPh(s, kf, "r2_secret_access_key", "ed_r2_secret_access_key");
          _fillOrPlaceholder(s, kf, "r2_bucket", "ed_r2_bucket");
          _fillOrPlaceholder(s, kf, "r2_public_base_url", "ed_r2_public_base_url");
          var nss = document.getElementById("ed_no_shared_settings");
          if (nss) nss.checked = !!s.no_shared_settings;
          var nsp = document.getElementById("ed_no_shared_prompts");
          if (nsp) nsp.checked = !!s.no_shared_prompts;
          var es = s.settings;
          _setV("ed_settings_json", (es && typeof es === "object") ? JSON.stringify(es, null, 2) : "");
          var stSet = document.getElementById("ed_settings_json");
          var mset = (snap && snap.settings) ? JSON.stringify(snap.settings, null, 2) : "";
          if (!es || typeof es !== "object") {{
            if (mset && mset.length) {{
              _setPh("ed_settings_json", mset.length > 520 ? mset.slice(0, 517) + "…" : mset);
              if (stSet) stSet.setAttribute("title", "Merged settings if you leave this field empty. Full JSON in hover.\\n" + mset);
            }} else {{
              _setPh("ed_settings_json", "Empty = use merged project settings. Paste {{}} to clear a previous row override.");
              if (stSet) stSet.removeAttribute("title");
            }}
          }} else {{ _setPh("ed_settings_json", ""); if (stSet) stSet.removeAttribute("title"); }}
          renderPromptsFromSchema(snap);
          _fillPromptFieldsFromRow(s);
          _applyProvenanceToForm(snap);
          _applyEffectivePromptPlaceholders(snap, s);
        }}
        function siteEditorReadMerged() {{
          var b = _siteEditorBase;
          if (!b) return null;
          var o = JSON.parse(JSON.stringify(b));
          var sk = ["display_name","out_dir","start_file","templates_dir","log_id","prompts_dir","wordpress_url","wordpress_user","openai_model"];
          sk.forEach(function(k) {{ var v = _getV("ed_" + k); if (v) o[k] = v; else if (b[k] !== undefined) o[k] = b[k]; }});
          var pass = _getV("ed_wordpress_app_password");
          if (pass) o.wordpress_app_password = pass;
          var sec = ["openai_api_key","useapi_token","useapi_midjourney_channel","r2_account_id","r2_access_key_id","r2_secret_access_key","r2_bucket","r2_public_base_url"];
          sec.forEach(function(k) {{
            var v = _getV("ed_" + k);
            if (v) o[k] = v; else if (b[k] !== undefined) o[k] = b[k];
          }});
          var nss = document.getElementById("ed_no_shared_settings");
          if (nss && nss.checked) o.no_shared_settings = true; else delete o.no_shared_settings;
          var nsp = document.getElementById("ed_no_shared_prompts");
          if (nsp && nsp.checked) o.no_shared_prompts = true; else delete o.no_shared_prompts;
          var stEl = document.getElementById("ed_settings_json");
          var st = (stEl && stEl.value) ? stEl.value.trim() : "";
          if (st) {{ try {{ o.settings = JSON.parse(st); }} catch (e) {{ throw new Error("settings JSON: " + e.message); }} }}
          else {{ if (b.settings) o.settings = JSON.parse(JSON.stringify(b.settings)); else delete o.settings; }}
          var bp = (b.prompts && typeof b.prompts === "object") ? JSON.parse(JSON.stringify(b.prompts)) : {{}};
          _readPromptsFromFormInto(bp, b);
          if (Object.keys(bp).length) o.prompts = bp; else delete o.prompts;
          if (!_getV("ed_id")) {{ throw new Error("id is required"); }}
          o.id = _getV("ed_id");
          return o;
        }}
        function showSiteConfigInfo(projectLabel) {{
          _siteEditorProject = projectLabel;
          _siteEditorBase = null;
          _siteEditorRawDirty = false;
          var fe = document.getElementById("siteEditorFetchErr");
          var ta = document.getElementById("siteEditorRaw");
          var saveBtn = document.getElementById("siteEditorSaveBtn");
          var formPane = document.getElementById("siteEditorFormPane");
          var formMissing = document.getElementById("siteEditorFormMissing");
          var titleEl = document.getElementById("siteConfigModalLabel");
          if (titleEl) titleEl.textContent = "Project — " + projectLabel;
          if (fe) {{ fe.classList.add("d-none"); fe.textContent = ""; }}
          if (ta) {{ ta.value = ""; ta.disabled = true; }}
          if (formPane) formPane.classList.add("d-none");
          if (formMissing) {{ formMissing.classList.add("d-none"); }}
          if (saveBtn) saveBtn.disabled = true;
          var mEl = document.getElementById("siteConfigModal");
          if (!mEl) return;
          if (typeof bootstrap === "undefined") {{
            if (fe) {{ fe.classList.remove("d-none"); fe.textContent = "Bootstrap is not loaded."; }}
            return;
          }}
          new bootstrap.Modal(mEl).show();
          var tabP = document.getElementById("tab-se-proj");
          if (tabP) {{ var tr = new bootstrap.Tab(tabP); tr.show(); }}
          fetch("/api/site-editor?project=" + encodeURIComponent(projectLabel))
            .then(function(r) {{ if (!r.ok) return r.text().then(function(t) {{ throw new Error(t || "HTTP " + r.status); }}); return r.json(); }})
            .then(function(data) {{
              if (fe) fe.classList.add("d-none");
              if (data.raw_site) {{
                _siteEditorBase = data.raw_site;
                _siteEditorRawDirty = false;
                siteEditorFillForm(data.raw_site, data.snapshot || null);
                if (ta) {{ ta.value = JSON.stringify(data.raw_site, null, 2); ta.disabled = false; }}
                if (formPane) formPane.classList.remove("d-none");
                if (formMissing) formMissing.classList.add("d-none");
                if (saveBtn) saveBtn.disabled = false;
              }} else {{
                if (ta) {{ ta.value = ""; }}
                if (formMissing) formMissing.classList.remove("d-none");
              }}
            }})
            .catch(function(e) {{ var fe2 = document.getElementById("siteEditorFetchErr"); if (fe2) {{ fe2.classList.remove("d-none"); fe2.textContent = "Error: " + (e && e.message ? e.message : e); }} }});
        }}
        function siteEditorSave() {{
          if (!_siteEditorProject) return;
          var obj = null;
          var ta0 = document.getElementById("siteEditorRaw");
          if (_siteEditorBase) {{
            if (_siteEditorRawDirty && ta0 && !ta0.disabled && (ta0.value || "").trim()) {{
              try {{
                obj = JSON.parse(ta0.value.trim());
                if (obj == null || typeof obj !== "object" || Array.isArray(obj)) throw new Error("Site must be a JSON object");
                siteEditorApplyKeepSecretsFromBase(_siteEditorBase, obj);
              }} catch (e) {{ alert(e.message || e); return; }}
            }} else {{
              try {{ obj = siteEditorReadMerged(); }} catch (e) {{ alert(e.message || e); return; }}
            }}
          }} else {{
            if (!ta0 || !ta0.value || ta0.disabled) {{ alert("Nothing to save"); return; }}
            try {{ obj = JSON.parse(ta0.value.trim()); }} catch (e) {{ alert("Invalid JSON: " + e); return; }}
          }}
          if (typeof obj !== "object" || obj === null || Array.isArray(obj)) {{ alert("Site must be a JSON object"); return; }}
          fetch("/api/site-editor", {{
            method: "POST",
            headers: {{ "Content-Type": "application/json" }},
            body: JSON.stringify({{ project: _siteEditorProject, raw_site: obj }})
          }})
            .then(function(r) {{ return r.json().then(function(j) {{ if (!r.ok) throw new Error(j.error || r.status); return j; }}); }})
            .then(function() {{ alert("Saved to config/sites.json. Refresh the page if the label changed."); }})
            .catch(function(e) {{ alert("Save failed: " + e); }});
        }}
        // -------------------- Themes + Cloudflare UI --------------------
        var _availableThemes = [];
        function loadThemesIntoAllPickers() {{
          return fetch("/api/themes").then(function(r) {{ return r.json(); }}).then(function(j) {{
            _availableThemes = (j && j.themes) || [];
            document.querySelectorAll("select.theme-picker").forEach(function(sel) {{
              var current = sel.getAttribute("data-current") || "";
              sel.innerHTML = "";
              var optBlank = document.createElement("option");
              optBlank.value = "";
              optBlank.textContent = "— none —";
              sel.appendChild(optBlank);
              _availableThemes.forEach(function(t) {{
                var o = document.createElement("option");
                o.value = t.slug;
                o.textContent = (t.display_name || t.slug) + (t.valid ? "" : " (incomplete)");
                if (!t.valid) o.disabled = true;
                sel.appendChild(o);
              }});
              if (current) sel.value = current;
            }});
          }}).catch(function(){{ /* silent */ }});
        }}
        function refreshProjectThemeState(project) {{
          if (!project) return Promise.resolve();
          var enc = encodeURIComponent(project);
          return fetch("/api/project-theme?project=" + enc).then(function(r) {{ return r.json(); }}).then(function(j) {{
            if (!j || !j.ok) return;
            var sel = document.querySelector('select.theme-picker[data-project="' + project.replace(/"/g, '\\\\"') + '"]');
            if (sel) {{
              sel.setAttribute("data-current", j.theme_slug || "");
              if (j.theme_slug) sel.value = j.theme_slug; else sel.value = "";
            }}
            var inp = document.querySelector('input.cf-project-input[data-project="' + project.replace(/"/g, '\\\\"') + '"]');
            if (inp) inp.value = j.cloudflare_project_name || "";
            var st = document.querySelector('.theme-status[data-project="' + project.replace(/"/g, '\\\\"') + '"]');
            if (st) {{
              if (j.cf_button_enabled) {{
                st.textContent = "✓ ready to deploy → " + (j.effective_cf_project_name || "");
                st.classList.remove("text-danger");
                st.classList.add("text-success");
              }} else {{
                st.textContent = (j.reasons && j.reasons[0]) || "Not deployable.";
                st.classList.add("text-danger");
                st.classList.remove("text-success");
              }}
            }}
            var cfBtn = document.querySelector('button.cf-upload-btn[data-project="' + project.replace(/"/g, '\\\\"') + '"]');
            if (cfBtn) {{
              cfBtn.disabled = !j.cf_button_enabled;
              cfBtn.title = j.cf_button_enabled
                ? "Build theme + deploy to Cloudflare Pages (creates new version)"
                : ((j.reasons || []).join(" · ") || "Not deployable");
            }}
            var verBtn = document.querySelector('button.cf-versions-btn[data-project="' + project.replace(/"/g, '\\\\"') + '"]');
            if (verBtn) {{
              verBtn.disabled = !(j.effective_cf_project_name && (j.has_global_cf_token || j.has_project_cf_override));
            }}
          }}).catch(function(){{ /* silent */ }});
        }}
        function saveProjectTheme(project) {{
          if (!project) return;
          var sel = document.querySelector('select.theme-picker[data-project="' + project.replace(/"/g, '\\\\"') + '"]');
          var inp = document.querySelector('input.cf-project-input[data-project="' + project.replace(/"/g, '\\\\"') + '"]');
          var payload = {{ project: project, theme_slug: (sel ? sel.value : ""), cloudflare_project_name: (inp ? inp.value.trim() : "") }};
          var btn = document.querySelector('.theme-save-btn[data-project="' + project.replace(/"/g, '\\\\"') + '"]');
          if (btn) {{ btn.disabled = true; btn.textContent = "Saving…"; }}
          fetch("/api/project-theme", {{
            method: "POST",
            headers: {{ "Content-Type": "application/json" }},
            body: JSON.stringify(payload)
          }}).then(function(r) {{ return r.json().then(function(j) {{ if (!r.ok) throw new Error(j.error || r.status); return j; }}); }})
            .then(function() {{ return refreshProjectThemeState(project); }})
            .catch(function(e) {{ alert("Save failed: " + e); }})
            .finally(function() {{ if (btn) {{ btn.disabled = false; btn.textContent = "Save"; }} }});
        }}
        function _bindThemeUI() {{
          document.querySelectorAll(".theme-save-btn").forEach(function(b) {{
            if (b.__themeBound) return;
            b.__themeBound = true;
            b.addEventListener("click", function() {{ saveProjectTheme(b.getAttribute("data-project") || ""); }});
          }});
          document.querySelectorAll("select.theme-picker").forEach(function(s) {{
            if (s.__themeBound) return;
            s.__themeBound = true;
            // No autosave; user clicks Save. But keep status synced if user already had it saved.
          }});
        }}
        function showCfVersions(project, logId) {{
          if (!project) return;
          var enc = encodeURIComponent(project);
          fetch("/api/cf-deployments?project=" + enc).then(function(r) {{ return r.json(); }}).then(function(j) {{
            var dlg = document.getElementById("cfVersionsModal");
            var body = document.getElementById("cfVersionsBody");
            var title = document.getElementById("cfVersionsTitle");
            if (title) title.textContent = "Cloudflare versions · " + project;
            if (!body) return;
            if (!j || !j.ok) {{ body.innerHTML = '<div class="alert alert-warning">' + ((j && j.error) || "Could not load deployments") + '</div>'; }}
            else if (!j.deployments || !j.deployments.length) {{ body.innerHTML = '<div class="text-muted">No deployments yet.</div>'; }}
            else {{
              var rows = j.deployments.map(function(d) {{
                var prod = d.is_production ? '<span class="badge bg-success">production</span>' : '';
                var url = d.url ? '<a href="' + d.url + '" target="_blank" rel="noopener">' + d.url + '</a>' : '';
                var disabled = d.is_production ? 'disabled' : '';
                return '<tr>'
                  + '<td><code>' + (d.short_id || d.id || '') + '</code> ' + prod + '</td>'
                  + '<td class="small text-muted">' + (d.created_on || '') + '</td>'
                  + '<td>' + url + '</td>'
                  + '<td><button class="btn btn-sm btn-outline-primary" ' + disabled
                  + ' onclick=\\'cfRollback("' + project.replace(/"/g, '\\\\"') + '","' + (d.id || '') + '")\\'>Make production</button></td>'
                  + '</tr>';
              }}).join("");
              body.innerHTML = '<table class="table table-sm"><thead><tr><th>ID</th><th>Date</th><th>URL</th><th></th></tr></thead><tbody>' + rows + '</tbody></table>';
            }}
            if (dlg && window.bootstrap && window.bootstrap.Modal) {{
              (new window.bootstrap.Modal(dlg)).show();
            }}
          }}).catch(function(e) {{ alert("Versions load failed: " + e); }});
        }}
        function cfRollback(project, deploymentId) {{
          if (!project || !deploymentId) return;
          if (!confirm("Promote deployment " + deploymentId.slice(0, 8) + " to production for " + project + "?")) return;
          fetch("/api/cf-rollback", {{
            method: "POST",
            headers: {{ "Content-Type": "application/json" }},
            body: JSON.stringify({{ project: project, deployment_id: deploymentId }})
          }}).then(function(r) {{ return r.json().then(function(j) {{ if (!r.ok) throw new Error(j.error || r.status); return j; }}); }})
            .then(function() {{ showCfVersions(project, ""); }})
            .catch(function(e) {{ alert("Rollback failed: " + e); }});
        }}
        window.showCfVersions = showCfVersions;
        window.cfRollback = cfRollback;

        document.addEventListener("DOMContentLoaded", function() {{
          refreshAllProjectStats();
          refreshActionButtonStates();
          requestBrowserNotifyPermission();
          if (_statsTickerId) clearInterval(_statsTickerId);
          _statsTickerId = setInterval(function() {{
            (projectUnits || []).forEach(function(u) {{
              refreshProjectStatsMetaOnly(u.log_id);
            }});
          }}, 1000);
          if (_statsAutoRefreshId) clearInterval(_statsAutoRefreshId);
          _statsAutoRefreshId = setInterval(function() {{
            refreshAllProjectStats();
            refreshActionButtonStates();
          }}, STATS_REFRESH_MS);
          var fp = document.getElementById("siteEditorFormPane");
          if (fp) fp.addEventListener("input", function() {{ _siteEditorRawDirty = false; }}, true);
          var taR = document.getElementById("siteEditorRaw");
          if (taR) taR.addEventListener("input", function() {{ _siteEditorRawDirty = true; }});

          // Themes / Cloudflare bootstrap
          _bindThemeUI();
          loadThemesIntoAllPickers().then(function() {{
            (projectUnits || []).forEach(function(u) {{ refreshProjectThemeState(u.label); }});
          }});
        }});
      </script>
    </head>
    <body>
      <div class="materio-sidebar">
        <div class="sidebar-title">
          <i class='bx bx-grid-alt'></i>
          <span>AUTOMATION</span>
        </div>
        <ul>
          <li><a href="/"><i class='bx bx-home-alt'></i> Dashboard</a></li>
          <li><a href="/manage_sites"><i class='bx bx-list-ul'></i> Projects (sites)</a></li>
          <li><a href="/manage_starts"><i class='bx bx-table'></i> Manage STARTS</a></li>
          <li><a href="/manage_recipes"><i class='bx bx-food-menu'></i> Manage Recipes</a></li>
          <li><a href="/manage_images"><i class='bx bx-image'></i> Manage Images</a></li>
          <li><a href="/manage_articles"><i class='bx bx-file'></i> Manage Articles</a></li>
        </ul>
      </div>

      <div class="navbar-materio">
        <h5>Dashboard</h5>
        <div>
          <i class='bx bx-bell' style="font-size: 20px; margin-right: 20px;'></i>
          <i class='bx bx-user-circle' style="font-size: 24px;"></i>
        </div>
      </div>

      <div class="materio-content">
        <div class="card mb-4">
          <div class="card-header">
            <strong>Actions</strong>
          </div>
          <div class="card-body">
            <button class="number-button" data-step="START" data-step-key="START" data-number="1" onclick="startAllStart(this)">START</button>
            <button class="number-button" data-step="JSON" data-step-key="JSON" data-number="2" onclick="startLog('/stream-all-json', this)">JSON</button>
            <button class="number-button" data-step="PROMPT" data-step-key="PROMPT" data-number="2" onclick="startLog('/stream-all-prompt', this)">PROMPT</button>
            <button class="number-button" data-step="IMAGINE" data-step-key="IMAGINE" data-number="ALL" onclick="startLog('/stream-imagine-all', this)">IMAGINE ALL</button>
            <button class="number-button" data-step="IMAGINE" data-step-key="CLEAR IMAGINE" onclick="clearFailed()">CLEAR IMAGINE</button>

            <button class="number-button" data-step="ARTICLE" data-step-key="ARTICLE" data-number="9" onclick="startLog('/stream-all-article', this)">ARTICLE</button>
            <button class="number-button" data-step="PIN DATA" data-step-key="PIN DATA" data-number="10" onclick="startLog('/stream-all-pin-data', this)">PIN DATA</button>
            <button class="number-button" data-step="PIN IMAGE" data-step-key="PIN IMAGE" data-number="11" onclick="startLog('/stream-all-pin-image', this)">PIN IMAGE</button>
            <button class="number-button" data-step="PIN IMAGE" data-step-key="PIN IMAGE HTML" data-number="11b" onclick="startLog('/stream-all-pin-image-html', this)">PIN IMAGE HTML</button>
            <button class="number-button" data-step="WP UPLOAD" data-step-key="WP UPLOAD" data-number="12" onclick="startLog('/stream-all-wp-upload', this)">WP UPLOAD</button>
            <button class="number-button" data-step-key="CF UPLOAD" data-number="12b" onclick="startLog('/stream-all-cf-upload', this)" style="background:#f48120;color:#fff;border:1px solid #d96e10;" title="Deploy ALL configured projects to Cloudflare Pages (skips projects without a theme or CF project name)">CF UPLOAD</button>
            <button class="number-button" data-step="PIN DATA" data-step-key="PIN BULK" data-number="13" onclick="startLog('/stream-all-pin-bulk', this)">PIN BULK</button>
            <button class="number-button" data-step-key="AUTO SAFE" data-number="AUTO" onclick="startLog('/stream-all-auto-safe', this)">AUTO SAFE (ALL STEPS)</button>
            <button class="number-button" type="button" data-step-key="CLEAR ALL LOGS" onclick="clearAllLogs()">CLEAR ALL LOGS</button>

            <form id="deleteAllForm" action="/delete-all-folder" method="post" style="display:inline-block;">
              <button type="button" class="btn btn-danger ms-2" id="deleteAllBtn" data-step-key="DELETE ALL" onclick="confirmDeleteAll()">Delete 'ALL'</button>
            </form>
            <button type="button" class="btn btn-outline-warning ms-2" id="cleanupLocalBtn" data-step-key="CLEANUP LOCAL" onclick="confirmCleanupLocal('')">Cleanup local</button>
            <button type="button" class="btn btn-outline-danger ms-2" id="cleanupR2Btn" data-step-key="CLEANUP R2" onclick="confirmCleanupR2('')">Cleanup R2 unused</button>
          </div>
        </div>

        <div class="row">
          {log_boxes}
        </div>

        <div class="modal fade" id="siteConfigModal" tabindex="-1" aria-hidden="true">
          <div class="modal-dialog modal-xl modal-dialog-scrollable">
            <div class="modal-content">
              <div class="modal-header">
                <h5 class="modal-title" id="siteConfigModalLabel">Project</h5>
                <button type="button" class="btn-close" data-bs-dismiss="modal" aria-label="Close"></button>
              </div>
              <div class="modal-body p-0">
                <div class="px-3 pt-2 pb-1 border-bottom bg-light">
                  <div class="alert alert-info py-2 small mb-2">
                    Edits the <strong>one site object</strong> in <code>config/sites.json</code>. Use the tabs: <strong>Project</strong> (ids, paths) → <strong>WordPress</strong> → <strong>API &amp; R2</strong> (IMAGINE / UseAPI) → <strong>Settings</strong> → <strong>Start (a1)</strong> → <strong>JSON (a2)</strong> (JSON + text prompts) → <strong>Pipeline (a4–a8)</strong> → <strong>Row JSON</strong>. <strong>Save</strong> stores the row. Empty fields and gray placeholders use merged values from shared + project + files; see “Layers” and hints under each prompt block.
                  </div>
                  <p id="siteEditorFetchErr" class="text-danger small mb-2 d-none" role="alert"></p>
                  <p id="siteEditorFormMissing" class="text-warning small d-none mb-0">No <code>config/sites.json</code> row for this card. Add sites in <a href="/manage_sites" target="_blank">Projects (sites)</a>.</p>
                </div>
                <div id="siteEditorFormPane" class="d-none">
                  <div class="overflow-x-auto border-bottom" style="white-space: nowrap;">
                    <ul class="nav nav-tabs site-editor-top-tabs border-0 flex-nowrap px-2 pt-1 mb-0" id="siteEditorTopTabs" role="tablist" style="min-width: min-content; flex-wrap: nowrap;">
                      <li class="nav-item" role="presentation">
                        <button class="nav-link active" id="tab-se-proj" data-bs-toggle="tab" data-bs-target="#pane-se-proj" type="button" role="tab" aria-selected="true">Project</button>
                      </li>
                      <li class="nav-item" role="presentation">
                        <button class="nav-link" id="tab-se-wp" data-bs-toggle="tab" data-bs-target="#pane-se-wp" type="button" role="tab">WordPress</button>
                      </li>
                      <li class="nav-item" role="presentation">
                        <button class="nav-link" id="tab-se-api" data-bs-toggle="tab" data-bs-target="#pane-se-api" type="button" role="tab">API &amp; R2</button>
                      </li>
                      <li class="nav-item" role="presentation">
                        <button class="nav-link" id="tab-se-st" data-bs-toggle="tab" data-bs-target="#pane-se-st" type="button" role="tab">Settings</button>
                      </li>
                      <li class="nav-item" role="presentation">
                        <button class="nav-link" id="tab-se-p1" data-bs-toggle="tab" data-bs-target="#pane-se-p1" type="button" role="tab">a1 Start</button>
                      </li>
                      <li class="nav-item" role="presentation">
                        <button class="nav-link" id="tab-se-p2" data-bs-toggle="tab" data-bs-target="#pane-se-p2" type="button" role="tab">a2 JSON + prompt</button>
                      </li>
                      <li class="nav-item" role="presentation">
                        <button class="nav-link" id="tab-se-pp" data-bs-toggle="tab" data-bs-target="#pane-se-pp" type="button" role="tab">Pipeline a4–a8</button>
                      </li>
                      <li class="nav-item" role="presentation">
                        <button class="nav-link" id="tab-se-adv" data-bs-toggle="tab" data-bs-target="#pane-se-adv" type="button" role="tab">Row JSON</button>
                      </li>
                    </ul>
                  </div>
                  <div class="tab-content site-editor-form-tab-content px-3 py-2" id="siteEditorFormTabContent" style="max-height: min(62vh, 680px); overflow-y: auto;">
                    <div class="tab-pane fade show active" id="pane-se-proj" role="tabpanel" aria-labelledby="tab-se-proj">
                      <p class="text-muted small">Site <code>id</code> and output paths. Matches <code>config/sites.json</code> for this dashboard card.</p>
                      <div class="row g-2">
                        <div class="col-md-2 col-6"><label class="form-label">id</label><input id="ed_id" class="form-control form-control-sm" required /></div>
                        <div class="col-md-3 col-6"><label class="form-label">display_name</label><input id="ed_display_name" class="form-control form-control-sm" /></div>
                        <div class="col-md-3 col-6"><label class="form-label">out_dir</label><input id="ed_out_dir" class="form-control form-control-sm" /></div>
                        <div class="col-md-2 col-6"><label class="form-label">start_file</label><input id="ed_start_file" class="form-control form-control-sm" placeholder="xlsx" /></div>
                        <div class="col-md-2 col-6"><label class="form-label">log_id</label><input id="ed_log_id" class="form-control form-control-sm" /></div>
                        <div class="col-md-3 col-6"><label class="form-label">prompts_dir</label><input id="ed_prompts_dir" class="form-control form-control-sm" placeholder="config/site_prompts/…" /></div>
                        <div class="col-md-3 col-6"><label class="form-label">templates_dir</label><input id="ed_templates_dir" class="form-control form-control-sm" placeholder="default: templates" /></div>
                      </div>
                    </div>
                    <div class="tab-pane fade" id="pane-se-wp" role="tabpanel" aria-labelledby="tab-se-wp">
                      <h6 class="text-secondary small text-uppercase">WordPress</h6>
                      <p class="text-muted small mb-2">Not in <code>shared_keys</code> — this row, or <code>WP_URL</code> / <code>WP_USER</code> / <code>WP_APP_PASSWORD</code> in env. Blank app password = keep the value already in <code>sites.json</code>.</p>
                      <div class="row g-2">
                        <div class="col-md-4">
                          <label class="form-label small mb-0">Site URL <span class="text-muted">(wordpress_url)</span></label>
                          <input id="ed_wordpress_url" class="form-control form-control-sm" autocomplete="off" />
                          <p class="form-text mb-0" id="ed_src_wordpress_url" style="font-size:0.7rem"></p>
                        </div>
                        <div class="col-md-3">
                          <label class="form-label small mb-0">User</label>
                          <input id="ed_wordpress_user" class="form-control form-control-sm" autocomplete="off" />
                          <p class="form-text mb-0" id="ed_src_wordpress_user" style="font-size:0.7rem"></p>
                        </div>
                        <div class="col-md-5">
                          <label class="form-label small mb-0">App password</label>
                          <input id="ed_wordpress_app_password" class="form-control form-control-sm" type="text" autocomplete="off" />
                          <p class="form-text mb-0" id="ed_src_wordpress_app_password" style="font-size:0.7rem"></p>
                        </div>
                      </div>
                    </div>
                    <div class="tab-pane fade" id="pane-se-api" role="tabpanel" aria-labelledby="tab-se-api">
                      <h6 class="text-secondary small text-uppercase">OpenAI · UseAPI (IMAGINE) · R2</h6>
                      <p class="text-muted small mb-2">IMAGINE / Midjourney actions use the <strong>MJ channel</strong> and <strong>UseAPI</strong> fields here. Gray placeholders: merged from <code>shared_keys.json</code> + project <code>keys.json</code> + this row. “Source” lines show where each key came from.</p>
                      <div class="row g-2 mb-1">
                        <div class="col-md-4">
                          <label class="form-label small mb-0">OpenAI API key</label>
                          <input id="ed_openai_api_key" class="form-control form-control-sm" type="text" spellcheck="false" autocomplete="off" />
                          <p class="form-text mb-0" id="ed_src_openai_api_key" style="font-size:0.7rem"></p>
                        </div>
                        <div class="col-md-2">
                          <label class="form-label small mb-0">Model</label>
                          <input id="ed_openai_model" class="form-control form-control-sm" spellcheck="false" autocomplete="off" />
                          <p class="form-text mb-0" id="ed_src_openai_model" style="font-size:0.7rem"></p>
                        </div>
                        <div class="col-md-3">
                          <label class="form-label small mb-0">UseAPI token</label>
                          <input id="ed_useapi_token" class="form-control form-control-sm" type="text" spellcheck="false" autocomplete="off" />
                          <p class="form-text mb-0" id="ed_src_useapi_token" style="font-size:0.7rem"></p>
                        </div>
                        <div class="col-md-3">
                          <label class="form-label small mb-0">Midjourney channel</label>
                          <input id="ed_useapi_midjourney_channel" class="form-control form-control-sm" spellcheck="false" autocomplete="off" />
                          <p class="form-text mb-0" id="ed_src_useapi_midjourney_channel" style="font-size:0.7rem"></p>
                        </div>
                      </div>
                      <div class="row g-2">
                        <div class="col-md-2 col-6">
                          <label class="form-label small mb-0">R2 account</label>
                          <input id="ed_r2_account_id" class="form-control form-control-sm" spellcheck="false" autocomplete="off" />
                          <p class="form-text mb-0" id="ed_src_r2_account_id" style="font-size:0.7rem"></p>
                        </div>
                        <div class="col-md-2 col-6">
                          <label class="form-label small mb-0">R2 access key</label>
                          <input id="ed_r2_access_key_id" class="form-control form-control-sm" type="text" spellcheck="false" autocomplete="off" />
                          <p class="form-text mb-0" id="ed_src_r2_access_key_id" style="font-size:0.7rem"></p>
                        </div>
                        <div class="col-md-2 col-6">
                          <label class="form-label small mb-0">R2 secret</label>
                          <input id="ed_r2_secret_access_key" class="form-control form-control-sm" type="text" spellcheck="false" autocomplete="off" />
                          <p class="form-text mb-0" id="ed_src_r2_secret_access_key" style="font-size:0.7rem"></p>
                        </div>
                        <div class="col-md-2 col-6">
                          <label class="form-label small mb-0">R2 bucket</label>
                          <input id="ed_r2_bucket" class="form-control form-control-sm" spellcheck="false" autocomplete="off" />
                          <p class="form-text mb-0" id="ed_src_r2_bucket" style="font-size:0.7rem"></p>
                        </div>
                        <div class="col-md-4">
                          <label class="form-label small mb-0">R2 public URL</label>
                          <input id="ed_r2_public_base_url" class="form-control form-control-sm" spellcheck="false" autocomplete="off" />
                          <p class="form-text mb-0" id="ed_src_r2_public_base_url" style="font-size:0.7rem"></p>
                        </div>
                      </div>
                    </div>
                    <div class="tab-pane fade" id="pane-se-st" role="tabpanel" aria-labelledby="tab-se-st">
                      <h6 class="text-secondary small text-uppercase">Settings merge + row object</h6>
                      <p class="text-muted small mb-2">Toggles and optional per-row <code>settings</code> JSON. Empty = use merged project settings; placeholder shows the merged object.</p>
                      <div class="border rounded p-2 mb-3 small">
                        <div class="form-check mb-1">
                          <input class="form-check-input" type="checkbox" id="ed_no_shared_settings" name="x" value="1" />
                          <label class="form-check-label" for="ed_no_shared_settings"><code>no_shared_settings</code> — do not load <code>config/shared_settings.json</code> (use project <code>settings.json</code> + this row only).</label>
                        </div>
                        <div class="form-check mb-0">
                          <input class="form-check-input" type="checkbox" id="ed_no_shared_prompts" name="x" value="1" />
                          <label class="form-check-label" for="ed_no_shared_prompts"><code>no_shared_prompts</code> — do not load repo <code>config/prompts/*.json</code> (project <code>config/prompts</code> + <code>site_prompts/…</code> + row inline <code>prompts</code> still apply).</label>
                        </div>
                      </div>
                      <label class="form-label small">Override <code>settings</code> (JSON) on this row</label>
                      <textarea id="ed_settings_json" class="form-control form-control-sm font-monospace" rows="8" spellcheck="false" placeholder="{{}}"></textarea>
                    </div>
                    <div class="tab-pane fade" id="pane-se-p1" role="tabpanel" aria-labelledby="tab-se-p1">
                      <h6 class="text-secondary small text-uppercase">START — a1_start</h6>
                      <p class="text-muted small mb-2">Fields from <code>config/prompts/a1_start.json</code> shape. Merged on save with file layers; “Layers” under each name shows the merge order.</p>
                      <div id="ed_pr_sub_start"></div>
                    </div>
                    <div class="tab-pane fade" id="pane-se-p2" role="tabpanel" aria-labelledby="tab-se-p2">
                      <h6 class="text-secondary small text-uppercase">a2 — JSON + prompt (PROMPT / IMAGINE text)</h6>
                      <p class="text-muted small mb-2"><code>a2_json</code> and <code>a2_prompt</code>: same names as the dashboard JSON / PROMPT steps. IMAGINE buttons use the UseAPI + Midjourney settings in <strong>API &amp; R2</strong>.</p>
                      <div id="ed_pr_sub_a2"></div>
                    </div>
                    <div class="tab-pane fade" id="pane-se-pp" role="tabpanel" aria-labelledby="tab-se-pp">
                      <h6 class="text-secondary small text-uppercase">a4, a5, a6b, a8 — articles, pin data, pin bulk</h6>
                      <p class="text-muted small mb-2">Pipeline prompts (ARTICLE, PIN DATA, PIN IMAGES HTML, PIN BULK). Under <strong>a5_pin_data</strong>, the <strong>Recipe → WordPress category</strong> block classifies titles for WordPress. All prompt text lives in <code>config/prompts/*.json</code> — editable here per project. <strong>app_title_filter</strong> (under Other) filters bulk title uploads on the dashboard.</p>
                      <div id="ed_pr_sub_pipe" class="mb-2"></div>
                      <p id="ed_pr_sub_other_lbl" class="text-muted small d-none">Other script prompt(s) in the repo</p>
                      <div id="ed_pr_sub_other" class="mb-1"></div>
                    </div>
                    <div class="tab-pane fade" id="pane-se-adv" role="tabpanel" aria-labelledby="tab-se-adv">
                      <h6 class="text-secondary small text-uppercase">Full row JSON</h6>
                      <p class="text-muted small mb-2">The exact object saved in <code>config/sites.json</code>. If you edit here, the raw text wins on <strong>Save</strong> (unless the form is edited afterward — then the form merge is used). Leave secrets blank in the form to keep existing file values when saving from the form.</p>
                      <textarea id="siteEditorRaw" class="form-control font-monospace" style="min-height: 200px; font-size: 11px;" spellcheck="false" disabled></textarea>
                    </div>
                  </div>
                </div>
              </div>
              <div class="modal-footer">
                <button type="button" class="btn btn-primary" id="siteEditorSaveBtn" onclick="siteEditorSave()" disabled>Save to <code>config/sites.json</code></button>
                <a href="/manage_sites" class="btn btn-outline-secondary" target="_blank" rel="noopener">Full projects form</a>
                <button type="button" class="btn btn-secondary" data-bs-dismiss="modal">Close</button>
              </div>
            </div>
          </div>
        </div>

        <div class="modal fade" id="columnDetailsModal" tabindex="-1" aria-hidden="true">
          <div class="modal-dialog modal-xl modal-dialog-scrollable">
            <div class="modal-content">
              <div class="modal-header">
                <h5 class="modal-title" id="columnDetailsTitle">Column Details</h5>
                <button type="button" class="btn-close" data-bs-dismiss="modal" aria-label="Close"></button>
              </div>
              <div class="modal-body" id="columnDetailsBody">
                <div class="text-muted">Loading...</div>
              </div>
              <div class="modal-footer">
                <button type="button" class="btn btn-secondary" data-bs-dismiss="modal">Close</button>
              </div>
            </div>
          </div>
        </div>

        <div class="modal fade" id="cfVersionsModal" tabindex="-1" aria-hidden="true">
          <div class="modal-dialog modal-lg modal-dialog-scrollable">
            <div class="modal-content">
              <div class="modal-header">
                <h5 class="modal-title" id="cfVersionsTitle">Cloudflare versions</h5>
                <button type="button" class="btn-close" data-bs-dismiss="modal" aria-label="Close"></button>
              </div>
              <div class="modal-body" id="cfVersionsBody">
                <div class="text-muted">Loading...</div>
              </div>
              <div class="modal-footer">
                <button type="button" class="btn btn-secondary" data-bs-dismiss="modal">Close</button>
              </div>
            </div>
          </div>
        </div>

      </div>
      <script src="https://cdn.jsdelivr.net/npm/bootstrap@5.0.2/dist/js/bootstrap.bundle.min.js"></script>
      <script>
        if (typeof applyStepButtonTooltips === "function") applyStepButtonTooltips();
      </script>
    </body>
    </html>
    """
# ---- Patch: Ensure IMAGINE ALL button turns green after batches complete ----
try:
    _orig_generate_log_in_batches = generate_log_in_batches
    def generate_log_in_batches(script_names, batch_size=3, env_extra=None):
        for chunk in _orig_generate_log_in_batches(script_names, batch_size=batch_size, env_extra=env_extra):
            yield chunk
        # السطر اللي كيسالي وكيخلي الزرّ يخضر
        yield "data: " + json.dumps({"folder": "all", "line": "Finished all processes."}) + "\n\n"
except NameError:
    pass


if __name__ == "__main__":
    app.run(debug=True, threaded=True)
